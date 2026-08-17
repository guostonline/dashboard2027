import os
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
import json
from flask import Flask, jsonify, request, render_template, redirect, send_from_directory
from data_processor import ExcelProcessor, get_categorie
import pandas as pd
import datetime
from generate_report import generate_report
import db_manager
try:
    db_manager.init_db()
    if os.path.exists("acm.xlsx"):
        db_manager.import_acm_file("acm.xlsx")
except Exception as e:
    print(f"Notice on init: {e}")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(
    __name__,
    static_folder=os.path.join(BASE_DIR, 'static'),
    static_url_path='/static',
    template_folder=os.path.join(BASE_DIR, 'templates')
)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(os.path.join(BASE_DIR, 'static'), filename)

@app.route('/logo.png')
def serve_logo_root():
    if os.path.exists(os.path.join(BASE_DIR, 'static', 'logo.png')):
        return send_from_directory(os.path.join(BASE_DIR, 'static'), 'logo.png')
    return send_from_directory(BASE_DIR, 'logo.png')

@app.route('/bg_anim.png')
def serve_bg_anim_root():
    return send_from_directory(BASE_DIR, 'bg_anim.png')

@app.route('/favicon.ico')
def serve_favicon():
    return serve_logo_root()

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# Ensure directories exist
try:
    os.makedirs("templates", exist_ok=True)
    os.makedirs("static/css", exist_ok=True)
    os.makedirs("static/js", exist_ok=True)
except Exception:
    pass

if getattr(db_manager, 'IS_SERVERLESS', False):
    CONFIG_PATH = os.path.join(getattr(db_manager, 'DB_DIR', '/tmp'), "config.json")
    _src_cfg = os.path.join(BASE_DIR, "config.json")
    if os.path.exists(_src_cfg) and not os.path.exists(CONFIG_PATH):
        try:
            shutil.copyfile(_src_cfg, CONFIG_PATH)
            try:
                os.chmod(CONFIG_PATH, 0o666)
            except Exception:
                pass
        except Exception:
            pass
else:
    CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

EXCEL_DIR = os.path.join(getattr(db_manager, 'DB_DIR', BASE_DIR), "excel")
try:
    os.makedirs(EXCEL_DIR, exist_ok=True)
except Exception:
    pass

def load_config():
    defaults = {
        "rest_days": 15,
        "exclude_families": [],
        "theme": "theme-1",
        "light_mode": True,
        "excluded_dates": [],
        "google_sheet_url": ""
    }
    if not os.path.exists(CONFIG_PATH):
        save_config(defaults)
        return defaults
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            # Ensure all default keys exist
            for k, v in defaults.items():
                if k not in data:
                    data[k] = v
            return data
    except Exception as e:
        print(f"Error loading config: {e}")
        return defaults

def save_config(config):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4)
        return True
    except Exception as e:
        print(f"Error saving config: {e}")
        return False

def get_processor():
    config = load_config()
    rest_days = int(config.get("rest_days", 15))
    exclude_families = config.get("exclude_families", [])
    return ExcelProcessor(rest_days=rest_days, exclude_families=exclude_families)


def process_and_save_suivi(date, file_content, force_extract_rest_days=False):
    temp_in = os.path.join(EXCEL_DIR, f"temp_upload_{date}.xlsx")
    temp_out = os.path.join(EXCEL_DIR, f"temp_finale_{date}.xlsx")
    
    # Save file content temporarily to run openpyxl/pandas
    try:
        os.makedirs(EXCEL_DIR, exist_ok=True)
    except Exception:
        pass
    with open(temp_in, "wb") as f:
        f.write(file_content)
        
    # Get settings for this date, default if not present
    settings = db_manager.get_suivi_settings(date)
    if settings and not force_extract_rest_days:
        rest_days = settings["rest_days"]
        exclude_families = settings["exclude_families"]
    else:
        rest_days = None
        exclude_families = settings["exclude_families"] if settings else []
        
    p = ExcelProcessor(path=temp_in, output_path=temp_out, rest_days=rest_days, exclude_families=exclude_families, date=date)
    
    try:
        # Extract default rest days from file if not specified
        elapsed, total = p.get_day_work()
        if rest_days is None:
            rest_days = total - elapsed
            db_manager.save_suivi_settings(date, rest_days, exclude_families)
            # Re-initialize processor with determined rest days
            p = ExcelProcessor(path=temp_in, output_path=temp_out, rest_days=rest_days, exclude_families=exclude_families, date=date)
            
        p.fix_sheet(jour_rest=rest_days)
        data = p.get_data()
        
        # Save parsed data to DB
        db_manager.save_suivi_data(date, data)
        return True
    finally:
        # Clean up temp files
        if os.path.exists(temp_in):
            try:
                os.remove(temp_in)
            except Exception:
                pass
        if os.path.exists(temp_out):
            try:
                os.remove(temp_out)
            except Exception:
                pass

def get_all_vendeurs_from_db():
    try:
        fdv_rows = db_manager.get_fdv_list()
        # Filter only vendors belonging to CHAKIB ELFIL's CDZ
        vendeurs = sorted(list(set([
            r["vendeur"].strip()
            for r in fdv_rows
            if r.get("vendeur") and r["vendeur"].strip() != "N/A"
            and (r.get("cdz", "") or "").strip().upper() == "CHAKIB ELFIL"
        ])))
        return vendeurs
    except Exception as e:
        import traceback
        print("Error fetching vendeurs from DB:", e)
        traceback.print_exc()
        return []

@app.route("/")
@app.route("/dashboard")
def index():
    # If Google callback
    code = request.args.get('code')
    if code:
        try:
            import requests
            with open('google.json', 'r') as f:
                secrets = json.load(f)['web']
            
            token_url = secrets['token_uri']
            payload = {
                'code': code,
                'client_id': secrets['client_id'],
                'client_secret': secrets['client_secret'],
                'redirect_uri': 'http://127.0.0.1:5000/',
                'grant_type': 'authorization_code'
            }
            res = requests.post(token_url, data=payload)
            tokens = res.json()
            
            if 'error' in tokens:
                return f"Erreur Google OAuth : {tokens.get('error_description', tokens['error'])}", 400
                
            token_data = {
                'token': tokens.get('access_token'),
                'refresh_token': tokens.get('refresh_token'),
                'token_uri': secrets['token_uri'],
                'client_id': secrets['client_id'],
                'client_secret': secrets['client_secret'],
                'scopes': ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/gmail.modify']
            }
            
            with open('token.json', 'w') as token_file:
                json.dump(token_data, token_file)
            
            # Successful auth redirect back to stock view
            return """
            <html>
                <body>
                    <script>
                        alert("Authentification Google Sheets réussie ! Le stock va être synchronisé.");
                        window.location.href = "/stock";
                    </script>
                </body>
            </html>
            """
        except Exception as e:
            return f"Erreur lors de la configuration Google Sheets : {str(e)}", 500

    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    all_vendeurs = get_all_vendeurs_from_db()
    return render_template("index.html", theme=theme, light_mode=light_mode, all_vendeurs=all_vendeurs)

@app.route("/login")
def login_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("login.html", theme=theme, light_mode=light_mode)

@app.route("/details")
@app.route("/vendeur360")
def details():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    active_sub_tab = request.args.get("view", "")
    if request.path == "/vendeur360":
        active_sub_tab = "vendeur360"
    all_vendeurs = get_all_vendeurs_from_db()
    print("INSIDE DETAILS ROUTE - VENDEURS COUNT:", len(all_vendeurs), all_vendeurs[:3] if all_vendeurs else 'EMPTY')
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="details", active_sub_tab=active_sub_tab, all_vendeurs=all_vendeurs)


@app.route("/api/vendeur360/tournees/<path:vendeur_name>")
def api_vendeur360_tournees(vendeur_name):
    try:
        from db_manager import get_vendeur_tournees_summary
        data = get_vendeur_tournees_summary(vendeur_name)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/import-all-secteurs", methods=["POST", "GET"])
def api_import_all_secteurs():
    try:
        from db_manager import import_all_secteurs_tournees
        count = import_all_secteurs_tournees()
        return jsonify({"success": True, "inserted_count": count})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/visites/import-all-secteurs", methods=["POST", "GET"])
def api_import_all_secteurs_visites():
    try:
        res = db_manager.import_all_secteurs_visites_rapports(save_tournee=False)
        return jsonify(res)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/clients")
def clients():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    active_sub_tab = request.args.get("view", "")
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="clients", active_sub_tab=active_sub_tab)

@app.route("/rapport")
def rapport():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="rapport")

@app.route("/stock")
def stock():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    active_sub_tab = request.args.get("view", "")
    if active_sub_tab == 'favorit':
        active_sub_tab = 'favorites'
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="stock", active_sub_tab=active_sub_tab)

@app.route("/anomalis")
@app.route("/anomalies")
def anomalis_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="anomalies")

@app.route("/visites")
def visites_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="visites")

@app.route("/api/anomalies/analysis", methods=["GET"])
def api_anomalies_analysis():
    """Analyze visits from visites_rapports against 4 strict anomaly rules:
    1. Durée < 3 min (Red)
    2. Multiple visits to same client on same date (Orange)
    3. First visit > 08:40:00 (Dark Red)
    4. Last visit < 14:45:00 (Dark Red)
    """
    try:
        from collections import defaultdict
        from datetime import datetime

        vendeur_filter = request.args.get("vendeur", "All").strip()
        date_filter = request.args.get("date", "All").strip()

        conn = db_manager.get_db_connection()
        cursor = conn.cursor()

        # Fast DISTINCT queries for dropdown filters
        cursor.execute("SELECT DISTINCT vendeur FROM visites_rapports WHERE vendeur IS NOT NULL AND vendeur != '' ORDER BY vendeur ASC")
        all_vendeurs = [r[0] for r in cursor.fetchall() if r[0]]

        cursor.execute("SELECT DISTINCT date_visite FROM visites_rapports WHERE date_visite IS NOT NULL AND date_visite != '' ORDER BY date_visite DESC")
        all_dates = [r[0] for r in cursor.fetchall() if r[0]]

        # Build SQL filter clauses to avoid scanning entire DB in Python
        where_clauses = ["date_visite IS NOT NULL AND date_visite != ''"]
        params = []

        if vendeur_filter and vendeur_filter != "All":
            vcode = vendeur_filter.split(" ")[0].strip().upper()
            clean_name = vendeur_filter.strip().upper()
            where_clauses.append("(UPPER(TRIM(vendeur)) = ? OR UPPER(vendeur) LIKE ? OR UPPER(vendeur) LIKE ?)")
            params.extend([clean_name, f"{vcode} %", f"{vcode}%"])

        if date_filter and date_filter != "All":
            where_clauses.append("(date_visite = ? OR date_visite LIKE ?)")
            params.extend([date_filter, f"%{date_filter}%"])

        where_str = " WHERE " + " AND ".join(where_clauses)

        cursor.execute(f"""
            SELECT id, file_name, vendeur, date_visite, tournee, agence, client_code, client_nom, heure, distance, motif, note
            FROM visites_rapports
            {where_str}
            ORDER BY date_visite DESC, vendeur ASC, heure ASC, id ASC
        """, params)
        rows = cursor.fetchall()

        # Fallback to vendeur_tournees_visits if visites_rapports is empty
        if not rows and not params:
            cursor.execute("""
                SELECT id, '' as file_name, vendeur_name as vendeur, date as date_visite, tournee, '' as agence,
                       client_code, client_name as client_nom,
                       (heure_debut || ' - ' || heure_fin) as heure, distance, motif, note
                FROM vendeur_tournees_visits
                WHERE date IS NOT NULL
                ORDER BY date DESC, vendeur_name ASC, heure_debut ASC, id ASC
            """)
            rows = cursor.fetchall()

        conn.close()

        groups = defaultdict(list)
        for r in rows:
            v = (r['vendeur'] or '').strip()
            d = (r['date_visite'] or '').strip()
            groups[(v, d)].append(dict(r))

        analyzed_visites = []
        stats = {
            "total_visites": 0,
            "count_less_3min": 0,
            "count_multiple": 0,
            "count_first_late": 0,
            "count_last_early": 0,
            "total_anomalies": 0
        }

        for (vendeur, date_visite), visits in groups.items():
            client_counts = defaultdict(int)
            for visit in visits:
                c_code = (visit.get('client_code') or '').strip().upper()
                if c_code:
                    client_counts[c_code] += 1

            total_v = len(visits)

            for idx, visit in enumerate(visits):
                c_code = (visit.get('client_code') or '').strip()
                c_nom = (visit.get('client_nom') or '').strip()
                heure_raw = (visit.get('heure') or '').strip()

                parts = [p.strip() for p in heure_raw.split('-')]
                h_start = parts[0] if len(parts) > 0 else ''
                h_end = parts[1] if len(parts) > 1 else h_start

                dur_sec = 0
                if h_start and h_end:
                    try:
                        t1 = datetime.strptime(h_start, '%H:%M:%S')
                        t2 = datetime.strptime(h_end, '%H:%M:%S')
                        dur_sec = max(0, int((t2 - t1).total_seconds()))
                    except Exception:
                        pass

                mins = dur_sec // 60
                secs = dur_sec % 60
                dur_str = f"{mins:02d}:{secs:02d}"

                motif_str = (visit.get('motif') or '').strip().upper()
                is_closed_motif = any(m in motif_str for m in ['FERM', 'ABSENT', 'NON VISITE', 'NON VISITÉ'])

                is_less_3min = (dur_sec < 180) if not is_closed_motif else False

                total_client_visites = client_counts[c_code.upper()] if c_code else 1
                is_closed_not_revisited = (is_closed_motif and total_client_visites < 2)
                is_multiple = (total_client_visites >= 2) if c_code else False

                is_first = (idx == 0)
                is_first_late = False
                if is_first and h_start and h_start > '08:40:00':
                    is_first_late = True

                is_last = (idx == total_v - 1)
                is_last_early = False
                if is_last and h_end and h_end < '14:45:00':
                    is_last_early = True

                anom_list = []
                if is_less_3min: anom_list.append('Durée < 3min')
                if is_closed_not_revisited: 
                    anom_list.append('1 Visite')
                if is_first_late: anom_list.append('1ère Visite > 08:40')
                if is_last_early: anom_list.append('Dernière Visite < 14:45')

                has_anom = len(anom_list) > 0

                item = {
                    "id": visit['id'],
                    "client_code": c_code,
                    "client_nom": c_nom,
                    "date_visite": date_visite,
                    "heure_debut": h_start,
                    "heure_fin": h_end,
                    "duree_seconds": dur_sec,
                    "duree_formatted": dur_str,
                    "distance": visit.get('distance') or 0,
                    "motif": visit.get('motif') or '',
                    "note": visit.get('note') or '',
                    "vendeur": vendeur,
                    "tournee": visit.get('tournee') or '',
                    "is_less_3min": is_less_3min,
                    "is_closed_not_revisited": is_closed_not_revisited,
                    "is_multiple": is_multiple,
                    "is_first_visit": is_first,
                    "is_last_visit": is_last,
                    "is_first_late": is_first_late,
                    "is_last_early": is_last_early,
                    "has_anomaly": has_anom,
                    "anomalies": anom_list
                }

                analyzed_visites.append(item)
                stats['total_visites'] += 1
                if is_less_3min: stats['count_less_3min'] += 1
                if is_closed_not_revisited or is_multiple: stats['count_multiple'] += 1
                if is_first_late: stats['count_first_late'] += 1
                if is_last_early: stats['count_last_early'] += 1
                if has_anom: stats['total_anomalies'] += 1

        analyzed_visites.sort(key=lambda x: (not x['has_anomaly'], x['heure_debut']))

        return jsonify({
            "status": "success",
            "stats": stats,
            "vendeurs": all_vendeurs,
            "dates": all_dates,
            "visites": analyzed_visites
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/anomalies", methods=["GET"])
def api_anomalies_list():
    try:
        anomalies = db_manager.get_all_anomalies()
        return jsonify({"status": "success", "anomalies": anomalies})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/anomalies", methods=["POST"])
def api_anomalies_create():
    try:
        data = request.get_json() or {}
        date = data.get("date")
        vendeur = data.get("vendeur")
        type_anomali = data.get("type_anomali")
        commentaire = data.get("commentaire")
        tag = data.get("tag")
        
        if not date or not vendeur or not type_anomali:
            return jsonify({"status": "error", "message": "Champs obligatoires manquants."}), 400
            
        success = db_manager.save_anomaly(date, vendeur, type_anomali, commentaire, tag)
        if success:
            return jsonify({"status": "success", "message": "Anomalie enregistrée avec succès."})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de l'enregistrement."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/anomalies/<int:anomaly_id>", methods=["DELETE"])
def api_anomalies_delete(anomaly_id):
    try:
        success = db_manager.delete_anomaly(anomaly_id)
        if success:
            return jsonify({"status": "success", "message": "Anomalie supprimée avec succès."})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de la suppression."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
@app.route("/tasks")
def tasks_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="tasks")

@app.route("/engagement")
def engagement_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="engagement")

@app.route("/api/engagements", methods=["GET"])
def api_engagements_list():
    try:
        engagements = db_manager.get_all_engagements()
        return jsonify({"status": "success", "engagements": engagements})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/engagements", methods=["POST"])
def api_engagements_create():
    try:
        data = request.get_json() or {}
        vendeur = (data.get("vendeur") or "").strip()
        periode = (data.get("periode") or "Jour").strip()
        date_engagement = (data.get("date_engagement") or "").strip()
        items = data.get("items") or []
        
        if not vendeur:
            return jsonify({"status": "error", "message": "Veuillez sélectionner un vendeur."}), 400
        if periode not in ["Jour", "Semaine", "Mois"]:
            return jsonify({"status": "error", "message": "Période invalide."}), 400
        if not date_engagement:
            from datetime import date
            date_engagement = date.today().strftime("%Y-%m-%d")
            
        engagement_id = db_manager.create_engagement(vendeur, periode, date_engagement, items)
        return jsonify({
            "status": "success",
            "message": "Engagement créé avec succès.",
            "engagement_id": engagement_id
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/engagements/<int:engagement_id>", methods=["DELETE"])
def api_engagements_delete(engagement_id):
    try:
        success = db_manager.delete_engagement(engagement_id)
        if success:
            return jsonify({"status": "success", "message": "Engagement supprimé avec succès."})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de la suppression."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/engagements/<int:engagement_id>", methods=["PUT", "POST"])
def api_engagements_update(engagement_id):
    try:
        data = request.get_json() or {}
        vendeur = (data.get("vendeur") or "").strip()
        periode = (data.get("periode") or "Jour").strip()
        date_engagement = (data.get("date_engagement") or "").strip()
        items = data.get("items") or []
        
        if not vendeur:
            return jsonify({"status": "error", "message": "Veuillez sélectionner un vendeur."}), 400
        if periode not in ["Jour", "Semaine", "Mois"]:
            return jsonify({"status": "error", "message": "Période invalide."}), 400
        if not date_engagement:
            from datetime import date
            date_engagement = date.today().strftime("%Y-%m-%d")
            
        db_manager.update_engagement(engagement_id, vendeur, periode, date_engagement, items)
        return jsonify({
            "status": "success",
            "message": "Engagement mis à jour avec succès."
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/vendeurs/raf", methods=["GET"])
def api_vendeur_raf():
    try:
        vendeur = request.args.get("vendeur", "").strip()
        if not vendeur:
            return jsonify({"status": "error", "message": "Vendeur non spécifié"}), 400

        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        
        # Get latest date in quantitative_data
        cursor.execute("SELECT date FROM quantitative_data ORDER BY date DESC LIMIT 1")
        latest_date_row = cursor.fetchone()
        if not latest_date_row:
            conn.close()
            return jsonify({"status": "success", "vendeur": vendeur, "raf_total": 0, "raf_jour": 0, "rest_days": 15, "ca": {"obj_mois": 0, "real": 0, "raf": 0, "raf_jour": 0}, "familles": {}})
            
        latest_date = latest_date_row[0]
        
        # Get rest_days for latest_date
        cursor.execute("SELECT rest_days FROM settings WHERE date = ?", (latest_date,))
        setting_row = cursor.fetchone()
        rest_days = setting_row[0] if (setting_row and setting_row[0] is not None) else 15
        if rest_days <= 0:
            rest_days = 15

        # Query quantitative rows for vendeur
        v_code = vendeur.split()[0].upper() if vendeur else ""
        cursor.execute(
            "SELECT famille, real, obj, obj_mois, raf FROM quantitative_data WHERE date = ? AND (vendeur = ? OR vendeur LIKE ?)",
            (latest_date, vendeur, f"{v_code}%")
        )
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()

        ca_info = {"obj_mois": 0, "real": 0, "raf": 0, "raf_jour": 0}
        familles_info = {}

        for r in rows:
            fam = r.get("famille", "")
            fam_upper = fam.upper()
            real = float(r.get("real") or 0)
            obj_partiel = float(r.get("obj") or 0)
            obj_mois = float(r.get("obj_mois") or 0)
            if obj_mois == 0 and obj_partiel > 0:
                obj_mois = round(obj_partiel * (24.0 / 10.0))
            
            raf_total_ttc = (obj_mois - real) * 1.2
            raf_jour = round(raf_total_ttc / rest_days) if rest_days > 0 else round(raf_total_ttc)

            if "C.A" in fam_upper or "CA" in fam_upper:
                ca_info = {
                    "obj_mois": obj_mois,
                    "real": real,
                    "raf": raf_total_ttc,
                    "raf_jour": raf_jour
                }
            else:
                familles_info[fam] = {
                    "obj_mois": obj_mois,
                    "real": real,
                    "raf": raf_total_ttc,
                    "raf_jour": raf_jour
                }

        if ca_info["raf"] == 0 and familles_info:
            sum_raf = sum(f["raf"] for f in familles_info.values())
            sum_obj = sum(f["obj_mois"] for f in familles_info.values())
            sum_real = sum(f["real"] for f in familles_info.values())
            ca_info = {
                "obj_mois": sum_obj,
                "real": sum_real,
                "raf": sum_raf,
                "raf_jour": round(sum_raf / rest_days) if rest_days > 0 else round(sum_raf)
            }

        return jsonify({
            "status": "success",
            "vendeur": vendeur,
            "date": latest_date,
            "rest_days": rest_days,
            "ca": ca_info,
            "familles": familles_info
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/tasks", methods=["GET"])
def api_tasks_list():
    try:
        tasks = db_manager.get_all_tasks()
        return jsonify({"status": "success", "tasks": tasks})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/tasks", methods=["POST"])
def api_tasks_create():
    try:
        data = request.get_json() or {}
        title = data.get("title")
        assignee_type = data.get("assignee_type")
        assignee = data.get("assignee")
        date = data.get("date")
        priority = data.get("priority")
        subtasks = data.get("subtasks", [])
        
        if not title or not assignee_type or not assignee or not date or not priority:
            return jsonify({"status": "error", "message": "Champs obligatoires manquants."}), 400
            
        success = db_manager.save_task(title, assignee_type, assignee, date, priority, status='Start', creator='me', subtasks=subtasks)
        if success:
            return jsonify({"status": "success", "message": "Tâche enregistrée avec succès."})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de l'enregistrement de la tâche."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/tasks/<int:task_id>", methods=["DELETE"])
def api_tasks_delete(task_id):
    try:
        success = db_manager.delete_task(task_id)
        if success:
            return jsonify({"status": "success", "message": "Tâche supprimée avec succès."})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de la suppression."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/tasks/<int:task_id>/status", methods=["PATCH"])
def api_tasks_update_status(task_id):
    try:
        data = request.get_json() or {}
        status = data.get("status")
        if not status:
            return jsonify({"status": "error", "message": "Statut manquant."}), 400
            
        success = db_manager.update_task_status(task_id, status)
        if success:
            return jsonify({"status": "success", "message": "Statut mis à jour."})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de la mise à jour."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/subtasks/<int:subtask_id>/toggle", methods=["PATCH"])
def api_subtasks_toggle(subtask_id):
    try:
        data = request.get_json() or {}
        completed = data.get("completed", False)
        
        success = db_manager.toggle_subtask_completed(subtask_id, completed)
        if success:
            return jsonify({"status": "success", "message": "Sous-tâche mise à jour."})
        else:
            return jsonify({"status": "error", "message": "Erreur de mise à jour."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/theme1")
def theme1():
    config = load_config()
    config["theme"] = "theme-1"
    save_config(config)
    return render_template("index.html", theme="theme-1", light_mode=config.get("light_mode", False))

@app.route("/theme2")
def theme2():
    config = load_config()
    config["theme"] = "theme-2"
    save_config(config)
    return render_template("index.html", theme="theme-2", light_mode=config.get("light_mode", False))

@app.route("/theme3")
def theme3():
    config = load_config()
    config["theme"] = "theme-3"
    save_config(config)
    return render_template("index.html", theme="theme-3", light_mode=config.get("light_mode", False))

@app.route("/theme4")
def theme4():
    config = load_config()
    config["theme"] = "theme-4"
    save_config(config)
    return render_template("index.html", theme="theme-4", light_mode=config.get("light_mode", False))

@app.route("/theme5")
def theme5():
    config = load_config()
    config["theme"] = "theme-5"
    save_config(config)
    return render_template("index.html", theme="theme-5", light_mode=config.get("light_mode", False))

@app.route("/favicon.ico")
def favicon():
    return app.send_static_file("logo.png")

@app.route("/api/data")
def get_all_data():
    try:
        category = request.args.get("category")
        date = request.args.get("date")
        
        if date and date != "default":
            data = db_manager.get_suivi_data(date)
            if not data:
                return jsonify({"status": "error", "message": f"Aucune donnée trouvée pour la date {date}."}), 404
        else:
            dates = db_manager.get_all_suivi_dates()
            if dates:
                data = db_manager.get_suivi_data(dates[0])
            else:
                processor = get_processor()
                data = processor.get_data()
        
        # Include FDV roster for dynamic Chef de Zone mapping
        data["fdv"] = db_manager.get_fdv_list()
        
        # Apply category filter if requested
        if category and category != "All":
            allowed_vendeurs = get_categorie(category)
            if not isinstance(allowed_vendeurs, list):
                allowed_vendeurs = [allowed_vendeurs]
                
            allowed_vendeurs_set = {v.strip().upper() for v in allowed_vendeurs if v}
            
            # Filter all relevant records
            data["quantitative"] = [r for r in data["quantitative"] if r["vendeur"].strip().upper() in allowed_vendeurs_set]
            data["qualitative"] = [r for r in data["qualitative"] if r["vendeur"].strip().upper() in allowed_vendeurs_set]
            data["focus_vmm"] = [r for r in data["focus_vmm"] if r["vendeur"].strip().upper() in allowed_vendeurs_set]
            data["focus_som"] = [r for r in data["focus_som"] if r["vendeur"].strip().upper() in allowed_vendeurs_set]
            
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/settings", methods=["POST"])
def update_settings():
    try:
        req_data = request.get_json() or {}
        rest_days = req_data.get("rest_days")
        date = req_data.get("date")
        
        if rest_days is None:
            return jsonify({"status": "error", "message": "Missing rest_days"}), 400
        
        rest_days = int(rest_days)
        
        exclude_families_raw = req_data.get("exclude_families", "")
        if isinstance(exclude_families_raw, str):
            exclude_families = [f.strip() for f in exclude_families_raw.split(",") if f.strip()]
        elif isinstance(exclude_families_raw, list):
            exclude_families = [str(f).strip() for f in exclude_families_raw if str(f).strip()]
        else:
            exclude_families = []
            
        if date and date != "default":
            # Save settings for this date
            db_manager.save_suivi_settings(date, rest_days, exclude_families)
            
            # Re-process file from db
            file_content, file_name = db_manager.get_suivi_file(date)
            if not file_content or not file_name:
                return jsonify({
                    "status": "error", 
                    "message": f"Le fichier Excel d'origine pour la date {date} est absent ou vide dans la base de données. Veuillez ré-importer le fichier pour cette date."
                }), 404

            process_and_save_suivi(date, file_content)
            return jsonify({"status": "success", "message": f"Paramètres de la date {date} mis à jour et recalculés."})
        else:
            # Load existing config
            config = load_config()
            
            config["rest_days"] = rest_days
            config["exclude_families"] = exclude_families
            
            # Save config
            save_config(config)
            
            # Re-run processor fix
            processor = ExcelProcessor(rest_days=rest_days, exclude_families=exclude_families)
            processor.fix_sheet(jour_rest=rest_days)
            
            return jsonify({"status": "success", "message": f"Rest days updated to {rest_days}, excluded families updated, and data recalculated."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/refresh", methods=["POST"])
def refresh_data():
    try:
        processor = get_processor()
        processor.get_day_work()
        processor.fix_sheet()
        return jsonify({"status": "success", "message": "Données rafraîchies depuis Excel avec succès."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/generate_report", methods=["POST"])
def run_ai_analysis():
    try:
        vendeur = request.args.get("vendeur")
        category = request.args.get("category")
        date = request.args.get("date")
        options_str = request.args.get("options")
        tax_mode = request.args.get("tax_mode", "TTC")
        report_type = request.args.get("report_type", "complet")
        language = request.args.get("language", "fr")
        model = request.args.get("model", "anthropic/claude-3.5-sonnet")
        
        # Parse options if provided
        options = None
        if options_str:
            selected_options = options_str.split(",")
            options = {
                "quanti": "quanti" in selected_options,
                "quali": "quali" in selected_options,
                "focus": "focus" in selected_options,
                "anomali": "anomali" in selected_options,
                "rappel": "rappel" in selected_options
            }

        report_content, summary_data = generate_report(vendeur=vendeur, category=category, date=date, options=options, tax_mode=tax_mode, report_type=report_type, language=language, model=model, return_data=True)
        focus_names = db_manager.get_focus_names()
        if report_content:
            return jsonify({"status": "success", "report": report_content, "summary_data": summary_data, "focus_names": focus_names})
        else:
            return jsonify({"status": "error", "message": "Erreur lors de la génération du rapport via OpenRouter."}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/send_whatsapp_image", methods=["POST"])
def send_whatsapp_image():
    try:
        import sys
        import base64
        import subprocess
        
        req_data = request.get_json() or {}
        phone = req_data.get("phone")
        image_data_b64 = req_data.get("image_data")
        
        if not phone or not image_data_b64:
            return jsonify({"status": "error", "message": "Numéro de téléphone et image_data requis."}), 400
            
        if "," in image_data_b64:
            image_data_b64 = image_data_b64.split(",")[1]
            
        img_bytes = base64.b64decode(image_data_b64)
        
        try:
            os.makedirs(EXCEL_DIR, exist_ok=True)
        except Exception:
            pass
        img_path = os.path.join(EXCEL_DIR, "temp_wa_card.png")
        
        with open(img_path, "wb") as f:
            f.write(img_bytes)
            
        python_executable = sys.executable
        cmd = [python_executable, "send_whatsapp.py", phone, img_path]
        
        subprocess.Popen(cmd, close_fds=True)
        
        return jsonify({
            "status": "success", 
            "message": "Envoi WhatsApp démarré. WhatsApp Web va s'ouvrir pour envoyer l'image."
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/send_bulk_whatsapp", methods=["POST"])
def send_bulk_whatsapp():
    try:
        import sys
        import time
        import json
        import subprocess
        
        req_data = request.get_json() or {}
        vendeurs = req_data.get("vendeurs", [])
        date = req_data.get("date", "")
        
        if not vendeurs:
            return jsonify({"status": "error", "message": "Liste de vendeurs requise."}), 400
            
        try:
            os.makedirs(EXCEL_DIR, exist_ok=True)
        except Exception:
            pass
        task_file_path = os.path.join(EXCEL_DIR, f"bulk_task_{int(time.time())}.json")
        
        with open(task_file_path, "w", encoding="utf-8") as f:
            json.dump({
                "vendeurs": vendeurs,
                "date": date
            }, f, indent=4)
            
        python_executable = sys.executable
        script_path = os.path.join(app_dir, "send_bulk_whatsapp_task.py")
        cmd = [python_executable, script_path, task_file_path]
        
        # Use DETACHED_PROCESS on Windows so the process survives independently
        try:
            CREATE_NO_WINDOW = 0x08000000
            subprocess.Popen(cmd, cwd=app_dir, creationflags=CREATE_NO_WINDOW)
        except TypeError:
            # Non-Windows fallback
            subprocess.Popen(cmd, cwd=app_dir)
        
        return jsonify({
            "status": "success", 
            "message": f"Envoi automatique en masse démarré pour {len(vendeurs)} vendeurs. Les rapports vont être générés et envoyés les uns après les autres."
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/clients/preview_visite_files", methods=["POST"])
def preview_visite_files_endpoint():
    """Extract vendedor code, name, total tournées and visit count for each uploaded file before user confirms saving."""
    try:
        import io
        import openpyxl
        files = request.files.getlist("files") or request.files.getlist("file")
        if not files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni."}), 400

        # Load client to tournee mapping for resolving tournées from client codes
        client_tournee_map = {}
        try:
            cv_conn = db_manager.get_cv_db_connection()
            for r in cv_conn.cursor().execute("SELECT c.code, l.name FROM clients c LEFT JOIN localites l ON c.localite_id = l.id"):
                if r[0] and r[1]:
                    client_tournee_map[str(r[0]).strip().upper()] = str(r[1]).strip()
            cv_conn.close()
        except Exception:
            pass

        preview_list = []
        vendeur_summary = {}

        for file in files:
            fname = file.filename
            if not fname or not fname.endswith(('.xlsx', '.xls')) or fname.startswith('~$'):
                continue

            file_stream = io.BytesIO(file.read())
            try:
                wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
                sheet_name = wb.sheetnames[0]
                for s in wb.sheetnames:
                    if "rapport" in s.lower() or "visite" in s.lower():
                        sheet_name = s
                        break
                ws = wb[sheet_name]

                v_code = ""
                v_name = ""
                agence = ""
                date_tournee = ""
                tournee_header = ""
                row_count = 0
                file_tournees = set()

                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 14:
                        for c_idx, val in enumerate(row):
                            if val is None:
                                continue
                            val_str = str(val).strip()
                            if "vendeur" in val_str.lower():
                                for offset in [1, 2, 3, 4, 5, 6]:
                                    if c_idx + offset < len(row) and row[c_idx + offset]:
                                        cand = str(row[c_idx + offset]).strip()
                                        if cand:
                                            if not v_code:
                                                v_code = cand
                                            elif not v_name and cand != v_code:
                                                v_name = cand
                            if "agence" in val_str.lower():
                                for offset in [1, 2, 3]:
                                    if c_idx + offset < len(row) and row[c_idx + offset]:
                                        agence = str(row[c_idx + offset]).strip()
                                        break
                            if "date" in val_str.lower() and "tourn" in val_str.lower():
                                for offset in [1, 2, 3, 4, 5]:
                                    if c_idx + offset < len(row) and row[c_idx + offset]:
                                        date_tournee = str(row[c_idx + offset]).split()[0]
                                        break
                            if val_str.lower() in ("tournée", "tournee"):
                                for offset in [1, 2, 3, 4, 5]:
                                    if c_idx + offset < len(row) and row[c_idx + offset]:
                                        tournee_header = str(row[c_idx + offset]).strip()
                                        if tournee_header:
                                            file_tournees.add(tournee_header)
                                        break
                    else:
                        if len(row) > 1 and row[1] is not None and str(row[1]).strip() != "":
                            row_count += 1
                            c_code = str(row[1]).strip().upper()
                            if c_code in client_tournee_map:
                                file_tournees.add(client_tournee_map[c_code])

                wb.close()

                resolved = None
                if v_code:
                    resolved_cand = db_manager.resolve_vendeur_full_name(v_code)
                    if resolved_cand and resolved_cand != "NON SPÉCIFIÉ" and len(resolved_cand.split()) > 1:
                        resolved = resolved_cand

                if not resolved:
                    resolved_fname = db_manager.resolve_vendeur_full_name(fname)
                    if resolved_fname and resolved_fname != "NON SPÉCIFIÉ" and len(resolved_fname.split()) > 1:
                        resolved = resolved_fname

                if not resolved and v_code and v_name:
                    resolved = f"{v_code} {v_name}".strip()
                elif not resolved and v_name:
                    resolved = v_name

                final_full = resolved if (resolved and resolved != "NON SPÉCIFIÉ") else (f"{v_code} {v_name}".strip() or fname.split('.')[0])
                parts = final_full.split(' ', 1)
                final_code = v_code or parts[0]
                final_clean_name = parts[1] if len(parts) > 1 else final_full

                # Group by vendeur
                v_key = final_full
                if v_key not in vendeur_summary:
                    vendeur_summary[v_key] = {
                        "vendeur_code": final_code,
                        "vendeur_name": final_clean_name,
                        "vendeur_full": final_full,
                        "tournees": set(),
                        "total_visites": 0,
                        "files_count": 0,
                        "dates": set()
                    }

                vendeur_summary[v_key]["total_visites"] += row_count
                vendeur_summary[v_key]["files_count"] += 1
                if date_tournee:
                    vendeur_summary[v_key]["dates"].add(date_tournee)
                for t in file_tournees:
                    if t and t != "-":
                        vendeur_summary[v_key]["tournees"].add(t)

                preview_list.append({
                    "file_name": fname,
                    "vendeur_code": final_code,
                    "vendeur_name": final_clean_name,
                    "vendeur_full": final_full,
                    "date": date_tournee or "-",
                    "tournee": tournee_header or (list(file_tournees)[0] if file_tournees else "-"),
                    "agence": agence or "AGADIR",
                    "total_visites": row_count
                })
            except Exception as f_err:
                resolved = db_manager.resolve_vendeur_full_name(fname)
                parts = resolved.split(' ', 1) if resolved else [fname.split('.')[0], "Vendeur"]
                f_code = parts[0]
                f_name = parts[1] if len(parts) > 1 else f_code
                
                if resolved not in vendeur_summary:
                    vendeur_summary[resolved] = {
                        "vendeur_code": f_code,
                        "vendeur_name": f_name,
                        "vendeur_full": resolved,
                        "tournees": set(),
                        "total_visites": 0,
                        "files_count": 0,
                        "dates": set()
                    }
                vendeur_summary[resolved]["files_count"] += 1

                preview_list.append({
                    "file_name": fname,
                    "vendeur_code": f_code,
                    "vendeur_name": f_name,
                    "vendeur_full": resolved,
                    "date": "-",
                    "tournee": "-",
                    "agence": "AGADIR",
                    "total_visites": 0
                })

        vendeurs_list = []
        for k, v in vendeur_summary.items():
            t_count = len(v["tournees"]) if v["tournees"] else max(1, v["files_count"])
            vendeurs_list.append({
                "vendeur_code": v["vendeur_code"],
                "vendeur_name": v["vendeur_name"],
                "vendeur_full": v["vendeur_full"],
                "total_tournees": t_count,
                "tournees_list": sorted(list(v["tournees"])),
                "total_visites": v["total_visites"],
                "files_count": v["files_count"],
                "dates": sorted(list(v["dates"]))
            })

        return jsonify({
            "status": "success",
            "files": preview_list,
            "vendeurs": vendeurs_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/analyse_visites", methods=["POST"])
def analyse_visites_endpoint():
    try:
        import io
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni."}), 400
        file = request.files["file"]
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"status": "error", "message": "Le fichier doit être un classeur Excel (.xlsx, .xls)."}), 400
            
        file_content = file.read()
        file_stream = io.BytesIO(file_content)
        
        # Smart detection: Check if uploaded file is ACM Master file
        file_name_lower = file.filename.lower()
        file_stream.seek(0)
        df_preview = pd.read_excel(file_stream, nrows=5)
        preview_text = str(df_preview.columns.tolist() + df_preview.values.tolist()).lower()
        
        if "acm" in file_name_lower or "role vendeur" in preview_text or "nbr clts inactifs" in preview_text or "clts inactifs" in preview_text:
            file_stream.seek(0)
            count = db_manager.import_acm_file(file_stream)
            if count > 0:
                stats = db_manager.get_acm_stats()
                return jsonify({
                    "status": "success",
                    "is_acm": True,
                    "message": f"Fichier ACM détecté et importé avec succès ! ({count:,} clients, Secteurs & Tournées enregistrés en base)",
                    "metadata": {
                        "vendeur": "ACM MASTER",
                        "date": "2026",
                        "tournee": "Toutes Tournées",
                        "file_name": file.filename
                    },
                    "summary": {
                        "total": count,
                        "ok": stats.get("active", 0),
                        "no_ok": stats.get("inactive", 0),
                        "acm": round((stats.get("active", 0) / count * 100), 1) if count > 0 else 0
                    },
                    "motifs": {"Clients Inactifs": stats.get("inactive", 0)},
                    "distance": {"average": 0, "anomalies": 0},
                    "clients_ok": [],
                    "clients_no_ok": []
                })

        # Standard visit report processing
        file_stream.seek(0)
        df_raw = pd.read_excel(file_stream, sheet_name=None)
        
        # Check sheet name
        sheet_name = None
        for name in df_raw.keys():
            if "rapport" in name.lower() or "visite" in name.lower():
                sheet_name = name
                break
        
        if not sheet_name:
            sheet_name = list(df_raw.keys())[0]
            
        df_sheet_raw = df_raw[sheet_name]
        
        # Safe extraction of metadata
        agence = "N/A"
        if len(df_sheet_raw) > 5 and len(df_sheet_raw.columns) > 5:
            val = df_sheet_raw.iloc[5, 5]
            if pd.notna(val):
                agence = str(val).strip()
                
        vendeur_code = ""
        if len(df_sheet_raw) > 8 and len(df_sheet_raw.columns) > 6:
            val = df_sheet_raw.iloc[8, 6]
            if pd.notna(val):
                vendeur_code = str(val).strip()
                
        vendeur_name = ""
        if len(df_sheet_raw) > 8 and len(df_sheet_raw.columns) > 8:
            val = df_sheet_raw.iloc[8, 8]
            if pd.notna(val):
                vendeur_name = str(val).strip()
                
        # Resolve seller full name from fdv database or filename (e.g. K91.xlsx -> K91 BAIZ MOHAMED)
        resolved_seller = db_manager.resolve_vendeur_full_name(file.filename)
        vendeur = resolved_seller if resolved_seller and resolved_seller != "NON SPÉCIFIÉ" else f"{vendeur_code} {vendeur_name}".strip() or "N/A"
        
        date_tournee = "N/A"
        if len(df_sheet_raw) > 4 and len(df_sheet_raw.columns) > 21:
            val = df_sheet_raw.iloc[4, 21]
            if pd.notna(val):
                date_tournee = str(val).split(' ')[0]
                
        tournee = "N/A"
        if len(df_sheet_raw) > 7 and len(df_sheet_raw.columns) > 21:
            val = df_sheet_raw.iloc[7, 21]
            if pd.notna(val):
                tournee = str(val).strip()
                
        # Skip metadata rows
        file_stream.seek(0)
        df_data = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=13)
        if df_data.empty:
            return jsonify({"status": "error", "message": "Le fichier Excel semble vide ou mal structuré."}), 400
            
        df_data.columns = [str(val).strip() if pd.notna(val) else f"col{i}" for i, val in enumerate(df_data.iloc[0])]
        df_data = df_data.iloc[1:]
        
        client_col = None
        for col in df_data.columns:
            if str(col).lower() == "client":
                client_col = col
                break
        if not client_col:
            for col in df_data.columns:
                if "col" not in str(col) and str(col).strip() != "":
                    client_col = col
                    break
        if not client_col:
            return jsonify({"status": "error", "message": "Impossible de trouver la colonne 'Client' dans le fichier."}), 400
            
        df_data = df_data.dropna(subset=[client_col])
        
        df_data = df_data.rename(columns={
            'Heure Dbut': 'Heure Début',
            'Heure Fin ': 'Heure Fin'
        })
        
        date_col = None
        for col in df_data.columns:
            if "date" in str(col).lower():
                date_col = col
                break
        if not date_col and 'col11' in df_data.columns:
            date_col = 'col11'

        h_dep = 'Heure Début' if 'Heure Début' in df_data.columns else ('col13' if 'col13' in df_data.columns else None)
        h_fin = 'Heure Fin' if 'Heure Fin' in df_data.columns else ('col14' if 'col14' in df_data.columns else None)
        dist_col = 'Distance' if 'Distance' in df_data.columns else ('col18' if 'col18' in df_data.columns else None)
        motif_col = 'Motif' if 'Motif' in df_data.columns else ('col19' if 'col19' in df_data.columns else None)
        note_col = 'Note' if 'Note' in df_data.columns else ('col24' if 'col24' in df_data.columns else None)
        nom_col = 'Nom' if 'Nom' in df_data.columns else ('col4' if 'col4' in df_data.columns else None)
        
        clients_ok = []
        clients_no_ok = []
        motifs_count = {}
        total_visits = len(df_data)
        ok_count = 0
        no_ok_count = 0
        
        sum_dist = 0
        valid_dist_count = 0
        dist_anomalies = 0
        
        for _, row in df_data.iterrows():
            c_code = str(row[client_col]).strip()
            c_name = str(row[nom_col]).strip() if nom_col else "N/A"
            c_date = date_tournee.split(' à ')[0].strip() if ' à ' in str(date_tournee) else date_tournee
            if date_col and pd.notna(row[date_col]):
                raw_d = str(row[date_col]).strip()
                if raw_d and raw_d.lower() not in ("nan", "none", "null", "nat"):
                    c_date = raw_d.split(' ')[0].strip()
            c_h_dep = str(row[h_dep]).strip() if h_dep else ""
            c_h_fin = str(row[h_fin]).strip() if h_fin else ""
            c_time = f"{c_h_dep} - {c_h_fin}" if c_h_dep or c_h_fin else "N/A"
            
            c_dist_str = str(row[dist_col]).split('.')[0].strip() if dist_col else "0"
            try:
                c_dist = int(c_dist_str)
                sum_dist += c_dist
                valid_dist_count += 1
                if c_dist > 100:
                    dist_anomalies += 1
            except:
                c_dist = 0
                
            c_motif = str(row[motif_col]).strip() if motif_col else "N/A"
            c_note = str(row[note_col]).strip() if note_col else ""
            if pd.isna(row[note_col]) or c_note.lower() in ("nan", "none", "null"):
                c_note = ""
                
            client_record = {
                "code": c_code,
                "name": c_name,
                "date": c_date,
                "time": c_time,
                "distance": f"{c_dist_str} m",
                "motif": c_motif,
                "note": c_note
            }
            
            if c_motif.upper() == "OK":
                ok_count += 1
                clients_ok.append(client_record)
            else:
                no_ok_count += 1
                clients_no_ok.append(client_record)
                motifs_count[c_motif] = motifs_count.get(c_motif, 0) + 1
                
        avg_dist = round(sum_dist / valid_dist_count, 1) if valid_dist_count > 0 else 0
        acm_pct = round((ok_count / total_visits) * 100, 1) if total_visits > 0 else 0

        all_row_dates = sorted(list(set(r["date"] for r in (clients_ok + clients_no_ok) if r.get("date") and r["date"] != "N/A")))
        display_date = f"{all_row_dates[0]} à {all_row_dates[-1]}" if len(all_row_dates) > 1 else (all_row_dates[0] if len(all_row_dates) == 1 else date_tournee)

        return jsonify({
            "status": "success",
            "metadata": {
                "agence": agence,
                "vendeur": vendeur,
                "date": display_date,
                "tournee": tournee,
                "file_name": file.filename
            },
            "summary": {
                "total": total_visits,
                "ok": ok_count,
                "no_ok": no_ok_count,
                "acm": acm_pct
            },
            "motifs": motifs_count,
            "distance": {
                "average": avg_dist,
                "anomalies": dist_anomalies
            },
            "clients_ok": clients_ok,
            "clients_no_ok": clients_no_ok
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Erreur d'analyse : {str(e)}"}), 500

@app.route("/api/clients/enregistrer_visites", methods=["POST"])
def enregistrer_visites_endpoint():
    try:
        data = request.json or {}
        metadata = data.get("metadata", {})
        clients_ok = data.get("clients_ok", [])
        clients_no_ok = data.get("clients_no_ok", [])
        
        file_name = metadata.get("file_name", "upload.xlsx")
        vendeur = db_manager.resolve_vendeur_full_name(metadata.get("vendeur") or file_name)
        date_tournee = metadata.get("date")
        ignore_tournee = data.get("ignore_tournee", True)
        clear_existing = data.get("clear_existing", True)
        tournee = "" if ignore_tournee else metadata.get("tournee", "")
        agence = metadata.get("agence")
        
        if not vendeur or not date_tournee:
            return jsonify({"status": "error", "message": "Métadonnées de visite incomplètes."}), 400
            
        all_records = clients_ok + clients_no_ok
        
        # Save to DB
        db_manager.save_visites_rapport(
            file_name,
            vendeur,
            date_tournee,
            tournee,
            agence,
            all_records,
            clear_all=clear_existing
        )
        
        return jsonify({
            "status": "success",
            "message": "Données de visites enregistrées avec succès dans la base de données !"
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/visites/upload-laptop-files", methods=["POST"])
def upload_laptop_visites_files():
    """Upload multiple Excel visit files directly from user's laptop and save file-by-file into visites_rapports without tournee."""
    try:
        import io
        files = request.files.getlist("files") or request.files.getlist("file")
        if not files or len(files) == 0 or (len(files) == 1 and files[0].filename == ""):
            return jsonify({"status": "error", "message": "Aucun fichier Excel sélectionné."}), 400

        ignore_tournee = request.form.get("ignore_tournee", "true").lower() in ("true", "1", "yes")

        # Clear existing visit reports before saving new uploaded batch
        db_manager.clear_all_visites_rapports()

        total_saved_records = 0
        imported_files = []

        for file in files:
            fname = file.filename
            if not fname or not fname.endswith(('.xlsx', '.xls')) or fname.startswith('~$'):
                continue

            file_content = file.read()
            file_stream = io.BytesIO(file_content)

            df_raw = pd.read_excel(file_stream, sheet_name=None)
            sheet_name = list(df_raw.keys())[0]
            for k in df_raw.keys():
                if "rapport" in k.lower() or "visite" in k.lower():
                    sheet_name = k
                    break

            df_sheet = df_raw[sheet_name]
            date_tournee = "2026-07-01"
            if len(df_sheet) > 4 and len(df_sheet.columns) > 21:
                val = df_sheet.iloc[4, 21]
                if pd.notna(val):
                    date_tournee = str(val).split(' ')[0]

            file_stream.seek(0)
            df_data = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=13)
            if df_data.empty:
                continue

            df_data.columns = [str(val).strip() if pd.notna(val) else f"col{i}" for i, val in enumerate(df_data.iloc[0])]
            df_data = df_data.iloc[1:]

            client_cols = [c for c in df_data.columns if str(c).lower() == "client"]
            if not client_cols:
                continue
            client_col = client_cols[0]
            df_data = df_data.dropna(subset=[client_col])

            date_col = [c for c in df_data.columns if "date" in str(c).lower()]
            date_col = date_col[0] if date_col else ("col11" if "col11" in df_data.columns else None)

            df_data = df_data.rename(columns={'Heure Dbut': 'Heure Début', 'Heure Fin ': 'Heure Fin'})
            h_dep = 'Heure Début' if 'Heure Début' in df_data.columns else ('col13' if 'col13' in df_data.columns else None)
            h_fin = 'Heure Fin' if 'Heure Fin' in df_data.columns else ('col14' if 'col14' in df_data.columns else None)
            dist_col = 'Distance' if 'Distance' in df_data.columns else ('col18' if 'col18' in df_data.columns else None)
            motif_col = 'Motif' if 'Motif' in df_data.columns else ('col19' if 'col19' in df_data.columns else None)
            note_col = 'Note' if 'Note' in df_data.columns else ('col24' if 'col24' in df_data.columns else None)
            nom_col = 'Nom' if 'Nom' in df_data.columns else ('col4' if 'col4' in df_data.columns else None)

            vendeur_full = db_manager.resolve_vendeur_full_name(fname)
            records = []

            for _, row in df_data.iterrows():
                c_code = str(row[client_col]).strip()
                c_name = str(row[nom_col]).strip() if nom_col and pd.notna(row[nom_col]) else "N/A"
                c_date = date_tournee
                if date_col and pd.notna(row[date_col]):
                    raw_d = str(row[date_col]).strip()
                    if raw_d and raw_d.lower() not in ("nan", "none", "null", "nat"):
                        c_date = raw_d.split(' ')[0].strip()
                c_h_dep = str(row[h_dep]).strip() if h_dep and pd.notna(row[h_dep]) else ""
                c_h_fin = str(row[h_fin]).strip() if h_fin and pd.notna(row[h_fin]) else ""
                c_time = f"{c_h_dep} - {c_h_fin}" if c_h_dep or c_h_fin else "N/A"
                dist_str = str(row[dist_col]).split('.')[0].strip() if dist_col and pd.notna(row[dist_col]) else "0"
                motif_val = str(row[motif_col]).strip() if motif_col and pd.notna(row[motif_col]) else "OK"
                note_val = str(row[note_col]).strip() if note_col and pd.notna(row[note_col]) else ""

                records.append({
                    "code": c_code,
                    "name": c_name,
                    "date": c_date,
                    "time": c_time,
                    "distance": dist_str,
                    "motif": motif_val,
                    "note": note_val
                })

            tournee_val = "" if ignore_tournee else ""
            ok = db_manager.save_visites_rapport(fname, vendeur_full, date_tournee, tournee_val, "", records)
            if ok:
                total_saved_records += len(records)
                imported_files.append({"file_name": fname, "vendeur": vendeur_full, "records": len(records)})

        return jsonify({
            "status": "success",
            "message": f"Importation réussie : {len(imported_files)} fichier(s) traité(s) ({total_saved_records:,} visites enregistrées)",
            "imported_files": imported_files,
            "total_records": total_saved_records
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Erreur lors de l'importation : {str(e)}"}), 500



@app.route("/api/clients/import_acm", methods=["POST"])
def import_acm_endpoint():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni."}), 400
        file = request.files["file"]
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"status": "error", "message": "Le fichier doit être au format Excel (.xlsx, .xls)."}), 400
            
        count = db_manager.import_acm_file(file)
        if count > 0:
            stats = db_manager.get_acm_stats()
            return jsonify({
                "status": "success",
                "message": f"{count:,} clients ACM (secteurs, tournées & inactifs) importés avec succès !",
                "stats": stats
            })
        else:
            return jsonify({"status": "error", "message": "Aucune donnée n'a pu être extraite du fichier ACM."}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/acm_stats", methods=["GET"])
def get_acm_stats_endpoint():
    try:
        stats = db_manager.get_acm_stats()
        return jsonify({"status": "success", "stats": stats})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/compare_visites", methods=["POST"])
def compare_visites_endpoint():
    try:
        import io
        files = []
        for key in ["file1", "file2", "file3"]:
            if key in request.files:
                f = request.files[key]
                if f and f.filename != "":
                    files.append(f)
                    
        if len(files) < 2:
            return jsonify({"status": "error", "message": "Veuillez charger au moins 2 rapports de visites pour pouvoir les comparer."}), 400
            
        tournees_data = []
        for i, file in enumerate(files):
            file_content = file.read()
            file_stream = io.BytesIO(file_content)
            df_raw = pd.read_excel(file_stream, sheet_name=None)
            
            sheet_name = None
            for name in df_raw.keys():
                if "rapport" in name.lower() or "visite" in name.lower():
                    sheet_name = name
                    break
            if not sheet_name:
                sheet_name = list(df_raw.keys())[0]
                
            df_sheet_raw = df_raw[sheet_name]
            
            # Extract metadata
            vendeur = "N/A"
            if len(df_sheet_raw) > 8 and len(df_sheet_raw.columns) > 8:
                v_code = str(df_sheet_raw.iloc[8, 6] or "").strip()
                v_name = str(df_sheet_raw.iloc[8, 8] or "").strip()
                vendeur = f"{v_code} {v_name}".strip() or "N/A"
                
            date_tournee = "N/A"
            if len(df_sheet_raw) > 4 and len(df_sheet_raw.columns) > 21:
                val = df_sheet_raw.iloc[4, 21]
                if pd.notna(val):
                    date_tournee = str(val).split(' ')[0]
                    
            # Skip metadata rows
            file_stream.seek(0)
            df_data = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=13)
            df_data.columns = [str(val).strip() if pd.notna(val) else f"col{i}" for i, val in enumerate(df_data.iloc[0])]
            df_data = df_data.iloc[1:]
            
            client_col = None
            for col in df_data.columns:
                if str(col).lower() == "client":
                    client_col = col
                    break
            if not client_col:
                for col in df_data.columns:
                    if "col" not in str(col) and str(col).strip() != "":
                        client_col = col
                        break
            if not client_col:
                continue
                
            df_data = df_data.dropna(subset=[client_col])
            
            motif_col = 'Motif' if 'Motif' in df_data.columns else ('col19' if 'col19' in df_data.columns else None)
            nom_col = 'Nom' if 'Nom' in df_data.columns else ('col4' if 'col4' in df_data.columns else None)
            
            clients_map = {}
            for _, row in df_data.iterrows():
                c_code = str(row[client_col]).strip()
                c_name = str(row[nom_col]).strip() if nom_col else "N/A"
                c_motif = str(row[motif_col]).strip() if motif_col else "N/A"
                
                if c_code not in clients_map:
                    clients_map[c_code] = {"name": c_name, "motif": c_motif}
                else:
                    curr_m = clients_map[c_code]["motif"]
                    if curr_m.upper() == "OK":
                        pass
                    elif c_motif.upper() == "OK":
                        clients_map[c_code] = {"name": c_name, "motif": "OK"}
                    elif not curr_m or curr_m == "N/A":
                        clients_map[c_code] = {"name": c_name, "motif": c_motif}
                    else:
                        clients_map[c_code] = {"name": c_name, "motif": c_motif}
                
            tournees_data.append({
                "date": date_tournee,
                "vendeur": vendeur,
                "clients": clients_map
            })
            
        # Compile all unique client codes
        all_clients = {}
        for td in tournees_data:
            for c_code, c_info in td["clients"].items():
                if c_code not in all_clients:
                    all_clients[c_code] = c_info["name"]
                    
        # Compare clients
        comparison_rows = []
        summary_stats = {
            "always_ok": 0,
            "never_ok": 0,
            "billing_loss": 0,
            "billing_gain": 0,
            "inconsistent": 0
        }
        
        for c_code, c_name in all_clients.items():
            motifs = []
            for td in tournees_data:
                if c_code in td["clients"]:
                    motifs.append(td["clients"][c_code]["motif"])
                else:
                    motifs.append("Non visité")
                    
            # Synthesis
            bools = []
            for m in motifs:
                if m.upper() == "OK":
                    bools.append(True)
                elif m == "Non visité":
                    bools.append(None)
                else:
                    bools.append(False)
                    
            synth = "Inconstant"
            visited_bools = [b for b in bools if b is not None]
            
            if len(visited_bools) == 0:
                synth = "Non visité"
            elif all(b is True for b in visited_bools):
                synth = "Toujours Facturé"
                summary_stats["always_ok"] += 1
            elif all(b is False for b in visited_bools):
                synth = "Jamais Facturé"
                summary_stats["never_ok"] += 1
            else:
                actual_vis = [(idx, val) for idx, val in enumerate(bools) if val is not None]
                if len(actual_vis) >= 2:
                    first_val = actual_vis[0][1]
                    last_val = actual_vis[-1][1]
                    if first_val is True and last_val is False:
                        synth = "Perte de facturation"
                        summary_stats["billing_loss"] += 1
                    elif first_val is False and last_val is True:
                        synth = "Gagné (Facturé en fin)"
                        summary_stats["billing_gain"] += 1
                    else:
                        synth = "Inconstant"
                        summary_stats["inconsistent"] += 1
                else:
                    synth = "Inconstant"
                    summary_stats["inconsistent"] += 1
                    
            comparison_rows.append({
                "code": c_code,
                "name": c_name,
                "motifs": motifs,
                "synthesis": synth
            })
            
        dates = [td["date"] for td in tournees_data]
        vendeur = tournees_data[0]["vendeur"]
        
        return jsonify({
            "status": "success",
            "vendeur": vendeur,
            "dates": dates,
            "summary": summary_stats,
            "comparison": comparison_rows
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Erreur de comparaison : {str(e)}"}), 500

@app.route("/api/clients/visites", methods=["GET"])
def get_visites_rapport_endpoint():
    try:
        vendeur = request.args.get("vendeur")
        date_visite = request.args.get("date")
        
        if not vendeur or not date_visite:
            return jsonify({"status": "error", "message": "Vendeur et Date requis."}), 400
            
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT file_name, tournee, agence, client_code, client_nom, heure, distance, motif, note
            FROM visites_rapports
            WHERE vendeur = ? AND date_visite = ?
        """, (vendeur, date_visite))
        rows = cursor.fetchall()
        conn.close()
        
        if len(rows) == 0:
            return jsonify({
                "status": "success",
                "message": "Aucun détail de visite enregistré dans la base pour ce vendeur et cette date.",
                "data": None
            })
            
        clients_ok = []
        clients_no_ok = []
        motifs_count = {}
        ok_count = 0
        no_ok_count = 0
        sum_dist = 0
        valid_dist_count = 0
        dist_anomalies = 0
        
        tournee = rows[0]["tournee"]
        agence = rows[0]["agence"]
        
        for row in rows:
            c_code = row["client_code"]
            c_name = row["client_nom"] or "N/A"
            c_time = row["heure"] or "N/A"
            c_dist = row["distance"] or 0
            c_motif = row["motif"] or "N/A"
            c_note = row["note"] or ""
            
            sum_dist += c_dist
            valid_dist_count += 1
            if c_dist > 100:
                dist_anomalies += 1
                
            client_record = {
                "code": c_code,
                "name": c_name,
                "time": c_time,
                "distance": f"{c_dist} m",
                "motif": c_motif,
                "note": c_note
            }
            
            if c_motif.upper() == "OK":
                ok_count += 1
                clients_ok.append(client_record)
            else:
                no_ok_count += 1
                clients_no_ok.append(client_record)
                motifs_count[c_motif] = motifs_count.get(c_motif, 0) + 1
                
        total_visits = len(rows)
        avg_dist = round(sum_dist / valid_dist_count, 1) if valid_dist_count > 0 else 0
        acm_pct = round((ok_count / total_visits) * 100, 1) if total_visits > 0 else 0
        
        return jsonify({
            "status": "success",
            "data": {
                "metadata": {
                    "vendeur": vendeur,
                    "date": date_visite,
                    "tournee": tournee,
                    "agence": agence
                },
                "summary": {
                    "total": total_visits,
                    "ok": ok_count,
                    "no_ok": no_ok_count,
                    "acm": acm_pct
                },
                "motifs": motifs_count,
                "distance": {
                    "average": avg_dist,
                    "anomalies": dist_anomalies
                },
                "clients_ok": clients_ok,
                "clients_no_ok": clients_no_ok
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Erreur de lecture : {str(e)}"}), 500

@app.route("/api/afacturer/tournees", methods=["GET"])
def get_afacturer_tournees():
    """Return vendeurs and secteurs with their distinct tournée+date entries.

    Sources:
      - fdv              → vendeur catalog (code + name + secteur)
      - clients          → localites assigned per vendeur (vendeur_som / vendeur_vmm)
      - visites_rapports → actual visit dates & tournée names
      - secteurs         → secteur master list
    """
    try:
        from collections import defaultdict
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()

        # ── 1. Vendeur catalog from fdv ──────────────────────────────────────
        cursor.execute("SELECT vendeur, secteur FROM fdv")
        fdv_rows = cursor.fetchall()

        # Parse "CODE NOM" format → { "D48": {"name": "D48 IBACH MOHAMED", "secteur": "TAROUDANT"} }
        fdv_map = {}
        for row in fdv_rows:
            vendeur_full = (row["vendeur"] or "").strip()
            secteur_fdv  = (row["secteur"] or "").strip()
            parts = vendeur_full.split(" ", 1)
            code = parts[0].strip().upper() if parts else vendeur_full.upper()
            if code:
                fdv_map[code] = {"name": vendeur_full, "secteur": secteur_fdv, "code": code}

        # ── 2. Localites per vendeur from clients table ──────────────────────
        cursor.execute("""
            SELECT
                c.vendeur_som,
                c.vendeur_vmm,
                l.name  AS localite,
                s.name  AS secteur,
                COUNT(*) AS client_count
            FROM clients c
            LEFT JOIN localites l ON l.id = c.localite_id
            LEFT JOIN secteurs  s ON s.id = c.secteur_id
            WHERE l.name IS NOT NULL
            GROUP BY c.vendeur_som, c.vendeur_vmm, l.name, s.name
            ORDER BY client_count DESC
        """)
        client_rows = cursor.fetchall()

        # Build: vcode → [ {localite, secteur, count} ]
        vendeur_localites = defaultdict(list)
        for row in client_rows:
            localite = row["localite"]
            secteur  = row["secteur"]
            count    = row["client_count"]
            for vfield in ("vendeur_som", "vendeur_vmm"):
                raw = (row[vfield] or "").strip()
                if not raw:
                    continue
                vcode = raw.split(" ", 1)[0].strip().upper()
                vendeur_localites[vcode].append({
                    "localite": localite,
                    "secteur":  secteur,
                    "count":    count
                })

        # ── 3. Visit dates & tournées from visites_rapports ─────────────────
        cursor.execute("""
            SELECT
                vendeur,
                date_visite AS date,
                tournee,
                COUNT(DISTINCT client_code) AS visit_count
            FROM visites_rapports
            WHERE date_visite IS NOT NULL
            GROUP BY vendeur, date_visite, tournee
            ORDER BY vendeur, date_visite
        """)
        visite_rows = cursor.fetchall()

        # Build: vcode → [ {date, tournee, secteur, count} ]
        vendeur_visits = defaultdict(list)
        for row in visite_rows:
            raw   = (row["vendeur"] or "").strip()
            vcode = raw.split(" ", 1)[0].strip().upper()
            tournee_name = (row["tournee"] or "TOURNÉE GÉNÉRALE").strip()
            secteur = fdv_map.get(vcode, {}).get("secteur", "")
            vendeur_visits[vcode].append({
                "date":    row["date"],
                "tournee": tournee_name,
                "secteur": secteur,
                "count":   row["visit_count"]
            })

        # ── 4. All secteurs ──────────────────────────────────────────────────
        cursor.execute("SELECT name FROM secteurs ORDER BY name")
        all_secteurs = [r["name"] for r in cursor.fetchall()]
        conn.close()

        # ── 5. Build vendeurs list ───────────────────────────────────────────
        all_vcodes = set(fdv_map.keys()) | set(vendeur_localites.keys()) | set(vendeur_visits.keys())
        vendeurs_list = []

        for vcode in sorted(all_vcodes):
            info = fdv_map.get(vcode, {"code": vcode, "name": vcode, "secteur": ""})

            # Deduplicate localites, keep top 20 by client count
            loc_seen = {}
            for loc in vendeur_localites.get(vcode, []):
                key = loc["localite"]
                if key not in loc_seen or loc["count"] > loc_seen[key]["count"]:
                    loc_seen[key] = loc
            sorted_localites = sorted(loc_seen.values(), key=lambda x: -x["count"])

            vendeurs_list.append({
                "code":          info["code"],
                "name":          info["name"],
                "secteur":       info["secteur"],
                "tournee_dates": vendeur_visits.get(vcode, []),
                "localites":     [l["localite"] for l in sorted_localites[:20]]
            })

        # ── 6. Build secteurs list ───────────────────────────────────────────
        secteur_vendeur_map = defaultdict(list)
        for v in vendeurs_list:
            sec = v.get("secteur", "")
            if sec:
                secteur_vendeur_map[sec].append({
                    "vendeur_code": v["code"],
                    "vendeur_name": v["name"],
                    "tournee_dates": v["tournee_dates"]
                })

        all_sec_names = set(secteur_vendeur_map.keys()) | set(all_secteurs)
        secteurs_list = []
        for sec in sorted(all_sec_names):
            fdv_vendeurs = secteur_vendeur_map.get(sec, [])
            secteurs_list.append({
                "name":         sec,
                "vendeurs":     fdv_vendeurs,
                "tournee_dates": [td for v in fdv_vendeurs for td in v["tournee_dates"]]
            })

        return jsonify({
            "status":   "success",
            "vendeurs": vendeurs_list,
            "secteurs": secteurs_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/afacturer/vendeur-tournees", methods=["GET"])
def get_vendeur_tournees():
    """Return all tournées for a specific vendeur.
    Sources:
      - visites_rapports (database.db)    → actual visit dates & tournée names
      - clients_vendeurs.db               → all localites/tournées assigned to this vendeur
    """
    try:
        vendeur_raw = request.args.get("vendeur", "").strip()
        if not vendeur_raw:
            return jsonify({"status": "error", "message": "vendeur param required"}), 400

        vcode = vendeur_raw.split(" ", 1)[0].strip().upper()

        # ── 1. Real visits from database.db → visites_rapports ───────────────
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                date_visite  AS date,
                tournee,
                COUNT(DISTINCT client_code) AS client_count
            FROM visites_rapports
            WHERE UPPER(vendeur) LIKE ? OR UPPER(vendeur) LIKE ?
            GROUP BY date_visite, tournee
            ORDER BY date_visite DESC, client_count DESC
        """, (f"{vcode}%", f"%{vcode}%"))
        visite_rows = cursor.fetchall()
        conn.close()

        tournee_dates = []
        for row in visite_rows:
            tournee_dates.append({
                "date":         row["date"],
                "tournee":      row["tournee"] or "TOURNÉE GÉNÉRALE",
                "client_count": row["client_count"],
                "source":       "visite"
            })

        # ── 2. Localites from clients_vendeurs.db ────────────────────────────
        cv_conn = db_manager.get_cv_db_connection()
        cv_cursor = cv_conn.cursor()
        cv_cursor.execute("""
            SELECT
                l.name  AS localite,
                s.name  AS secteur,
                COUNT(*) AS client_count
            FROM clients c
            LEFT JOIN localites l ON l.id = c.localite_id
            LEFT JOIN secteurs  s ON s.id = c.secteur_id
            WHERE (UPPER(c.vendeur_som) LIKE ? OR UPPER(c.vendeur_vmm) LIKE ?)
              AND l.name IS NOT NULL
            GROUP BY l.name, s.name
            ORDER BY client_count DESC
        """, (f"{vcode}%", f"{vcode}%"))
        localite_rows = cv_cursor.fetchall()
        cv_conn.close()

        localites = []
        for row in localite_rows:
            localites.append({
                "localite":     row["localite"],
                "secteur":      row["secteur"],
                "client_count": row["client_count"]
            })

        return jsonify({
            "status":          "success",
            "vendeur":         vendeur_raw,
            "tournee_dates":   tournee_dates,
            "localites":       localites,
            "total_visites":   len(tournee_dates),
            "total_localites": len(localites)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()

@app.route("/api/visites/filters", methods=["GET"])
def api_visites_filters():
    """Retrieve distinct vendeurs from FDV table for Chakib Elfil team only and dates strictly from visites_rapports database table."""
    try:
        vendeur_filter = request.args.get("vendeur", "").strip()
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        
        # 1. Fetch exclusively from FDV table for Chakib Elfil team
        cursor.execute("SELECT DISTINCT vendeur FROM fdv WHERE vendeur IS NOT NULL AND vendeur != '' AND UPPER(cdz) LIKE '%CHAKIB%' ORDER BY vendeur ASC")
        chakib_vendeurs = [r[0].strip() for r in cursor.fetchall() if r[0] and r[0].strip()]
        
        # Fallback if empty
        if not chakib_vendeurs:
            cursor.execute("SELECT DISTINCT vendeur FROM fdv WHERE vendeur IS NOT NULL AND vendeur != '' ORDER BY vendeur ASC")
            chakib_vendeurs = [r[0].strip() for r in cursor.fetchall() if r[0] and r[0].strip()]
        
        # 2. Fetch distinct dates and their primary tournee strictly from visites_rapports
        if vendeur_filter and vendeur_filter != "All":
            vcode = vendeur_filter.split(" ", 1)[0].strip().upper()
            cursor.execute("""
                SELECT date_visite, tournee, COUNT(*) as cnt
                FROM visites_rapports 
                WHERE date_visite IS NOT NULL AND date_visite != '' AND date_visite != '-'
                  AND (UPPER(TRIM(vendeur)) = ? OR UPPER(vendeur) LIKE ? OR UPPER(vendeur) LIKE ?)
                GROUP BY date_visite, tournee
                ORDER BY date_visite DESC, cnt DESC
            """, (vendeur_filter.upper(), f"{vcode} %", f"{vcode}%"))
        else:
            cursor.execute("""
                SELECT date_visite, tournee, COUNT(*) as cnt
                FROM visites_rapports 
                WHERE date_visite IS NOT NULL AND date_visite != '' AND date_visite != '-'
                GROUP BY date_visite, tournee
                ORDER BY date_visite DESC, cnt DESC
            """)

        date_tournee_map = {}
        for r in cursor.fetchall():
            d_val = r[0].strip() if r[0] else ''
            t_val = r[1].strip() if r[1] else ''
            if d_val and d_val not in date_tournee_map:
                date_tournee_map[d_val] = t_val

        dates = []
        for d, t in date_tournee_map.items():
            dates.append({
                "date": d,
                "tournee": t,
                "label": f"{d} - {t}" if t else d
            })

        conn.close()
        
        return jsonify({"status": "success", "vendeurs": chakib_vendeurs, "dates": dates})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e), "vendeurs": [], "dates": []})


@app.route("/api/visites/tournee-stats", methods=["GET"])
def get_visites_tournee_stats():
    """Retrieve visit statistics per tournée by matching visites_rapports with clients_vendeurs.db.
    Filters by selected vendeur and date_visite (day).
    For each client_code, looks up secteur, tourné, and vendeur in clients_vendeurs.db.
    Returns breakdown statistics per tourné: total clients, magasin fermé, ok, responsable absent, stock suffisant.
    """
    try:
        vendeur_raw = request.args.get("vendeur", "").strip()
        date_val    = request.args.get("date", "").strip() or request.args.get("date_visite", "").strip()

        conn = db_manager.get_db_connection()
        try:
            conn.execute(f"ATTACH DATABASE '{db_manager.CV_DB_PATH}' AS cv")
        except Exception:
            pass

        cursor = conn.cursor()
        vcode = vendeur_raw.split(" ", 1)[0].strip().upper() if vendeur_raw else ""

        params = []
        where_clauses = []

        if vendeur_raw:
            vcode = vendeur_raw.split(" ", 1)[0].strip().upper()
            clean_name = vendeur_raw.strip().upper()
            where_clauses.append("(UPPER(TRIM(vr.vendeur)) = ? OR UPPER(vr.vendeur) LIKE ? OR UPPER(vr.vendeur) LIKE ?)")
            params.extend([clean_name, f"{vcode} %", f"{vcode}%"])

        if date_val:
            where_clauses.append("(vr.date_visite = ? OR vr.date_visite LIKE ?)")
            params.extend([date_val, f"%{date_val}%"])

        where_str = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
        SELECT 
            COALESCE(vr.tournee, 'Tournée non spécifiée') AS tournee_name,
            COALESCE(vr.agence, 'Secteur non spécifié') AS secteur_name,
            vr.vendeur AS cv_vendeur,
            MAX(vr.date_visite) AS date_visite,
            GROUP_CONCAT(DISTINCT vr.date_visite) AS dates_list,
            COUNT(DISTINCT vr.client_code) AS total_visites,
            SUM(CASE WHEN UPPER(TRIM(vr.motif)) = 'OK' THEN 1 ELSE 0 END) AS ok_count,
            SUM(CASE WHEN UPPER(TRIM(vr.motif)) LIKE '%FERME%' OR UPPER(TRIM(vr.motif)) LIKE '%FERMÉ%' THEN 1 ELSE 0 END) AS magasin_ferme_count,
            SUM(CASE WHEN UPPER(TRIM(vr.motif)) LIKE '%RESPONSABLE%' OR UPPER(TRIM(vr.motif)) LIKE '%ABSENT%' THEN 1 ELSE 0 END) AS responsable_absent_count,
            SUM(CASE WHEN UPPER(TRIM(vr.motif)) LIKE '%STOCK%' OR UPPER(TRIM(vr.motif)) LIKE '%SUFISANT%' OR UPPER(TRIM(vr.motif)) LIKE '%SUFFISANT%' THEN 1 ELSE 0 END) AS stock_suffisant_count,
            SUM(CASE WHEN UPPER(TRIM(vr.motif)) NOT IN ('OK') 
                     AND UPPER(TRIM(vr.motif)) NOT LIKE '%FERME%' AND UPPER(TRIM(vr.motif)) NOT LIKE '%FERMÉ%'
                     AND UPPER(TRIM(vr.motif)) NOT LIKE '%RESPONSABLE%' AND UPPER(TRIM(vr.motif)) NOT LIKE '%ABSENT%'
                     AND UPPER(TRIM(vr.motif)) NOT LIKE '%STOCK%' AND UPPER(TRIM(vr.motif)) NOT LIKE '%SUFISANT%' AND UPPER(TRIM(vr.motif)) NOT LIKE '%SUFFISANT%'
                THEN 1 ELSE 0 END) AS autres_count
        FROM visites_rapports vr
        {where_str}
        GROUP BY tournee_name
        ORDER BY date_visite DESC, total_visites DESC
        """

        cursor.execute(query, params)
        rows = cursor.fetchall()

        registered_map = {}
        try:
            reg_rows = cursor.execute("SELECT l.name, COUNT(c.id) FROM cv.clients c JOIN cv.localites l ON c.localite_id = l.id GROUP BY l.name").fetchall()
            registered_map = {r[0]: r[1] for r in reg_rows if r[0]}
        except Exception:
            pass

        results = []
        for r in rows:
            t_name = r["tournee_name"]
            total_registered = registered_map.get(t_name, 0)

            results.append({
                "tournee": t_name,
                "date": r["date_visite"] or "",
                "dates_list": r["dates_list"] or "",
                "secteur": r["secteur_name"],
                "vendeur": r["cv_vendeur"],
                "total_clients_enregistres": total_registered,
                "total_clients_visites": r["total_visites"],
                "statistics": {
                    "ok": r["ok_count"],
                    "magasin_ferme": r["magasin_ferme_count"],
                    "responsable_absent": r["responsable_absent_count"],
                    "stock_suffisant": r["stock_suffisant_count"],
                    "autres": r["autres_count"]
                }
            })

        conn.close()

        return jsonify({
            "status": "success",
            "vendeur": vendeur_raw,
            "date": date_val,
            "total_tournees": len(results),
            "tournees": results
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/afacturer/clients", methods=["GET"])
def get_afacturer_clients():
    """Return clients and their visit data for a specific (vendeur_code, date, localite).
    Matches sellers across short codes and full names, checks both visit tables,
    and returns full data for UI rendering.
    """
    try:
        vendeur_raw = request.args.get("vendeur", "").strip()
        date        = request.args.get("date", "").strip() or request.args.get("dates", "").strip()
        localite    = request.args.get("localite", "").strip() or request.args.get("tournee", "").strip()

        if not date or not localite:
            return jsonify({"status": "error", "message": "date and localite required"}), 400

        vendeur_code = vendeur_raw.upper()
        v_short = vendeur_code.split()[0] if vendeur_code else ""

        conn = db_manager.get_db_connection()
        cursor = conn.cursor()

        # 1. Get ALL client codes in this localite from clients table
        cursor.execute("""
            SELECT c.code, c.name, l.name as localite, s.name as secteur,
                   c.vendeur_som, c.vendeur_vmm
            FROM clients c
            JOIN localites l ON c.localite_id = l.id
            JOIN secteurs s ON c.secteur_id = s.id
            WHERE UPPER(l.name) = UPPER(?) OR UPPER(l.name) LIKE UPPER(?)
            ORDER BY c.code
        """, (localite, f"%{localite}%"))
        acm_clients = {r["code"]: dict(r) for r in cursor.fetchall()}

        # Fallback: if no clients found in localites table, check visites_rapports
        if not acm_clients:
            cursor.execute("""
                SELECT DISTINCT client_code as code, client_nom as name, tournee as localite, agence as secteur,
                       '' as vendeur_som, '' as vendeur_vmm
                FROM visites_rapports
                WHERE (UPPER(tournee) = UPPER(?) OR UPPER(tournee) LIKE UPPER(?))
                  AND (date_visite = ? OR date_visite LIKE ?)
            """, (localite, f"%{localite}%", date, f"%{date}%"))
            acm_clients = {r["code"]: dict(r) for r in cursor.fetchall() if r["code"]}

        acm_codes = list(acm_clients.keys())
        visited_map = {}

        if acm_codes:
            placeholders = ",".join(["?" for _ in acm_codes])

            # Query 2a: vendeur_tournees_visits
            query_vtv = f"""
                SELECT client_code, client_name, heure_debut, heure_fin, duree_minutes, motif, note, facture_status, distance
                FROM vendeur_tournees_visits
                WHERE (UPPER(vendeur_code) = UPPER(?) OR UPPER(vendeur_code) = UPPER(?) OR UPPER(vendeur_name) LIKE UPPER(?))
                  AND (date = ? OR date_visite = ?)
                  AND client_code IN ({placeholders})
                ORDER BY heure_debut
            """
            params_vtv = [vendeur_code, v_short, f"%{vendeur_code}%", date, date] + acm_codes
            try:
                cursor.execute(query_vtv, params_vtv)
                for r in cursor.fetchall():
                    code = r["client_code"]
                    if code not in visited_map: visited_map[code] = []
                    visited_map[code].append({
                        "heure_debut":    r["heure_debut"] or "",
                        "heure_fin":      r["heure_fin"] or "",
                        "duree_minutes":  r["duree_minutes"] or 0,
                        "motif":          r["motif"] or "",
                        "note":           r["note"] or "",
                        "facture_status": r["facture_status"] or ("AVEC FACTURE" if (r["motif"] or "").upper() == "OK" else "SANS FACTURE"),
                        "distance":       r["distance"] or ""
                    })
            except Exception as ex:
                print(f"vendeur_tournees_visits fetch error: {ex}")

            # Query 2b: visites_rapports
            query_vr = f"""
                SELECT client_code, client_nom as client_name, heure, motif, note, distance
                FROM visites_rapports
                WHERE (UPPER(vendeur) LIKE UPPER(?) OR UPPER(vendeur) LIKE UPPER(?))
                  AND (date_visite = ? OR date_visite LIKE ?)
                  AND client_code IN ({placeholders})
            """
            params_vr = [f"%{v_short}%", f"%{vendeur_code}%", date, f"%{date}%"] + acm_codes
            try:
                cursor.execute(query_vr, params_vr)
                for r in cursor.fetchall():
                    code = r["client_code"]
                    if code not in visited_map: visited_map[code] = []
                    motif_val = r["motif"] or ""
                    fact_status = "AVEC FACTURE" if motif_val.upper() == "OK" else "SANS FACTURE"
                    visited_map[code].append({
                        "heure_debut":    r["heure"] or "",
                        "heure_fin":      "",
                        "duree_minutes":  0,
                        "motif":          motif_val,
                        "note":           r["note"] or "",
                        "facture_status": fact_status,
                        "distance":       r["distance"] or ""
                    })
            except Exception as ex:
                print(f"visites_rapports fetch error: {ex}")

        conn.close()

        # Build result: all ACM clients + their visit status
        clients_result = []
        total_avec = 0
        total_sans = 0
        total_non_visite = 0
        anomalie = 0
        big_facture = 0
        small_facture = 0

        motifs_counter = {}

        for code, acm in acm_clients.items():
            visits = visited_map.get(code, [])
            is_visited = len(visits) > 0

            # Determine best facture status
            if visits:
                priority = {"AVEC FACTURE": 5, "BIG FACTURE": 4, "SMALL FACTURE": 3,
                            "ANOMALIE AVEC FACTURE": 2, "SANS FACTURE": 1}
                best = max(visits, key=lambda v: priority.get(v["facture_status"], 0))
                status = best["facture_status"]
                motif  = best["motif"]
            else:
                status = "NON VISITÉ"
                motif  = "Non visité"

            # Count KPIs
            if status in ("AVEC FACTURE", "BIG FACTURE", "SMALL FACTURE", "ANOMALIE AVEC FACTURE"):
                total_avec += 1
                if status == "BIG FACTURE":           big_facture += 1
                if status == "SMALL FACTURE":         small_facture += 1
                if status == "ANOMALIE AVEC FACTURE": anomalie += 1
            elif status == "SANS FACTURE":
                total_sans += 1
            else:
                total_non_visite += 1

            motifs_counter[motif] = motifs_counter.get(motif, 0) + 1

            clients_result.append({
                "code":           code,
                "name":           acm["name"],
                "localite":       localite,
                "secteur":        acm.get("secteur", ""),
                "vendeur_som":    acm.get("vendeur_som", ""),
                "vendeur_vmm":    acm.get("vendeur_vmm", ""),
                "visits":         visits,
                "visit_count":    len(visits),
                "motifs":         [motif],
                "motif":          motif,
                "facture_status": status,
                "is_visited":     is_visited,
                "has_facture":    status in ("AVEC FACTURE", "BIG FACTURE", "SMALL FACTURE", "ANOMALIE AVEC FACTURE"),
            })

        # Sort: visited first (by heure_debut), then non-visited
        clients_result.sort(key=lambda x: (
            0 if x["is_visited"] else 1,
            x["visits"][0]["heure_debut"] if x["visits"] else "99:99"
        ))

        taux = round((total_avec / len(clients_result)) * 100, 1) if clients_result else 0

        return jsonify({
            "status":              "success",
            "vendeur_code":        vendeur_code,
            "date":                date,
            "localite":            localite,
            "total_clients":       len(clients_result),
            "total_avec":          total_avec,
            "total_avec_facture":  total_avec,
            "toujours_facture":    total_avec,
            "pertes":              0,
            "gains":               0,
            "jamais_facture":      total_sans,
            "total_inactifs":      total_non_visite,
            "total_sans":          total_sans,
            "total_non_visite":    total_non_visite,
            "anomalie":            anomalie,
            "big_facture":         big_facture,
            "small_facture":       small_facture,
            "taux_facturation":    taux,
            "motifs":              motifs_counter,
            "comparison":          clients_result,
            "clients":             clients_result
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


# Alias for backward compat with existing JS triggerAnalysis call
@app.route("/api/clients/analyse", methods=["GET"])
def get_clients_analyse():
    """Redirect-like alias: delegates to /api/afacturer/clients."""
    from flask import redirect
    vendeur = request.args.get("vendeur", "")
    date    = request.args.get("dates", "")
    tournee = request.args.get("tournee", "")
    return redirect(f"/api/afacturer/clients?vendeur={vendeur}&date={date}&localite={tournee}")


@app.route("/api/clients/visites_disponibles", methods=["GET"])
def get_visites_disponibles_endpoint():
    try:
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        
        # 1. Fetch visits from DB
        cursor.execute("""
            SELECT DISTINCT tournee, date_visite, vendeur
            FROM visites_rapports
            WHERE date_visite IS NOT NULL AND date_visite != ''
            ORDER BY tournee, date_visite ASC
        """)
        rows = cursor.fetchall()
        
        # 2. Fetch Secteur -> Tournee mapping from localites JOIN secteurs
        cursor.execute("""
            SELECT s.name AS role_vendeur, l.name AS tournee 
            FROM localites l
            JOIN secteurs s ON l.secteur_id = s.id
            ORDER BY s.name, l.name ASC
        """)
        st_rows = cursor.fetchall()
        
        secteur_to_tournees = {}
        tournee_to_secteur = {}
        for r in st_rows:
            sec = r["role_vendeur"]
            tn = r["tournee"]
            if sec not in secteur_to_tournees:
                secteur_to_tournees[sec] = []
            secteur_to_tournees[sec].append(tn)
            tournee_to_secteur[tn] = sec

        cursor.execute("SELECT DISTINCT date_visite FROM visites_rapports WHERE date_visite IS NOT NULL AND date_visite != '' ORDER BY date_visite ASC")
        all_dates = [r["date_visite"] for r in cursor.fetchall() if r["date_visite"]]
        
        cursor.execute("SELECT DISTINCT vendeur, cdz, secteur FROM fdv WHERE vendeur IS NOT NULL AND vendeur != '' ORDER BY vendeur ASC")
        fdv_vendeurs_rows = cursor.fetchall()
        conn.close()
        
        tournees_map = {}
        vendeurs_map = {}
        secteurs_map = {}

        # Pre-populate vendeurs from FDV database table
        for fv in fdv_vendeurs_rows:
            v_name = fv["vendeur"]
            vendeurs_map[v_name] = {
                "dates": set(),
                "dates_details": [],
                "cdz": fv["cdz"] or "",
                "secteur": fv["secteur"] or ""
            }

        # Initialize secteurs_map with tournees
        for sec, t_list in secteur_to_tournees.items():
            secteurs_map[sec] = {
                "tournees": t_list,
                "dates": set(),
                "dates_details": []
            }
        
        for row in rows:
            t_name = row["tournee"]
            d_val = row["date_visite"]
            v_name = row["vendeur"]
            
            if t_name and t_name != "N/A":
                if t_name not in tournees_map:
                    tournees_map[t_name] = {"dates": set(), "dates_details": []}
                if d_val not in tournees_map[t_name]["dates"]:
                    tournees_map[t_name]["dates"].add(d_val)
                    tournees_map[t_name]["dates_details"].append({"date": d_val, "tournee": t_name, "vendeur": v_name or ""})
                    
            if v_name and v_name != "N/A":
                if v_name not in vendeurs_map:
                    vendeurs_map[v_name] = {"dates": set(), "dates_details": [], "cdz": "", "secteur": ""}
                if d_val not in vendeurs_map[v_name]["dates"]:
                    vendeurs_map[v_name]["dates"].add(d_val)
                    vendeurs_map[v_name]["dates_details"].append({"date": d_val, "tournee": t_name or "", "vendeur": v_name})

            sec = tournee_to_secteur.get(t_name)
            if sec and sec in secteurs_map:
                secteurs_map[sec]["dates"].add(d_val)
                existing_d = [dt["date"] for dt in secteurs_map[sec]["dates_details"] if dt["tournee"] == t_name]
                if d_val not in existing_d:
                    secteurs_map[sec]["dates_details"].append({"date": d_val, "tournee": t_name, "vendeur": v_name or ""})

        # Add ACM tournées if missing
        for t in tournee_to_secteur.keys():
            if t not in tournees_map:
                tournees_map[t] = {"dates": set(), "dates_details": []}

        # Build final lists
        tournees_list = []
        for t, info in sorted(tournees_map.items()):
            dates_sorted = sorted(list(info["dates"]))
            tournees_list.append({"name": t, "dates": dates_sorted, "dates_details": sorted(info["dates_details"], key=lambda x: x["date"])})

        vendeurs_list = []
        for v, info in sorted(vendeurs_map.items()):
            dates_sorted = sorted(list(info["dates"]))
            
            # Build tournées details for this seller
            v_tournees = {}
            for dt in info.get("dates_details", []):
                tn = dt.get("tournee")
                d_date = dt.get("date")
                if tn:
                    if tn not in v_tournees:
                        v_tournees[tn] = set()
                    if d_date:
                        v_tournees[tn].add(d_date)
                        
            v_tournees_objs = []
            for tn, t_dates_set in v_tournees.items():
                t_dates = sorted(list(t_dates_set))
                v_tournees_objs.append({
                    "name": tn,
                    "dates": t_dates,
                    "min_date": t_dates[0] if t_dates else "9999-12-31",
                    "max_date": t_dates[-1] if t_dates else "9999-12-31"
                })
            v_tournees_objs.sort(key=lambda x: (x["min_date"], x["name"]))

            vendeurs_list.append({
                "name": v,
                "cdz": info.get("cdz", ""),
                "secteur": info.get("secteur", ""),
                "tournees": [x["name"] for x in v_tournees_objs],
                "tournees_details": v_tournees_objs,
                "dates": dates_sorted,
                "dates_details": sorted(info["dates_details"], key=lambda x: x["date"])
            })

        secteurs_list = []
        for s, info in sorted(secteurs_map.items()):
            dates_sorted = sorted(list(info["dates"]))
            if not dates_sorted:
                dates_sorted = list(all_dates)

            # Build date-sorted tournées for this secteur
            sec_tournees_objs = []
            for t in info["tournees"]:
                t_dates = sorted(list(tournees_map[t]["dates"])) if (t in tournees_map and tournees_map[t].get("dates")) else []
                m_date = t_dates[0] if t_dates else "9999-12-31"
                mx_date = t_dates[-1] if t_dates else "9999-12-31"
                sec_tournees_objs.append({
                    "name": t,
                    "dates": t_dates,
                    "min_date": m_date,
                    "max_date": mx_date
                })
            
            # Sort tournees chronologically by min_date, then by name
            sec_tournees_objs.sort(key=lambda x: (x["min_date"], x["name"]))
            sorted_tournees = [x["name"] for x in sec_tournees_objs]

            secteurs_list.append({
                "name": s,
                "tournees": sorted_tournees,
                "tournees_details": sec_tournees_objs,
                "dates": dates_sorted,
                "dates_details": sorted(info["dates_details"], key=lambda x: x["date"])
            })

        return jsonify({
            "status": "success",
            "tournees": tournees_list,
            "vendeurs": vendeurs_list,
            "secteurs": secteurs_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/clients/compare_visites_db", methods=["POST"])
def compare_visites_db_endpoint():
    try:
        data = request.json or {}
        tournee = data.get("tournee")
        vendeur = data.get("vendeur")
        secteur = data.get("secteur")
        client_status = data.get("client_status", "all") # 'all', 'actif', 'inactif'
        dates = sorted(data.get("dates", []))
        
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()

        if tournee and (not dates or len(dates) < 1):
            cursor.execute("SELECT DISTINCT date_visite FROM visites_rapports WHERE tournee = ? AND date_visite IS NOT NULL AND date_visite != '' ORDER BY date_visite ASC", (tournee,))
            dates = [r["date_visite"] for r in cursor.fetchall() if r["date_visite"]]
            if not dates:
                cursor.execute("SELECT DISTINCT date_visite FROM visites_rapports WHERE date_visite IS NOT NULL AND date_visite != '' ORDER BY date_visite ASC LIMIT 3")
                dates = [r["date_visite"] for r in cursor.fetchall() if r["date_visite"]]

        if not dates or len(dates) < 1:
            conn.close()
            return jsonify({"status": "error", "message": "Veuillez sélectionner au moins 1 date à analyser."}), 400
            
        if not tournee and len(dates) > 3:
            conn.close()
            return jsonify({"status": "error", "message": "Vous pouvez comparer jusqu'à 3 dates au maximum."}), 400
        
        dates_data = []
        dates_details = []
        for d in dates:
            if tournee:
                cursor.execute("""
                    SELECT client_code, client_nom, motif, note, vendeur, tournee
                    FROM visites_rapports
                    WHERE tournee = ? AND date_visite = ?
                """, (tournee, d))
            elif secteur:
                cursor.execute("""
                    SELECT client_code, client_nom, motif, note, vendeur, tournee
                    FROM visites_rapports
                    WHERE (vendeur = ? OR tournee IN (SELECT DISTINCT l.name FROM localites l JOIN secteurs s ON l.secteur_id = s.id WHERE s.name = ?)) AND date_visite = ?
                """, (secteur, secteur, d))
            else:
                cursor.execute("""
                    SELECT client_code, client_nom, motif, note, vendeur, tournee
                    FROM visites_rapports
                    WHERE vendeur = ? AND date_visite = ?
                """, (vendeur, d))
            rows = cursor.fetchall()
            t_found = (rows[0]["tournee"] if rows and rows[0]["tournee"] else (tournee or ""))
            v_found = (rows[0]["vendeur"] if rows and rows[0]["vendeur"] else (vendeur or secteur or ""))
            dates_details.append({"date": d, "tournee": t_found, "vendeur": v_found})
            dates_data.append((d, rows))
            
        conn.close()
        
        # Allow dates with 0 visited rows so ACM clients (active/inactive) can still be analyzed
                
        all_clients = {}
        for idx, (d, rows) in enumerate(dates_data):
            for row in rows:
                code = row["client_code"]
                name = row["client_nom"] or "N/A"
                motif = row["motif"] or "Non visité"
                
                if code not in all_clients:
                    all_clients[code] = {
                        "code": code,
                        "name": name,
                        "motifs": ["Non visité"] * len(dates),
                        "is_inactif": 0,
                        "is_actif": 0
                    }
                curr_m = all_clients[code]["motifs"][idx]
                if curr_m.upper() == "OK":
                    pass
                elif motif.upper() == "OK":
                    all_clients[code]["motifs"][idx] = "OK"
                elif curr_m == "Non visité" or not curr_m:
                    all_clients[code]["motifs"][idx] = motif
                else:
                    all_clients[code]["motifs"][idx] = motif
                
        # Merge clients from clients table
        acm_clients_map = {}
        conn = db_manager.get_db_connection()
        c_cursor = conn.cursor()
        
        base_client_sql = """
            SELECT c.code, c.name, 0 AS is_actif, 0 AS is_inactif, s.name AS role_vendeur, l.name AS tournee 
            FROM clients c
            JOIN secteurs s ON c.secteur_id = s.id
            JOIN localites l ON c.localite_id = l.id
        """

        if tournee:
            c_cursor.execute(f"{base_client_sql} WHERE l.name = ?", (tournee,))
        elif secteur:
            if dates:
                # Find active tournées executed on the selected dates in this secteur
                placeholders = ','.join(['?'] * len(dates))
                c_cursor.execute(f"""
                    SELECT DISTINCT tournee 
                    FROM visites_rapports 
                    WHERE (vendeur = ? OR tournee IN (SELECT DISTINCT l.name FROM localites l JOIN secteurs s ON l.secteur_id = s.id WHERE s.name = ?))
                    AND date_visite IN ({placeholders})
                    AND tournee IS NOT NULL AND tournee != ''
                """, [secteur, secteur] + dates)
                active_tournees = [r[0] for r in c_cursor.fetchall() if r[0]]
                
                if active_tournees:
                    tn_placeholders = ','.join(['?'] * len(active_tournees))
                    c_cursor.execute(f"""
                        {base_client_sql}
                        WHERE s.name = ? AND l.name IN ({tn_placeholders})
                    """, [secteur] + active_tournees)
                else:
                    c_cursor.execute(f"{base_client_sql} WHERE s.name = ?", (secteur,))
            else:
                c_cursor.execute(f"{base_client_sql} WHERE s.name = ?", (secteur,))
        elif vendeur:
            v_code = vendeur.split()[0] if vendeur else ""
            c_cursor.execute(f"{base_client_sql} WHERE s.name = ? OR s.name LIKE ?", (vendeur, f"%{v_code}%"))
            
        for r in c_cursor.fetchall():
            acm_clients_map[r["code"]] = dict(r)
        conn.close()

        for code, acm_info in acm_clients_map.items():
            if client_status == "actif" and acm_info.get("is_actif") != 1:
                continue
            if client_status == "inactif" and acm_info.get("is_inactif") != 1:
                continue

            if code not in all_clients:
                all_clients[code] = {
                    "code": code,
                    "name": acm_info["name"],
                    "motifs": ["Non visité"] * len(dates),
                    "is_inactif": acm_info.get("is_inactif", 0),
                    "is_actif": acm_info.get("is_actif", 0)
                }
            else:
                all_clients[code]["is_inactif"] = acm_info.get("is_inactif", 0)
                all_clients[code]["is_actif"] = acm_info.get("is_actif", 0)

        # Strict Tournée Filtering:
        # If a specific tournée is selected, restrict all_clients ONLY to clients belonging to that tournée in clients
        if tournee:
            conn = db_manager.get_db_connection()
            acm_check_cursor = conn.cursor()
            acm_check_cursor.execute("""
                SELECT c.code 
                FROM clients c
                JOIN localites l ON c.localite_id = l.id
                WHERE l.name = ?
            """, (tournee,))
            tournee_acm_codes = set(r["code"] for r in acm_check_cursor.fetchall() if r["code"])
            conn.close()

            if tournee_acm_codes:
                all_clients = {code: c_data for code, c_data in all_clients.items() if code in tournee_acm_codes}

        # Filter out if client_status applies to existing visited clients
        if client_status in ("actif", "inactif"):
            filtered_ac = {}
            for code, c in all_clients.items():
                if client_status == "actif" and c.get("is_actif") == 1:
                    filtered_ac[code] = c
                elif client_status == "inactif" and c.get("is_inactif") == 1:
                    filtered_ac[code] = c
            all_clients = filtered_ac

        client_codes = list(all_clients.keys())
        localites_map = {}
        if client_codes:
            conn = db_manager.get_db_connection()
            c_cursor = conn.cursor()
            placeholders = ",".join(["?" for _ in client_codes])
            c_cursor.execute(f"""
                SELECT c.code, l.name AS localite 
                FROM clients c
                JOIN localites l ON c.localite_id = l.id
                WHERE c.code IN ({placeholders})
            """, client_codes)
            for r in c_cursor.fetchall():
                if r["code"]:
                    localites_map[r["code"]] = r["localite"] or ""
            conn.close()

        comparison_rows = []
        summary_stats = {
            "always_ok": 0,
            "billing_loss": 0,
            "billing_gain": 0,
            "never_ok": 0,
            "inconsistent": 0,
            "at_least_one": 0,
            "no_facture": 0,
            "total_inactive": 0
        }
        
        for code, c in all_clients.items():
            motifs = c["motifs"]
            is_ok = [m.upper() == "OK" for m in motifs]
            facture_cnt = sum(1 for ok in is_ok if ok)
            has_fact = (facture_cnt >= 1)
            localite = localites_map.get(code, "")
            
            if c.get("is_inactif") == 1:
                summary_stats["total_inactive"] += 1

            if has_fact:
                summary_stats["at_least_one"] += 1
            else:
                summary_stats["no_facture"] += 1

            if all(is_ok):
                synth = "Toujours Facturé"
                summary_stats["always_ok"] += 1
            elif not any(is_ok):
                synth = "Jamais Facturé"
                summary_stats["never_ok"] += 1
            elif is_ok[0] and not is_ok[-1]:
                synth = "Perte de facturation"
                summary_stats["billing_loss"] += 1
            elif not is_ok[0] and is_ok[-1]:
                synth = "Gagné (Facturé en fin)"
                summary_stats["billing_gain"] += 1
            else:
                synth = "Inconstant"
                summary_stats["inconsistent"] += 1
                
            comparison_rows.append({
                "code": c["code"],
                "name": c["name"],
                "localite": localite,
                "motifs": motifs,
                "synthesis": synth,
                "facture_count": facture_cnt,
                "has_facture": has_fact,
                "is_inactif": c.get("is_inactif", 0),
                "is_actif": c.get("is_actif", 0)
            })
            
        vendeur_name = vendeur or secteur or tournee or "N/A"
        for _, rows in dates_data:
            if len(rows) > 0 and rows[0]["vendeur"]:
                vendeur_name = rows[0]["vendeur"]
                break
                
        v_phone = db_manager.get_vendeur_phone_from_fdv(vendeur_name) or ""
        v_phone_raw = v_phone

        # Retrieve ACM RAF (raf_acm) for this seller from qualitative_data
        raf_acm = 20
        if vendeur_name and vendeur_name != "N/A":
            try:
                conn_q = db_manager.get_db_connection()
                cur_q = conn_q.cursor()
                v_code = vendeur_name.split()[0]
                cur_q.execute("SELECT raf_acm FROM qualitative_data WHERE vendeur LIKE ? ORDER BY date DESC LIMIT 1", (f"%{v_code}%",))
                q_row = cur_q.fetchone()
                if q_row and q_row["raf_acm"] is not None:
                    raf_acm = int(round(q_row["raf_acm"]))
                conn_q.close()
            except Exception as e:
                pass

        return jsonify({
            "status": "success",
            "vendeur": vendeur_name,
            "vendeur_phone": v_phone,
            "vendeur_phone_raw": v_phone_raw,
            "raf_acm": raf_acm,
            "dates": dates,
            "dates_details": dates_details,
            "summary": summary_stats,
            "comparison": comparison_rows
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/upload", methods=["POST"])
def upload_suivi():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni."}), 400
        file = request.files["file"]
        date = request.form.get("date")
        rest_days = request.form.get("rest_days")
        
        if not date:
            return jsonify({"status": "error", "message": "Aucune date fournie."}), 400
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"status": "error", "message": "Le fichier doit être un classeur Excel (.xlsx, .xls)."}), 400
            
        file_content = file.read()
        db_manager.save_suivi_file(date, file.filename, file_content)
        
        # Process and save structured JSON, always extracting rest days from the file
        process_and_save_suivi(date, file_content, force_extract_rest_days=True)
        
        return jsonify({"status": "success", "message": f"Fichier importé et recalculé avec succès pour le {date}."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/suivi_dates", methods=["GET"])
def get_suivi_dates():
    try:
        dates = db_manager.get_all_suivi_dates()
        return jsonify({"status": "success", "dates": dates})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/reset_db", methods=["POST"])
def reset_db():
    try:
        data = request.get_json(silent=True) or {}
        tables = data.get("tables", [])
        if tables:
            db_manager.reset_specific_tables(tables)
            message = "Les tables spécifiées ont été réinitialisées avec succès."
        else:
            db_manager.reset_all_database_tables()
            message = "La base de données a été réinitialisée avec succès."
        return jsonify({"status": "success", "message": message})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/recreate_db_file", methods=["POST"])
def recreate_db_file():
    try:
        db_manager.delete_and_recreate_db_file()
        return jsonify({"status": "success", "message": "Le fichier database.db a été supprimé et récréé avec succès."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500




@app.route("/api/trends", methods=["GET"])
def get_trends():
    try:
        family = request.args.get("family", "LEVURE").strip().upper()
        category = request.args.get("category", "All").strip()
        records = db_manager.get_all_suivi_data_records()
        
        allowed_vendeurs_set = None
        if category and category != "All":
            allowed_vendeurs = get_categorie(category)
            if not isinstance(allowed_vendeurs, list):
                allowed_vendeurs = [allowed_vendeurs]
            allowed_vendeurs_set = {v.strip().upper() for v in allowed_vendeurs if v}
        
        dates = []
        vendeurs_set = set()
        trends = {}
        quali_trends = {}
        
        for r in records:
            date_str = r["date"]
            dates.append(date_str)
            quanti = r["data"].get("quantitative", [])
            quanti = [dict(i) if not isinstance(i, dict) else i for i in quanti]
            quali_list = r["data"].get("qualitative", [])
            quali_list = [dict(i) if not isinstance(i, dict) else i for i in quali_list]
            
            # Extract records for the selected family (handle C.A (TTC) and C.A (HT) interchangeably)
            if family in ("C.A (TTC)", "C.A (HT)", "C.A(HT)", "C.A(TTC)"):
                family_records = [item for item in quanti if item["famille"].strip().upper() in ("C.A (HT)", "C.A (TTC)", "C.A(HT)", "C.A(TTC)")]
            else:
                family_records = [item for item in quanti if item["famille"].strip().upper() == family]
            
            for item in family_records:
                v = item["vendeur"].strip()
                if not v or v.upper() == 'AUTRE':
                    continue
                if allowed_vendeurs_set is not None and v.upper() not in allowed_vendeurs_set:
                    continue
                vendeurs_set.add(v)
                
                if v not in trends:
                    trends[v] = {}
                
                pct = round((item["real"] / item["obj"]) * 100) if item["obj"] > 0 else 0
                trends[v][date_str] = {
                    "real": item["real"],
                    "obj": item["obj"],
                    "pct": pct,
                    "encours": item.get("encours", 0)
                }
                
            for item in quali_list:
                v = item["vendeur"].strip()
                if not v or v.upper() == 'AUTRE':
                    continue
                if allowed_vendeurs_set is not None and v.upper() not in allowed_vendeurs_set:
                    continue
                
                if v not in quali_trends:
                    quali_trends[v] = {}
                    
                quali_trends[v][date_str] = {
                    "clt_programme": item.get("clt_programme", 0),
                    "clt_facture": item.get("clt_facture", 0),
                    "acm": round(item.get("acm", 0.0) * 100, 1),
                    "tsm": round(item.get("tsm", 0.0) * 100, 1),
                    "line": round(item.get("line", 0.0) * 100, 1) if item.get("line") is not None else 0.0,
                    "raf_tsm": item.get("raf_tsm", 0),
                    "raf_acm": item.get("raf_acm", 0)
                }
                
        # Fill missing dates with 0s for each vendeur
        sorted_dates = sorted(list(set(dates)))
        formatted_trends = {}
        formatted_quali = {}
        
        for v in vendeurs_set:
            v_data = []
            v_quali = []
            for d in sorted_dates:
                # Quantitative
                if d in trends.get(v, {}):
                    v_data.append({
                        "date": d,
                        "real": trends[v][d]["real"],
                        "obj": trends[v][d]["obj"],
                        "pct": trends[v][d]["pct"],
                        "encours": trends[v][d]["encours"]
                    })
                else:
                    v_data.append({
                        "date": d,
                        "real": 0,
                        "obj": 0,
                        "pct": 0,
                        "encours": 0
                    })
                
                # Qualitative
                if v in quali_trends and d in quali_trends[v]:
                    v_quali.append({
                        "date": d,
                        "clt_programme": quali_trends[v][d]["clt_programme"],
                        "clt_facture": quali_trends[v][d]["clt_facture"],
                        "acm": quali_trends[v][d]["acm"],
                        "tsm": quali_trends[v][d]["tsm"],
                        "line": quali_trends[v][d]["line"],
                        "raf_tsm": quali_trends[v][d]["raf_tsm"],
                        "raf_acm": quali_trends[v][d]["raf_acm"]
                    })
                else:
                    v_quali.append({
                        "date": d,
                        "clt_programme": 0,
                        "clt_facture": 0,
                        "acm": 0.0,
                        "tsm": 0.0,
                        "line": 0.0,
                        "raf_tsm": 0,
                        "raf_acm": 0
                    })
            formatted_trends[v] = v_data
            formatted_quali[v] = v_quali
            
        # Get all distinct families for the dropdown (excluding empty)
        all_families = set()
        for r in records:
            quanti = r["data"].get("quantitative", [])
            for item in quanti:
                fam = item["famille"].strip()
                if fam:
                    all_families.add(fam)
                    
        return jsonify({
            "status": "success",
            "dates": sorted_dates,
            "vendeurs": sorted(list(vendeurs_set)),
            "families": sorted(list(all_families)),
            "trends": formatted_trends,
            "qualitative_trends": formatted_quali
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
@app.route("/api/reload-check", methods=["GET"])
def reload_check():
    import os
    paths = ["templates", "static"]
    max_mtime = 0
    for path in paths:
        if os.path.exists(path):
            for root, dirs, files in os.walk(path):
                for file in files:
                    fp = os.path.join(root, file)
                    try:
                        mtime = os.path.getmtime(fp)
                        if mtime > max_mtime:
                            max_mtime = mtime
                    except OSError:
                        pass
    return jsonify({"last_modified": max_mtime})

@app.route("/api/config", methods=["GET"])
def get_app_config():
    try:
        config = load_config()
        return jsonify({"status": "success", "config": config})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/config", methods=["POST"])
def post_app_config():
    try:
        req_data = request.get_json() or {}
        config = load_config()
        
        # Update keys if present
        if "rest_days" in req_data:
            config["rest_days"] = int(req_data["rest_days"])
        if "exclude_families" in req_data:
            config["exclude_families"] = req_data["exclude_families"]
        if "theme" in req_data:
            config["theme"] = req_data["theme"]
        if "light_mode" in req_data:
            config["light_mode"] = bool(req_data["light_mode"])
        if "excluded_dates" in req_data:
            config["excluded_dates"] = req_data["excluded_dates"]
            
        save_config(config)
        return jsonify({"status": "success", "config": config})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ------------------------------------------------------------------
# Stock API endpoint
# ------------------------------------------------------------------
@app.route("/api/stock", methods=["GET"])
def get_stock_data():
    try:
        search = request.args.get("search", "").strip() or None
        sites_raw = request.args.get("sites", "")
        sites = [s.strip() for s in sites_raw.split(",") if s.strip()] if sites_raw else None
        
        socs_raw = request.args.get("socs", "")
        socs = [s.strip() for s in socs_raw.split(",") if s.strip()] if socs_raw else None
        
        fournisseurs_raw = request.args.get("fournisseurs", "")
        fournisseurs = [f.strip() for f in fournisseurs_raw.split(",") if f.strip()] if fournisseurs_raw else None
        
        sort_by = request.args.get("sort_by", "Produit").strip()
        sort_dir = request.args.get("sort_dir", "ASC").strip().upper()
        
        date = request.args.get("date", "").strip()
        if date == "default" or date == "null" or date == "undefined":
            date = ""
            
        # Get data from database
        result = db_manager.get_stock_data_from_db(
            date=date or None,
            search=search,
            sites=sites,
            socs=socs,
            fournisseurs=fournisseurs,
            sort_by=sort_by,
            sort_dir=sort_dir
        )
        
        # Get all distinct stock dates
        all_dates = db_manager.get_stock_dates()
        
        # Get filter choices specific to the selected date
        selected_date = result.get("date")
        if selected_date:
            filters = db_manager.get_stock_filters_from_db(selected_date)
        else:
            filters = {"sites": [], "socs": [], "fournisseurs": []}
            
        return jsonify({
            "status": "success",
            "rows": result["rows"],
            "summary": result["summary"],
            "filters": filters,
            "dates": all_dates,
            "selected_date": selected_date
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/stock/favorites", methods=["GET"])
def get_stock_favorites_api():
    try:
        favs = db_manager.get_stock_favorites()
        return jsonify({"status": "success", "favorites": favs})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/stock/favorites/toggle", methods=["POST"])
def toggle_stock_favorites_api():
    try:
        payload = request.json or {}
        produit = payload.get("produit")
        action = payload.get("action") # "add" or "remove"
        if not produit:
            return jsonify({"status": "error", "message": "Code produit requis."}), 400
        
        if action == "add":
            db_manager.add_stock_favorite(produit)
        elif action == "remove":
            db_manager.remove_stock_favorite(produit)
        else:
            return jsonify({"status": "error", "message": "Action non spécifiée ou incorrecte (doit être 'add' ou 'remove')."}), 400
            
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/stock/upload", methods=["POST"])
def upload_stock_file():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni."}), 400
        file = request.files["file"]
        date = request.form.get("date")
        
        if not date:
            return jsonify({"status": "error", "message": "Aucune date fournie."}), 400
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"status": "error", "message": "Le fichier doit être un classeur Excel (.xlsx, .xls)."}), 400
            
        # Read Excel using pandas
        df = pd.read_excel(file)
        df.columns = [c.strip() for c in df.columns]
        df = df.fillna("")
        
        # Clean text values
        def clean_val(val):
            if not isinstance(val, str):
                return val
            # Specific corrections for encoding
            val = val.replace('PLATEAU  TH EN INOX_HPC', 'PLATEAU À THÉ EN INOX_HPC')
            val = val.replace('PRSENTOIR', 'PRÉSENTOIR')
            val = val.replace('\ufffd', 'É')
            return val.strip()
            
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(clean_val)
                
        rows = df.to_dict(orient="records")
        
        # Save to database
        saved_count = db_manager.save_stock_data(date, rows)
        
        return jsonify({
            "status": "success",
            "message": f"Fichier stock importé avec succès. {saved_count} articles enregistrés pour le {date}."
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


def send_telegram_notification(text, chat_id=None):
    """Sends a message via Telegram bot."""
    import requests
    import os
    import json
    token = "8932059052:AAEbwgRvpDlofG49OxY-9TWVdwT7MfaWdJk"
    
    # Try to load chat_id from environment or cache file if not provided
    if not chat_id:
        try:
            from dotenv import load_dotenv
            load_dotenv()
        except Exception:
            pass
            
        chat_id = os.environ.get("TELEGRAM_CHAT_ID")
        cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "telegram_chat_cache.json")
        
        if not chat_id and os.path.exists(cache_path):
            try:
                with open(cache_path, "r") as f:
                    chat_id = json.load(f).get("chat_id")
            except Exception:
                pass
            
    # Fetch from getUpdates if not cached
    if not chat_id:
        try:
            res = requests.get(f"https://api.telegram.org/bot{token}/getUpdates", timeout=5).json()
            results = res.get("result", [])
            if font_scopes := results:  # checking if results is not empty
                for r in reversed(results):
                    cid = r.get("message", {}).get("chat", {}).get("id") or r.get("my_chat_member", {}).get("chat", {}).get("id")
                    if cid:
                        chat_id = cid
                        try:
                            with open(cache_path, "w") as f:
                                json.dump({"chat_id": chat_id}, f)
                        except Exception:
                            pass
                        break
        except Exception:
            pass
            
    if chat_id:
        try:
            url = f"https://api.telegram.org/bot{token}/sendMessage"
            payload = {
                "chat_id": chat_id,
                "text": text,
                "parse_mode": "Markdown"
            }
            requests.post(url, json=payload, timeout=5)
            return True
        except Exception:
            pass
    return False


def sync_to_google_sheet(date):
    import os
    import google.oauth2.credentials
    from googleapiclient.discovery import build
    import sqlite3
    
    SPREADSHEET_ID = "17Q3DoTjLdGwAmztk3LWaC2Z69_ZrGYlAsLHXTusrHIk"
    SHEET_NAME = "STock Speed-X3"
    
    if not os.path.exists('token.json'):
        return False, "Aucun jeton d'autorisation Google trouvé. Veuillez vous authentifier."
        
    try:
        creds = google.oauth2.credentials.Credentials.from_authorized_user_file('token.json', ['https://www.googleapis.com/auth/spreadsheets'])
        
        if creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request
            creds.refresh(Request())
            with open('token.json', 'w') as token_file:
                token_file.write(creds.to_json())
                
        service = build('sheets', 'v4', credentials=creds)
        
        # Check sheet tab existence
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        sheets = [s['properties']['title'] for s in spreadsheet.get('sheets', [])]
        if SHEET_NAME not in sheets:
            body = {
                'requests': [
                    {
                        'addSheet': {
                            'properties': {
                                'title': SHEET_NAME
                            }
                        }
                    }
                ]
            }
            service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
            
        # Get all rows for date from database
        conn = db_manager.get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT act_code, site, soc, fournisseur, gamme, famille, produit, designation, statut, stk_qte, source FROM stock WHERE date = ? AND act_code = 'AG_AGDR'", (date,))
            db_rows = cursor.fetchall()
        finally:
            conn.close()
        
        # Build payload
        headers = ['ACT CODE', 'Site', 'SOC', 'Fournisseur', 'GAMME', 'FAMILLE', 'Produit', 'DESIGNATION', 'Statut', 'STK QTE', 'Source']
        values = [headers]
        for row in db_rows:
            values.append([
                row[0], # act_code
                row[1], # site
                row[2], # soc
                row[3], # fournisseur
                row[4], # gamme
                row[5], # famille
                row[6], # produit
                row[7], # designation
                row[8], # statut
                int(row[9] or 0), # stk_qte
                row[10] # source
            ])
            
        # Clear existing range
        service.spreadsheets().values().clear(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{SHEET_NAME}'!A1:Z20000",
            body={}
        ).execute()
        
        # Update
        body = {'values': values}
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{SHEET_NAME}'!A1",
            valueInputOption='RAW',
            body=body
        ).execute()
        
        msg = f"Stock synchronisé avec succès vers Google Sheet '{SHEET_NAME}' ({len(db_rows)} articles)."
        send_telegram_notification(f"📦 *Stock Sync Successful*\n\n{msg}")
        
        return True, msg
    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, f"Erreur Google Sheets : {str(e)}"


@app.route("/api/google/authorize")
def google_authorize():
    import json
    import urllib.parse
    with open('google.json', 'r') as f:
        secrets = json.load(f)['web']
        
    params = {
        'client_id': secrets['client_id'],
        'redirect_uri': 'http://127.0.0.1:5000/',
        'response_type': 'code',
        'scope': 'https://www.googleapis.com/auth/spreadsheets https://www.googleapis.com/auth/gmail.modify',
        'access_type': 'offline',
        'prompt': 'consent'
    }
    authorization_url = secrets['auth_uri'] + '?' + urllib.parse.urlencode(params)
    return redirect(authorization_url)


@app.route("/api/google/sync", methods=["POST"])
def google_sync_stock():
    try:
        data = request.get_json() or {}
        date = data.get("date")
        
        if not date:
            dates = db_manager.get_stock_dates()
            if not dates:
                return jsonify({"status": "error", "message": "Aucune date de stock trouvée dans la base de données."}), 400
            date = dates[0]
            
        if not os.path.exists('token.json'):
            return jsonify({
                "status": "auth_required", 
                "auth_url": "/api/google/authorize",
                "message": "Authentification Google Sheets requise."
            })
            
        success, msg = sync_to_google_sheet(date)
        if success:
            return jsonify({"status": "success", "message": msg})
        else:
            return jsonify({"status": "error", "message": msg})
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/google/gmail-webhook", methods=["POST"])
def google_gmail_webhook():
    import io
    import base64
    import os
    import json
    from googleapiclient.discovery import build
    
    # 1. Resolve spreadsheet config
    SPREADSHEET_ID = "17Q3DoTjLdGwAmztk3LWaC2Z69_ZrGYlAsLHXTusrHIk"
    SHEET_NAME = "STock Speed-X3"
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/gmail.modify"
    ]
    
    # Helper to load credentials
    def _load_app_creds():
        if not os.path.exists('token.json'):
            return None
        try:
            with open('token.json', 'r') as f:
                token_data = json.load(f)
            token_scopes = token_data.get("scopes", SCOPES)
            import google.oauth2.credentials
            creds = google.oauth2.credentials.Credentials.from_authorized_user_file('token.json', token_scopes)
            if creds.expired and creds.refresh_token:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                refreshed_data = json.loads(creds.to_json())
                if "scopes" not in refreshed_data:
                    refreshed_data["scopes"] = token_scopes
                with open('token.json', 'w') as token_file:
                    json.dump(refreshed_data, token_file)
            return creds
        except Exception:
            return None

    try:
        # Check if file is uploaded in the request (multipart/form-data)
        file_stream = None
        filename = "Fichier uploadé via webhook"
        
        if 'file' in request.files:
            file_stream = request.files['file'].stream
            filename = request.files['file'].filename
        elif request.data and not request.is_json:
            # Maybe raw bytes sent in body
            file_stream = io.BytesIO(request.data)
            filename = "Raw bytes payload"
            
        # If no file was sent, retrieve it from Gmail using the query
        if not file_stream:
            creds = _load_app_creds()
            if not creds:
                return jsonify({"status": "error", "message": "Token Google introuvable. Veuillez vous authentifier."}), 400
                
            token_scopes = getattr(creds, "scopes", [])
            if "https://www.googleapis.com/auth/gmail.modify" not in token_scopes:
                return jsonify({"status": "error", "message": "Permissions Gmail manquantes. Re-authentifiez-vous d'abord."}), 400
                
            gmail_service = build("gmail", "v1", credentials=creds)
            query = 'has:attachment (filename:xlsx OR filename:xls) "STock Speed-X3"'
            results = gmail_service.users().messages().list(userId="me", q=query, maxResults=5).execute()
            messages = results.get("messages", [])
            
            if not messages:
                return jsonify({"status": "error", "message": "Aucun e-mail correspondant trouvé dans Gmail."}), 404
                
            webhook_msg_id = None
            for msg_summary in messages:
                msg_id = msg_summary["id"]
                webhook_msg_id = msg_id
                message = gmail_service.users().messages().get(userId="me", id=msg_id).execute()
                
                payload = message.get("payload", {})
                parts = [payload]
                attachments = []
                while parts:
                    part = parts.pop()
                    if part.get("parts"):
                        parts.extend(part["parts"])
                    fn = part.get("filename", "")
                    if fn and (fn.lower().endswith(".xlsx") or fn.lower().endswith(".xls")):
                        att_id = part["body"].get("attachmentId")
                        if att_id:
                            attachments.append((fn, att_id))
                            
                if attachments:
                    filename, att_id = attachments[0]
                    attachment = gmail_service.users().messages().attachments().get(
                        userId="me", messageId=msg_id, id=att_id
                    ).execute()
                    file_data = base64.urlsafe_b64decode(attachment["data"].encode("UTF-8"))
                    file_stream = io.BytesIO(file_data)
                    break
                    
            if not file_stream:
                return jsonify({"status": "error", "message": "Aucune pièce jointe valide trouvée dans les derniers messages Gmail."}), 404
                
        # 2. Parse and filter the Excel file using Pandas
        try:
            df = pd.read_excel(file_stream)
        except Exception as e:
            return jsonify({"status": "error", "message": f"Erreur de lecture du fichier Excel : {e}"}), 400
            
        df.columns = [c.strip() for c in df.columns]
        df = df.fillna("")
        
        act_col = next((c for c in df.columns if c.upper() == "ACT CODE"), None)
        if act_col:
            df = df[df[act_col].astype(str).str.strip() == "AG_AGDR"].reset_index(drop=True)
            
        if df.empty:
            return jsonify({"status": "warning", "message": "Aucune ligne avec ACT CODE = AG_AGDR trouvée."}), 200
            
        # 3. Push to Google Sheets
        creds = _load_app_creds()
        if not creds:
            return jsonify({"status": "error", "message": "Token Google introuvable."}), 400
            
        sheets_service = build("sheets", "v4", credentials=creds)
        
        # Ensure sheet tab exists
        meta = sheets_service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        existing = [s["properties"]["title"] for s in meta.get("sheets", [])]
        if SHEET_NAME not in existing:
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=SPREADSHEET_ID,
                body={"requests": [{"addSheet": {"properties": {"title": SHEET_NAME}}}]},
            ).execute()
            
        headers = ["ACT CODE", "Site", "SOC", "Fournisseur", "GAMME",
                   "FAMILLE", "Produit", "DESIGNATION", "Statut", "STK QTE", "Source"]
                   
        col_map = {}
        for h in headers:
            for c in df.columns:
                if c.strip().upper() == h.upper():
                    col_map[h] = c
                    break
                    
        def safe_int(v):
            try: return int(v)
            except: return 0
            
        rows = [headers]
        for _, row in df.iterrows():
            rows.append([
                str(row.get(col_map.get("ACT CODE",  ""), "")).strip(),
                str(row.get(col_map.get("Site",       ""), "")).strip(),
                str(row.get(col_map.get("SOC",        ""), "")).strip(),
                str(row.get(col_map.get("Fournisseur",""), "")).strip(),
                str(row.get(col_map.get("GAMME",      ""), "")).strip(),
                str(row.get(col_map.get("FAMILLE",    ""), "")).strip(),
                str(row.get(col_map.get("Produit",    ""), "")).strip(),
                str(row.get(col_map.get("DESIGNATION",""), "")).strip(),
                str(row.get(col_map.get("Statut",     ""), "")).strip(),
                safe_int(row.get(col_map.get("STK QTE", ""), 0)),
                str(row.get(col_map.get("Source",     ""), "")).strip(),
            ])
            
        sheets_service.spreadsheets().values().clear(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{SHEET_NAME}'!A1:Z20000",
            body={},
        ).execute()
        
        sheets_service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{SHEET_NAME}'!A1",
            valueInputOption="RAW",
            body={"values": rows},
        ).execute()
        
        # Trash processed Gmail message
        if not 'file' in request.files and webhook_msg_id:
            try:
                gmail_service.users().messages().trash(userId="me", id=webhook_msg_id).execute()
            except Exception as tr_err:
                print(f"Error trashing message {webhook_msg_id}: {tr_err}")
                
        msg = f"Webhook traité avec succès. {len(rows) - 1} articles synchronisés depuis '{filename}' vers '{SHEET_NAME}'."
        send_telegram_notification(f"✉️ *Gmail / Webhook Stock Sync Successful*\n\n{msg}")
        
        return jsonify({
            "status": "success",
            "message": msg
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Erreur interne : {e}"}), 500


# ------------------------------------------------------------------
# Client (full list with duplicates) API endpoints

# ------------------------------------------------------------------

def _parse_csv_list(value):
    """Convert a comma-separated query-string value into a clean list."""
    if value is None:
        return None
    if isinstance(value, list):
        return [v for v in value if v]
    return [v.strip() for v in str(value).split(",") if v.strip()]


@app.route("/api/clients_full", methods=["GET"])
def list_clients_full():
    try:
        search = request.args.get("search", "").strip() or None
        secteurs = _parse_csv_list(request.args.get("secteurs"))
        localites = _parse_csv_list(request.args.get("localites"))
        vendeurs_som = _parse_csv_list(request.args.get("vendeurs_som"))
        vendeurs_vmm = _parse_csv_list(request.args.get("vendeurs_vmm"))
        is_repeat_raw = request.args.get("is_repeat")
        if is_repeat_raw is None or is_repeat_raw == "":
            is_repeat = None
        else:
            is_repeat = is_repeat_raw in ("1", "true", "True", "yes", "oui", "OUI")
        unique_raw = request.args.get("unique")
        unique = unique_raw in ("1", "true", "True", "yes", "oui", "OUI") if unique_raw else False
        sort_by = request.args.get("sort_by", "row_index")
        sort_dir = request.args.get("sort_dir", "ASC")
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 25))

        result = db_manager.get_clients_full(
            search=search,
            secteurs=secteurs,
            localites=localites,
            vendeurs_som=vendeurs_som,
            vendeurs_vmm=vendeurs_vmm,
            is_repeat=is_repeat,
            unique=unique,
            sort_by=sort_by,
            sort_dir=sort_dir,
            page=page,
            per_page=per_page,
        )
        return jsonify({"status": "success", **result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients_full/filters", methods=["GET"])
def clients_full_filters():
    try:
        return jsonify({
            "status": "success",
            "filters": db_manager.get_clients_full_filters(),
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients_full/stats", methods=["GET"])
def clients_full_stats():
    try:
        return jsonify({
            "status": "success",
            "stats": db_manager.get_clients_full_stats(),
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients_full/export", methods=["GET"])
def clients_full_export():
    """Export the (filtered) clients_full list to CSV."""
    try:
        import csv
        import io

        search = request.args.get("search", "").strip() or None
        secteurs = _parse_csv_list(request.args.get("secteurs"))
        localites = _parse_csv_list(request.args.get("localites"))
        vendeurs_som = _parse_csv_list(request.args.get("vendeurs_som"))
        vendeurs_vmm = _parse_csv_list(request.args.get("vendeurs_vmm"))
        is_repeat_raw = request.args.get("is_repeat")
        if is_repeat_raw is None or is_repeat_raw == "":
            is_repeat = None
        else:
            is_repeat = is_repeat_raw in ("1", "true", "True", "yes", "oui", "OUI")

        # Pull every row that matches (server-side paginated internally
        # with a large per_page to keep things simple)
        result = db_manager.get_clients_full(
            search=search,
            secteurs=secteurs,
            localites=localites,
            vendeurs_som=vendeurs_som,
            vendeurs_vmm=vendeurs_vmm,
            is_repeat=is_repeat,
            sort_by="row_index",
            sort_dir="ASC",
            page=1,
            per_page=10000,
        )

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "Code", "Nom", "Secteur", "Localité",
            "Vendeur SOM", "Vendeur VMM",
        ])
        for r in result["rows"]:
            writer.writerow([
                r["code"],
                r["name"],
                r["secteur"],
                r["localite"],
                r["vendeur_som"],
                r["vendeur_vmm"],
            ])

        from flask import Response
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": 'attachment; filename="clients_full.csv"'
            },
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



# ------------------------------------------------------------------
# Client management routes (edit / delete / assign vendeur)
# ------------------------------------------------------------------

@app.route("/api/clients/<int:client_id>", methods=["GET"])
def get_client(client_id):
    try:
        client = db_manager.get_client_by_id(client_id)
        if not client:
            return jsonify({"status": "error", "message": "Client introuvable"}), 404
        return jsonify({"status": "success", "client": client})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/<int:client_id>", methods=["PUT"])
def edit_client(client_id):
    try:
        data = request.get_json(force=True) or {}
        updated = db_manager.update_client(
            client_id,
            name=data.get("name"),
            secteur_id=data.get("secteur_id"),
            localite_id=data.get("localite_id"),
            vendeur_som=data.get("vendeur_som"),
            vendeur_vmm=data.get("vendeur_vmm"),
        )
        if not updated:
            return jsonify({"status": "error", "message": "Client introuvable ou aucune modification"}), 404
        client = db_manager.get_client_by_id(client_id)
        return jsonify({"status": "success", "client": client})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/<int:client_id>", methods=["DELETE"])
def delete_client_route(client_id):
    try:
        deleted = db_manager.delete_client(client_id)
        if not deleted:
            return jsonify({"status": "error", "message": "Client introuvable"}), 404
        return jsonify({"status": "success", "message": "Client supprimé"})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/<int:client_id>/assign-vendeur", methods=["POST"])
def assign_vendeur_to_client_secteur(client_id):
    """Assign a vendeur to ALL clients in the same secteur as the given client."""
    try:
        data = request.get_json(force=True) or {}
        channel = (data.get("channel") or "").lower()   # 'som' or 'vmm'
        vendeur = (data.get("vendeur") or "").strip()

        if channel not in ("som", "vmm"):
            return jsonify({"status": "error", "message": "channel doit être 'som' ou 'vmm'"}), 400
        if not vendeur:
            return jsonify({"status": "error", "message": "vendeur requis"}), 400

        # Get the secteur_id of the target client
        client = db_manager.get_client_by_id(client_id)
        if not client:
            return jsonify({"status": "error", "message": "Client introuvable"}), 404

        secteur_id = client["secteur_id"]
        secteur_name = client["secteur"]
        affected = db_manager.assign_vendeur_to_secteur(secteur_id, channel, vendeur)

        return jsonify({
            "status": "success",
            "affected": affected,
            "secteur": secteur_name,
            "channel": channel,
            "vendeur": vendeur,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/vendeurs-select", methods=["GET"])
def clients_vendeurs_select():
    """Return FDV vendeurs grouped by SOM / VMM for dropdowns."""
    try:
        data = db_manager.get_fdv_vendeurs_for_select()
        return jsonify({"status": "success", **data})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/clients/sync_vendeurs", methods=["POST"])
def sync_clients_vendeurs():
    """Sync SOM and VMM vendors across all secteurs for all clients."""
    try:
        db_manager.sync_clients_vendeurs_from_fdv()
        return jsonify({"status": "success", "message": "Les vendeurs SOM et VMM ont été mis à jour pour tous les secteurs."})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/secteurs", methods=["GET"])
def list_secteurs():
    """Return all secteurs."""
    try:
        conn = db_manager.get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, name FROM secteurs ORDER BY name")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"status": "success", "secteurs": rows})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/localites", methods=["GET"])
def list_localites():
    """Return localites, optionally filtered by secteur_id."""
    try:
        secteur_id = request.args.get("secteur_id")
        conn = db_manager.get_db_connection()
        cur = conn.cursor()
        if secteur_id:
            cur.execute("SELECT id, name, secteur_id FROM localites WHERE secteur_id = ? ORDER BY name", (secteur_id,))
        else:
            cur.execute("SELECT id, name, secteur_id FROM localites ORDER BY name")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"status": "success", "localites": rows})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/vendeur/360", methods=["GET"])
def get_vendeur_360_endpoint():
    try:
        vendeur_name = request.args.get("vendeur", "").strip()
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        
        # 1. Sellers list for dropdown - only CHAKIB ELFIL CDZ
        cursor.execute("""
            SELECT DISTINCT vendeur FROM fdv
            WHERE vendeur IS NOT NULL AND vendeur != ''
            AND UPPER(TRIM(cdz)) = 'CHAKIB ELFIL'
        """)
        all_vendeurs = sorted([r[0] for r in cursor.fetchall() if r[0] and r[0].strip() != "N/A"])
        
        if not vendeur_name and all_vendeurs:
            vendeur_name = all_vendeurs[0]
            
        # 2. FDV Master Info for selected seller
        cursor.execute("SELECT * FROM fdv WHERE vendeur LIKE ?", (f"%{vendeur_name}%",))
        fdv_row = cursor.fetchone()
        vendeur_info = {
            "name": vendeur_name,
            "code": vendeur_name.split(' ')[0] if vendeur_name else "",
            "role": (fdv_row["role"] if fdv_row and "role" in fdv_row.keys() else "") or "VENDEUR",
            "type_role": (fdv_row["type_role"] if fdv_row and "type_role" in fdv_row.keys() else "") or "VENDEUR",
            "cdz": (fdv_row["cdz"] if fdv_row and "cdz" in fdv_row.keys() else "") or "N/A",
            "activite": (fdv_row["activite"] if fdv_row and "activite" in fdv_row.keys() else "") or "GMS",
            "secteur": (fdv_row["secteur"] if fdv_row and "secteur" in fdv_row.keys() else "") or "AGADIR",
            "telephone": (fdv_row["telephone"] if fdv_row and "telephone" in fdv_row.keys() else "") or "",
            "whatsapp": (fdv_row["whatsapp"] if fdv_row and "whatsapp" in fdv_row.keys() else "") or ""
        }
        if not vendeur_info["whatsapp"] and vendeur_info["telephone"]:
            vendeur_info["whatsapp"] = vendeur_info["telephone"]

        # 3. Visites & Tournées Data from visites_rapports
        cursor.execute("""
            SELECT client_code, client_nom, tournee, motif, date_visite, distance, agence
            FROM visites_rapports
            WHERE vendeur LIKE ? OR vendeur = ?
        """, (f"%{vendeur_name}%", vendeur_name))
        visites_rows = cursor.fetchall()
        
        # Get localites map from clients_full
        client_codes = list(set([r["client_code"] for r in visites_rows if r["client_code"] and r["client_code"].strip()]))
        localites_map = {}
        if client_codes:
            placeholders = ','.join(['?'] * len(client_codes))
            cursor.execute(f"SELECT c.code, l.name AS localite FROM clients c JOIN localites l ON c.localite_id = l.id WHERE c.code IN ({placeholders})", client_codes)
            for r in cursor.fetchall():
                if r["code"]:
                    localites_map[r["code"]] = r["localite"] or ""
                    
        # Consolidate per client
        clients_map = {}
        tournees_map = {}
        anomalies_list = []
        
        for r in visites_rows:
            code = r["client_code"]
            name = r["client_nom"] or "N/A"
            t_name = r["tournee"] or "N/A"
            motif = r["motif"] or "Non visité"
            loc = localites_map.get(code) or t_name.replace("VMM", "").replace("SOM", "").strip()
            dist = 0
            try:
                dist = int(str(r["distance"]).split('.')[0])
            except:
                pass
                
            if dist > 100:
                anomalies_list.append({
                    "client_code": code,
                    "client_nom": name,
                    "date": r["date_visite"],
                    "distance": dist,
                    "motif": motif
                })

            if code not in clients_map:
                clients_map[code] = {
                    "code": code,
                    "name": name,
                    "localite": loc,
                    "tournee": t_name,
                    "has_ok": False,
                    "latest_motif": motif,
                    "visite_count": 0
                }
                
            clients_map[code]["visite_count"] += 1
            if str(motif).upper() == "OK":
                clients_map[code]["has_ok"] = True
                clients_map[code]["latest_motif"] = "OK"
            elif not clients_map[code]["has_ok"]:
                clients_map[code]["latest_motif"] = motif

        # Also fetch registered anomalies from anomalies table
        try:
            cursor.execute("""
                SELECT id, date, vendeur, type_anomali, commentaire, tag
                FROM anomalies
                WHERE vendeur LIKE ? OR vendeur = ?
            """, (f"%{vendeur_name}%", vendeur_name))
            for ar in cursor.fetchall():
                anomalies_list.append({
                    "client_code": ar["type_anomali"] or "Anomalie",
                    "client_nom": ar["commentaire"] or ar["tag"] or "Anomalie enregistrée",
                    "date": ar["date"] or "",
                    "distance": 0,
                    "motif": ar["tag"] or ar["type_anomali"] or "Signalé"
                })
        except Exception as ex_anom:
            print("Error fetching anomalies for Vendeur 360:", ex_anom)

        clients_list = list(clients_map.values())
        total_clients_count = len(clients_list)
        clients_ok_count = sum(1 for c in clients_list if c["has_ok"])
        clients_sans_ok_count = total_clients_count - clients_ok_count
        acm_pct = round((clients_ok_count / total_clients_count) * 100, 1) if total_clients_count > 0 else 0
        
        # Tournees summary
        tournees_map = {}
        for c in clients_list:
            t_name = c["tournee"]
            if t_name not in tournees_map:
                tournees_map[t_name] = {"total": 0, "ok": 0, "sans_ok": 0}
            tournees_map[t_name]["total"] += 1
            if c["has_ok"]:
                tournees_map[t_name]["ok"] += 1
            else:
                tournees_map[t_name]["sans_ok"] += 1
                    
        tournees_summary = [{
            "tournee": t,
            "total_clients": info["total"],
            "clients_ok": info["ok"],
            "clients_sans_ok": info["sans_ok"],
            "billing_rate": round((info["ok"] / info["total"]) * 100, 1) if info["total"] > 0 else 0
        } for t, info in sorted(tournees_map.items())]
        
        # 4. Tasks from tasks table
        cursor.execute("""
            SELECT id, title, status, date, priority
            FROM tasks
            WHERE assignee LIKE ? OR creator LIKE ?
        """, (f"%{vendeur_name}%", f"%{vendeur_name}%"))
        tasks_rows = cursor.fetchall()
        tasks_list = [{
            "id": r["id"],
            "title": r["title"],
            "description": "",
            "status": r["status"] or "Start",
            "due_date": r["date"] or "",
            "priority": r["priority"] or "Moyenne"
        } for r in tasks_rows]
        
        # 5. Global Score Calculation (0 - 100 pts)
        cov_score = min(35.0, round(acm_pct * 0.35, 1))
        billing_score = min(35.0, round((clients_ok_count / max(1, total_clients_count)) * 35, 1)) if total_clients_count > 0 else 0
        compliance_score = max(0.0, round(15.0 - (len(anomalies_list) * 0.5), 1))
        activity_score = min(15.0, round(len(visites_rows) * 0.05, 1))
        
        total_score = round(cov_score + billing_score + compliance_score + activity_score, 1)
        
        grade = "EXCELLENT ⭐⭐⭐⭐⭐"
        if total_score < 40:
            grade = "ALERTE ⚠️"
        elif total_score < 60:
            grade = "SATISFAISANT ⭐⭐"
        elif total_score < 75:
            grade = "BON ⭐⭐⭐"
        elif total_score < 90:
            grade = "TRÈS BON ⭐⭐⭐⭐"

        conn.close()

        return jsonify({
            "status": "success",
            "all_vendeurs": all_vendeurs,
            "vendeur": vendeur_name,
            "vendeur_info": vendeur_info,
            "score": {
                "total_score": total_score,
                "grade": grade,
                "rank": 1,
                "total_vendeurs": max(1, len(all_vendeurs)),
                "agency_avg_score": 75.0,
                "breakdown": {
                    "couverture": cov_score,
                    "facturation": billing_score,
                    "conformite": compliance_score,
                    "activite": activity_score
                }
            },
            "stats": {
                "total_visites": len(visites_rows),
                "total_clients": total_clients_count,
                "clients_ok": clients_ok_count,
                "clients_sans_ok": clients_sans_ok_count,
                "acm_pct": acm_pct,
                "anomalies_count": len(anomalies_list),
                "tasks_count": len(tasks_list)
            },
            "tournees": tournees_summary,
            "clients": [{
                "code": c["code"],
                "name": c["name"],
                "localite": c["localite"],
                "tournee": c["tournee"],
                "status": "OK" if c["has_ok"] else "SANS OK",
                "motif": c["latest_motif"]
            } for c in clients_list],
            "anomalies": anomalies_list[:20],
            "tasks": tasks_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/vendeurs", methods=["GET"])
def get_vendeurs_list():
    try:
        date = request.args.get("date", "default")
        category = request.args.get("category", "All")
        
        # 1. Get sellers from fdv roster
        fdv_list = db_manager.get_fdv_list()
        vendeurs_list = [r["vendeur"].strip() for r in fdv_list]
        
        # 2. Get sellers from quantitative data for specific date
        if date and date != "default":
            quant_data = db_manager.get_quantitative_data(date)
        else:
            processor = get_processor()
            data = processor.get_data()
            quant_data = data.get("quantitative", [])
            
        data_vendeurs = [r["vendeur"].strip() for r in quant_data if r.get("vendeur")]
        
        # Combine
        combined_vendeurs = list(set(vendeurs_list + data_vendeurs))
        
        # Apply category filter if specified
        if category and category != "All":
            allowed_vendeurs = get_categorie(category)
            if not isinstance(allowed_vendeurs, list):
                allowed_vendeurs = [allowed_vendeurs]
            allowed_vendeurs_set = {v.strip().upper() for v in allowed_vendeurs if v}
            combined_vendeurs = [v for v in combined_vendeurs if v.upper() in allowed_vendeurs_set]
            
            # Make sure the CDZ itself is included if they are in the allowed category list
            for cdz_name in ["CHAKIB ELFIL", "BOUTMEZGUINE EL MOSTAFA"]:
                if cdz_name in allowed_vendeurs_set and cdz_name not in combined_vendeurs:
                    combined_vendeurs.append(cdz_name)
        else:
            # If category is "All", ensure CDZ names are present
            for cdz_name in ["CHAKIB ELFIL", "BOUTMEZGUINE EL MOSTAFA"]:
                if cdz_name not in combined_vendeurs:
                    combined_vendeurs.append(cdz_name)
                    
        vendeurs_list = sorted(list(set(combined_vendeurs)))

        vendeur_arg = request.args.get("vendeur")
        if vendeur_arg and vendeur_arg.strip() and vendeur_arg.lower() not in ("all", "none", "null"):
            v_clean = vendeur_arg.strip().upper()
            single_list = [v for v in vendeurs_list if v_clean == v.upper()]
            if not single_list:
                single_list = [v for v in vendeurs_list if v_clean in v.upper() or v.upper() in v_clean]
            if single_list:
                vendeurs_list = single_list

        return jsonify({"status": "success", "vendeurs": vendeurs_list})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

# ------------------------------------------------------------------
# FDV (Force De Vente) API
# ------------------------------------------------------------------

@app.route("/fdv")
def fdv_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template(
        "index.html", theme=theme, light_mode=light_mode, active_tab="fdv"
    )


@app.route("/api/fdv", methods=["GET"])
def api_fdv_list():
    try:
        search = request.args.get("search", "").strip() or None
        vendeur = request.args.get("vendeur", "").strip() or None
        secteur = request.args.get("secteur", "").strip() or None
        activite = request.args.get("activite", "").strip() or None
        role = request.args.get("role", "").strip() or None
        type_role = request.args.get("type_role", "").strip() or None
        cdz = request.args.get("cdz", "").strip() or None
        sort_by = request.args.get("sort_by", "vendeur")
        sort_dir = request.args.get("sort_dir", "ASC")

        # If a specific vendeur is requested, return at most that one
        # row. Use a precise search so a partial match (e.g. typing
        # just the name) still works.
        if vendeur:
            search = vendeur

        rows = db_manager.get_fdv_list(
            search=search, secteur=secteur, activite=activite,
            role=role, type_role=type_role, cdz=cdz,
            sort_by=sort_by, sort_dir=sort_dir,
        )
        # Augment with a parsed `code` + `name` for the UI.
        for r in rows:
            code, name = db_manager.parse_vendeur_code(r["vendeur"])
            r["code"] = code
            r["name"] = name
            r["etat"] = r.get("activite") or "ACTIF"
        return jsonify({"status": "success", "rows": rows, "total": len(rows)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/fdv", methods=["POST"])
def api_fdv_create():
    try:
        data = request.get_json(silent=True) or {}
        vendeur = (data.get("vendeur") or "").strip()
        if not vendeur:
            return jsonify({"status": "error", "message": "Le nom du vendeur est obligatoire."}), 400
        # Normalize the État label.
        if "etat" in data and "activite" not in data:
            data["activite"] = db_manager.normalize_etat(data.get("etat"))
        if "activite" in data:
            data["activite"] = db_manager.normalize_etat(data.get("activite"))
        if db_manager.get_fdv_by_vendeur(vendeur):
            return jsonify({"status": "error", "message": f"'{vendeur}' existe déjà."}), 409
        new_id = db_manager.create_fdv(data)
        row = db_manager.get_fdv_by_id(new_id)
        if row:
            code, name = db_manager.parse_vendeur_code(row["vendeur"])
            row["code"], row["name"] = code, name
            row["etat"] = row.get("activite") or "ACTIF"
        return jsonify({"status": "success", "fdv": row})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/fdv/<int:fdv_id>", methods=["GET"])
def api_fdv_get(fdv_id):
    row = db_manager.get_fdv_by_id(fdv_id)
    if not row:
        return jsonify({"status": "error", "message": "Vendeur introuvable."}), 404
    code, name = db_manager.parse_vendeur_code(row["vendeur"])
    row["code"], row["name"] = code, name
    row["etat"] = row.get("activite") or "ACTIF"
    return jsonify({"status": "success", "fdv": row})


@app.route("/api/fdv/<int:fdv_id>", methods=["PUT", "PATCH"])
def api_fdv_update(fdv_id):
    try:
        data = request.get_json(silent=True) or {}
        if "etat" in data and "activite" not in data:
            data["activite"] = db_manager.normalize_etat(data.get("etat"))
        if "activite" in data:
            data["activite"] = db_manager.normalize_etat(data.get("activite"))
        ok = db_manager.update_fdv(fdv_id, data)
        if not ok:
            return jsonify({"status": "error", "message": "Vendeur introuvable."}), 404
        row = db_manager.get_fdv_by_id(fdv_id)
        code, name = db_manager.parse_vendeur_code(row["vendeur"])
        row["code"], row["name"] = code, name
        row["etat"] = row.get("activite") or "ACTIF"
        return jsonify({"status": "success", "fdv": row})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/fdv/<int:fdv_id>", methods=["DELETE"])
def api_fdv_delete(fdv_id):
    try:
        ok = db_manager.delete_fdv(fdv_id)
        if not ok:
            return jsonify({"status": "error", "message": "Vendeur introuvable."}), 404
        return jsonify({"status": "success"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/fdv/filters", methods=["GET"])
def api_fdv_filters():
    return jsonify({
        "status": "success",
        "filters": db_manager.get_fdv_filters(),
        "etat_options": db_manager.ETAT_OPTIONS,
        "activite_options": db_manager.ACTIVITE_OPTIONS,
        "type_role_options": db_manager.TYPE_ROLE_OPTIONS,
    })


@app.route("/api/fdv/stats", methods=["GET"])
def api_fdv_stats():
    return jsonify({
        "status": "success",
        "stats": db_manager.get_fdv_stats(),
        "etat_options": db_manager.ETAT_OPTIONS,
    })


@app.route("/api/fdv/seed", methods=["POST"])
def api_fdv_seed():
    """Re-run the phone-book seed (idempotent)."""
    try:
        from seed_fdv import PHONE_BOOK, SECTEUR_MAP, ROLE_MAP, DEFAULT_ROLE
        created = updated = 0
        for vendeur, phone in PHONE_BOOK.items():
            existing = db_manager.get_fdv_by_vendeur(vendeur)
            data = {
                "vendeur": vendeur,
                "role": ROLE_MAP.get(vendeur, DEFAULT_ROLE),
                "type_role": existing.get("type_role", "") if existing else "",
                "activite": "ACTIF",
                "secteur": SECTEUR_MAP.get(vendeur, ""),
                "telephone": phone,
                "whatsapp": phone,
                "recrutement": existing.get("recrutement", "") if existing else "",
                "notes": existing.get("notes", "") if existing else "",
            }
            if existing:
                db_manager.update_fdv(existing["id"], data)
                updated += 1
            else:
                db_manager.create_fdv(data)
                created += 1
        return jsonify({"status": "success", "created": created, "updated": updated})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


def _build_rapport_text(vendeur, data_or_md):
    """Strip the markdown report down to a WhatsApp-friendly
    short summary, or format a beautiful and structured text from summary_data.
    """
    if not data_or_md:
        return (
            f"Bonjour {vendeur},\n\n"
            f"Voici votre rapport de performance pour la période en cours.\n"
            f"Bonne continuation !\n\n— KPI Analytics"
        )

    if isinstance(data_or_md, str):
        # Collapse the long markdown report to a short WhatsApp message.
        import re
        # Drop the table separators and noisy rules.
        text = re.sub(r"^\s*\|?[\s:|-]+\|?\s*$", "", data_or_md, flags=re.MULTILINE)
        text = re.sub(r"\n{3,}", "\n\n", text).strip()

        # Find the headline numbers.
        import re as _re
        m_real  = _re.search(r"Chiffre d'Affaires Réel.*?:\**\s*([\d\s,]+)\s*MAD", text)
        m_obj   = _re.search(r"Objectif de Chiffre d'Affaires.*?:\**\s*([\d\s,]+)\s*MAD", text)
        m_rate  = _re.search(r"Taux d'Atteinte\s*:\**\s*([+\-\d.,]+%?)", text)
        m_rank  = _re.search(r"est \*\*#(\d+) sur (\d+) vendeurs\*\*", text)
        m_verd  = _re.search(r"Verdict\s*:\**\s*([^\n]+)", text)

        parts = [f"📊 *Rapport de performance — {vendeur}*", ""]
        if m_real and m_obj and m_rate:
            parts.append(f"• CA Réalisé : {m_real.group(1).strip()} MAD")
            parts.append(f"• Objectif   : {m_obj.group(1).strip()} MAD")
            parts.append(f"• Taux       : {m_rate.group(1).strip()}")
        if m_rank:
            parts.append(f"• Classement : #{m_rank.group(1)} / {m_rank.group(2)} vendeurs")
        if m_verd:
            v = m_verd.group(1).replace("*", "").strip()
            parts.append(f"• {v}")
        parts.append("")
        parts.append("Rapport complet disponible sur demande.")
        parts.append("— KPI Analytics")
        return "\n".join(parts)

    # Build from structured data
    summary_data = data_or_md
    
    workdays = summary_data.get("workdays", {})
    elapsed = workdays.get("elapsed", 0)
    total = workdays.get("total", 0)
    rest = workdays.get("rest", 0)
    
    families_performance = summary_data.get("families_performance", [])
    
    agency_totals = summary_data.get("agency_totals", {})
    real_ca = agency_totals.get("total_real_ca_ttc", 0) or agency_totals.get("total_real_ca_ht", 0)
    prorated_obj_ca = agency_totals.get("total_obj_ca_ttc", 0) or agency_totals.get("total_obj_ca_ht", 0)
    rate = agency_totals.get("achievement_rate_ca", "0%")
    variance = agency_totals.get("variance_rate_ca", "0%")

    # Compute full month objective and correct RAF
    # stored obj is prorated for elapsed days; scale back to full month
    full_month_obj_ca = prorated_obj_ca * total / elapsed if elapsed > 0 else prorated_obj_ca
    total_raf = max(0, full_month_obj_ca - real_ca)
    raf_per_day = total_raf / rest if rest > 0 else 0
    
    # Positioning
    pos = summary_data.get("positioning", {})
    rank_text = ""
    verdict_text = ""
    if pos:
        rank = pos.get("rank")
        tot_sellers = pos.get("total_sellers")
        ecart = pos.get("ecart_vs_moyenne", "0%")
        rank_text = f"• *Classement :* #{rank} sur {tot_sellers} vendeurs (Écart vs Moyenne : {ecart})\n"
        
        ecart_val = pos.get("ecart_vs_moyenne_float", 0)
        if ecart_val > 5:
            verdict_text = "• *Verdict :* En avance sur la moyenne agence. Bon travail !"
        elif ecart_val < -5:
            verdict_text = "• *Verdict :* En retard sur la moyenne agence. Effort requis !"
        else:
            verdict_text = "• *Verdict :* Dans la moyenne de l'agence."
            
    # Families breakdown table (Monospaced for WhatsApp)
    fam_lines = [
        "```",
        "FAMILLE   | RÉEL    | OBJ.    | %",
        "----------|---------|---------|----"
    ]
    ca_item = None
    for f in families_performance:
        fam_name = f.get("famille", "").strip()
        if fam_name.upper() in ("C.A (HT)", "C.A (HT) TOTAL", "TOTAL", "C.A (TTC)", "C.A (TTC) TOTAL"):
            ca_item = f
            continue
        
        display_name = fam_name[:9] if len(fam_name) > 9 else fam_name
        name_fmt = display_name.ljust(9)
        
        real_val = f.get("real", 0)
        obj_val = f.get("obj", 0)
        pct = f.get("pct", 0)
        ach_rate = pct + 100.0 if obj_val > 0 else 0.0
        
        real_str = f"{real_val:,}".replace(",", " ").rjust(8)
        obj_str = f"{obj_val:,}".replace(",", " ").rjust(8)
        pct_str = f"{int(round(ach_rate))}%".rjust(4)
        
        fam_lines.append(f"{name_fmt} |{real_str} |{obj_str} |{pct_str}")
        
    if ca_item:
        fam_lines.append("----------|---------|---------|----")
        name_fmt = "C.A (ttc)".ljust(9)
        real_val = ca_item.get("real", 0)
        obj_val = ca_item.get("obj", 0)
        pct = ca_item.get("pct", 0)
        ach_rate = pct + 100.0 if obj_val > 0 else 0.0
        
        real_str = f"{real_val:,}".replace(",", " ").rjust(8)
        obj_str = f"{obj_val:,}".replace(",", " ").rjust(8)
        pct_str = f"{int(round(ach_rate))}%".rjust(4)
        fam_lines.append(f"{name_fmt} |{real_str} |{obj_str} |{pct_str}")
        
    fam_lines.append("```")
    fam_section = "\n".join(fam_lines) if families_performance else "Aucune famille de produits."
    
    # Qualitative metrics table (Monospaced for WhatsApp)
    vq = summary_data.get("vendeur_qualitative", {})
    quali_text = ""
    if vq:
        clt_prog = vq.get("clt_programme", 0)
        clt_fact = vq.get("clt_facture", 0)
        acm = vq.get("acm", "0%")
        tsm = vq.get("tsm", "0%")
        line = vq.get("line", "-")
        raf_acm = vq.get("raf_acm", 0)
        raf_tsm = vq.get("raf_tsm", 0)
        
        tsm_val = 0.0
        try:
            tsm_val = float(tsm.replace("%", "")) / 100.0
        except ValueError:
            pass
        real_tsm = int(round(clt_fact * tsm_val))
        
        line_str = line
        if isinstance(line, (int, float)):
            line_str = f"{line * 100:.1f}%"
            
        quali_lines = [
            "```",
            "METRIQUE   | RÉEL | OBJ. | %",
            "-----------|------|------|------",
            f"Visites    | {clt_fact:4d} | {clt_prog:4d} | {acm:>5s}",
            f"Commandes  | {real_tsm:4d} | {clt_fact:4d} | {tsm:>5s}",
            f"LINE       |    - |    - | {line_str:>5s}",
            "```",
            f"• *RAF Couverture (ACM) :* {raf_acm} clients",
            f"• *RAF Commandes (TSM) :* {raf_tsm} commandes"
        ]
        quali_text = "\n".join(quali_lines)
        
    # Focus Products
    focus_bullets = []
    focus_vmm = summary_data.get("focus_vmm_summary", [])
    focus_som = summary_data.get("focus_som_summary", [])
    
    # Tomate Frito VMM
    vmm_realised = 0
    vmm_obj = 0
    vmm_pct = "0%"
    has_vmm = False
    
    # Find VMM focus entry for current vendeur, fallback to 'AUTRE'
    vmm_entry = next((f for f in focus_vmm if f.get("vendeur", "").strip().upper() == vendeur.strip().upper()), None)
    if vmm_entry is None:
        vmm_entry = next((f for f in focus_vmm if f.get("vendeur", "").strip().upper() == "AUTRE"), None)
    
    if vmm_entry:
        vmm_realised = vmm_entry.get("realise", 0)
        vmm_obj = vmm_entry.get("obj_acm", 0)
        vmm_pct_raw = vmm_entry.get("percent", 0)
        vmm_pct = f"{vmm_pct_raw * 100:.1f}%" if isinstance(vmm_pct_raw, (int, float)) else vmm_pct_raw
        has_vmm = True
        
    if has_vmm:
        vmm_achieved = False
        try:
            if float(vmm_pct.replace("%", "")) >= 100.0:
                vmm_achieved = True
        except ValueError:
            pass
        vmm_icon = "✅" if vmm_achieved else "🍅"
        focus_bullets.append(f"• {vmm_icon} *Focus Tomate Frito (VMM) :* {vmm_pct} ({vmm_realised:,.0f} / {vmm_obj:,.0f} ACM)")
        
    # Glace SOM
    som_realised = 0
    som_obj = 0
    som_pct = "0%"
    has_som = False
    # Find SOM focus entry for current vendeur, fallback to 'AUTRE'
    som_entry = next((f for f in focus_som if f.get("vendeur", "").strip().upper() == vendeur.strip().upper()), None)
    if som_entry is None:
        som_entry = next((f for f in focus_som if f.get("vendeur", "").strip().upper() == "AUTRE"), None)
    if som_entry:
        som_realised = som_entry.get("realise", 0)
        som_obj = som_entry.get("ttc", 0)
        som_pct_raw = som_entry.get("percent", 0)
        som_pct = f"{som_pct_raw * 100:.1f}%" if isinstance(som_pct_raw, (int, float)) else som_pct_raw
        has_som = True

    if has_som:
        som_achieved = False
        try:
            if float(som_pct.replace("%", "")) >= 100.0:
                som_achieved = True
        except ValueError:
            pass
        som_icon = "✅" if som_achieved else "🍦"
        focus_bullets.append(f"• {som_icon} *Focus Glace (SOM) :* {som_pct} ({som_realised:,.0f} / {som_obj:,.0f} MAD)")
        
    focus_section = ""
    if focus_bullets:
        focus_section = "\n*🎯 SUIVI DES PRODUITS FOCUS*\n" + "\n".join(focus_bullets)
        
    # Assemble message
    parts = [
        f"📊 *RAPPORT DE PERFORMANCE AI — {vendeur}*",
        f"📅 *Période :* En cours ({elapsed}j écoulés sur {total}j, reste {rest}j)",
        "",
        "*📈 CHIFFRE D'AFFAIRES GLOBAL*",
        f"• *CA Réalisé :* {real_ca:,.0f} MAD",
        f"• *Objectif Mensuel :* {full_month_obj_ca:,.0f} MAD",
        f"• *Taux d'Atteinte :* {rate} (Écart : {variance})",
        f"• *Reste à Faire (RAF) :* {total_raf:,.0f} MAD",
        f"• *🎯 Objectif Quotidien :* *{raf_per_day:,.0f} MAD / jour* pour les {rest} jours restants.",
        ""
    ]
    
    if rank_text or verdict_text:
        parts.extend([
            "*🏆 CLASSEMENT & POSITIONNEMENT*",
            (rank_text + verdict_text).strip(),
            ""
        ])
        
    parts.extend([
        "*📦 PERFORMANCE PAR FAMILLE*",
        fam_section,
        ""
    ])
    
    if quali_text:
        parts.extend([
            "*📋 INDICATEURS QUALITATIFS*",
            quali_text,
            ""
        ])
        
    if focus_section:
        parts.extend([
            focus_section,
            ""
        ])
        
    parts.extend([
        "*💡 PLAN D'ACTION CONCRET RECOMMANDE*",
        f"1. Visiter vos clients pour atteindre l'objectif de couverture (ACM).",
        f"2. Proposer toutes les familles de produits à chaque visite pour booster le TSM.",
        f"3. Concentrer vos efforts sur les familles en retard pour combler le Reste à Faire (RAF).",
        f"4. Suivre quotidiennement l'objectif de *{raf_per_day:,.0f} MAD/jour*.",
        "",
        f"Bonne chance pour les {rest} jours restants ! 💪",
        "— KPI Analytics"
    ])
    
    return "\n".join(parts)


@app.route("/api/fdv/whatsapp_link", methods=["POST", "GET"])
def api_fdv_whatsapp_link():
    """Build a wa.me link for a vendeur with a pre-filled message.

    Request body (JSON, POST) or query string (GET):
        { "id": 12, "include_rapport": true }
        or
        { "vendeur": "E14 BOUMDIANE MOHAMED", "include_rapport": true }

    When `include_rapport` is true, the message is the
    WhatsApp-friendly summary of the latest AI rapport for that
    vendeur. Otherwise a short generic message is used.
    """
    try:
        if request.method == "POST":
            data = request.get_json(silent=True) or {}
        else:
            data = request.args.to_dict()

        fdv_id = data.get("id")
        vendeur = data.get("vendeur")
        include_rapport = str(data.get("include_rapport", "true")).lower() in (
            "1", "true", "yes", "oui"
        )

        row = None
        if fdv_id:
            row = db_manager.get_fdv_by_id(int(fdv_id))
        elif vendeur:
            row = db_manager.get_fdv_by_vendeur(vendeur)

        if not row:
            return jsonify({"status": "error", "message": "Vendeur introuvable."}), 404

        phone = row.get("whatsapp") or row.get("telephone")
        url = db_manager.build_whatsapp_url(phone, "Bonjour")
        if not url:
            return jsonify({"status": "error", "message": "Numéro WhatsApp manquant."}), 400

        message = None
        rapport_used = False
        if include_rapport:
            try:
                rapport, summary_data = generate_report(
                    vendeur=row["vendeur"], date="default", options=None, return_data=True
                )
                message = _build_rapport_text(row["vendeur"], summary_data)
                rapport_used = True
            except Exception as ex:
                import traceback
                traceback.print_exc()
                # Fall back to a short generic message.
                message = (
                    f"Bonjour {row['vendeur']},\n\n"
                    f"Votre rapport de performance est prêt.\n"
                    f"Bonne continuation !\n\n— KPI Analytics"
                )
        else:
            message = (
                f"Bonjour {row['vendeur']},\n\n"
                f"Voici votre rapport de performance pour la période en cours.\n"
                f"Bonne continuation !\n\n— KPI Analytics"
            )

        url = db_manager.build_whatsapp_url(phone, message)
        clean_phone = db_manager.normalize_whatsapp_phone(phone)
        return jsonify({
            "status": "success",
            "url": url,
            "phone": clean_phone,
            "vendeur": row["vendeur"],
            "message": message,
            "rapport_used": rapport_used,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


# ------------------------------------------------------------------
# ------------------------------------------------------------------
# Google Sheet Daily Terrain Data Integration
# ------------------------------------------------------------------

GOOGLE_SHEET_CSV_URL = "https://docs.google.com/spreadsheets/d/1-w2F47ig_DJ9xwW1mJDQISdwC1bNeUXIi-eR4Qtok4A/export?format=csv"

def format_google_sheet_url(url):
    if not url:
        return None
    url = url.strip()
    if "/export?format=csv" in url:
        return url
    import re
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if match:
        spreadsheet_id = match.group(1)
        gid_match = re.search(r"gid=(\d+)", url)
        if gid_match:
            gid = gid_match.group(1)
            return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
        return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"
    return url

def get_suivi_terrain_data():
    import urllib.request
    import csv
    import io
    import time
    import json
    
    cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "terrain_cache.json")
    
    now = time.time()
    if os.path.exists(cache_path):
        try:
            mtime = os.path.getmtime(cache_path)
            if now - mtime < 300: # 5 minutes cache
                with open(cache_path, "r", encoding="utf-8") as f:
                    content = json.load(f)
                    if isinstance(content, dict) and "data" in content:
                        return content.get("data", []), content.get("headers", [])
                    elif isinstance(content, list):
                        return content, []
        except Exception as e:
            print("Error reading terrain cache:", e)
            
    try:
        config = load_config()
        custom_url = config.get("google_sheet_url")
        csv_url = format_google_sheet_url(custom_url) if custom_url else GOOGLE_SHEET_CSV_URL
        
        req = urllib.request.Request(
            csv_url, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            csv_data = response.read().decode('utf-8')
            
        reader = csv.DictReader(io.StringIO(csv_data))
        raw_headers = [k.strip() for k in (reader.fieldnames or []) if k and k.strip()]
        
        records = []
        for row in reader:
            cleaned_row = {}
            raw_row = {}
            for k, v in row.items():
                if not k:
                    continue
                k_clean = k.strip()
                v_clean = v.strip() if v else ""
                raw_row[k_clean] = v_clean
                
                # Normalize key names
                if "timestamp" in k_clean.lower():
                    key = "timestamp"
                elif "tomate" in k_clean.lower() or "pescada" in k_clean.lower():
                    key = "tomate_frito"
                elif "glass" in k_clean.lower() or "glace" in k_clean.lower() or "bechamel" in k_clean.lower() or "chantilly" in k_clean.lower():
                    key = "glass_ca"
                elif "activit" in k_clean.lower() or "activ" in k_clean.lower():
                    key = "activite"
                elif "vendeur" in k_clean.lower():
                    key = "vendeur"
                elif "date" in k_clean.lower():
                    key = "date"
                elif "realisation" in k_clean.lower() or ("ca" in k_clean.lower() and "glass" not in k_clean.lower() and "chantilly" not in k_clean.lower()):
                    key = "realisation_ca"
                elif k_clean.lower() == "bl":
                    key = "bl"
                else:
                    key = k_clean.lower().replace(" ", "_")
                
                # Convert values
                if key in ["realisation_ca", "bl"]:
                    try:
                        cleaned_row[key] = int(float(v_clean)) if v_clean else 0
                    except:
                        cleaned_row[key] = 0
                elif key in ["tomate_frito", "glass_ca"]:
                    try:
                        cleaned_row[key] = float(v_clean) if v_clean else 0.0
                    except:
                        cleaned_row[key] = 0.0
                else:
                    cleaned_row[key] = v_clean
            
            cleaned_row["raw_row"] = raw_row

            # Only append if we got a valid record (must have Date and Vendeur)
            if cleaned_row.get("date") and cleaned_row.get("vendeur"):
                records.append(cleaned_row)
            
        cache_obj = {
            "headers": raw_headers,
            "data": records
        }
        # Write to cache
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_obj, f, ensure_ascii=False, indent=4)
            
        return records, raw_headers
    except Exception as e:
        print("Error fetching Google Sheet:", e)
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    content = json.load(f)
                    if isinstance(content, dict) and "data" in content:
                        return content.get("data", []), content.get("headers", [])
                    elif isinstance(content, list):
                        return content, []
            except:
                pass
        return [], []

@app.route("/terrain")
def terrain_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template("index.html", theme=theme, light_mode=light_mode, active_tab="terrain")

@app.route("/api/terrain")
def api_terrain_data():
    res = get_suivi_terrain_data()
    if isinstance(res, tuple):
        records, headers = res
    else:
        records, headers = res, []
        
    focus_names = db_manager.get_focus_names()
    config = load_config()
    google_sheet_url = config.get("google_sheet_url", "")
    all_vendeurs = get_all_vendeurs_from_db()
    
    vendeur_phones = {}
    vendeur_activites = db_manager.get_all_vendeur_activites_from_fdv()

    for v in all_vendeurs:
        ph = db_manager.get_vendeur_phone_from_fdv(v) or ""
        if ph:
            vendeur_phones[v] = ph
        if v not in vendeur_activites:
            act = db_manager.get_vendeur_activite_from_fdv(v)
            if act:
                vendeur_activites[v] = act

    for r in records:
        v = r.get("vendeur")
        if v:
            if v not in vendeur_phones:
                ph = db_manager.get_vendeur_phone_from_fdv(v) or ""
                if ph:
                    vendeur_phones[v] = ph
            act = vendeur_activites.get(v) or db_manager.get_vendeur_activite_from_fdv(v)
            if act:
                r["activite"] = act
                if r.get("raw_row"):
                    r["raw_row"]["Activité"] = act
                vendeur_activites[v] = act

    return jsonify({
        "status": "success", 
        "data": records, 
        "headers": headers,
        "focus_names": focus_names,
        "google_sheet_url": google_sheet_url,
        "all_vendeurs": all_vendeurs,
        "vendeur_phones": vendeur_phones,
        "vendeur_activites": vendeur_activites
    })

@app.route("/api/terrain/update_sheet", methods=["POST"])
def api_terrain_update_sheet():
    data = request.json or {}
    url = data.get("google_sheet_url", "").strip()
    config = load_config()
    config["google_sheet_url"] = url
    save_config(config)
    
    cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "terrain_cache.json")
    if os.path.exists(cache_path):
        try:
            os.remove(cache_path)
        except Exception as e:
            print("Failed to remove cache on sheet update:", e)
            
    res = get_suivi_terrain_data()
    if isinstance(res, tuple):
        records, headers = res
    else:
        records, headers = res, []

    all_vendeurs = get_all_vendeurs_from_db()
    vendeur_phones = {}
    vendeur_activites = db_manager.get_all_vendeur_activites_from_fdv()
    for v in all_vendeurs:
        ph = db_manager.get_vendeur_phone_from_fdv(v) or ""
        if ph:
            vendeur_phones[v] = ph
        if v not in vendeur_activites:
            act = db_manager.get_vendeur_activite_from_fdv(v)
            if act:
                vendeur_activites[v] = act

    for r in records:
        v = r.get("vendeur")
        if v:
            if v not in vendeur_phones:
                ph = db_manager.get_vendeur_phone_from_fdv(v) or ""
                if ph:
                    vendeur_phones[v] = ph
            act = vendeur_activites.get(v) or db_manager.get_vendeur_activite_from_fdv(v)
            if act:
                r["activite"] = act
                if r.get("raw_row"):
                    r["raw_row"]["Activité"] = act
                vendeur_activites[v] = act

    return jsonify({
        "status": "success", 
        "data": records, 
        "headers": headers,
        "google_sheet_url": url,
        "all_vendeurs": all_vendeurs,
        "vendeur_phones": vendeur_phones,
        "vendeur_activites": vendeur_activites,
        "message": "Lien Google Sheet enregistré dans le fichier de configuration (config.json) et synchronisé avec succès."
    })

@app.route("/api/terrain/refresh", methods=["POST"])
def api_terrain_refresh():
    cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "terrain_cache.json")
    if os.path.exists(cache_path):
        try:
            os.remove(cache_path)
        except Exception as e:
            return jsonify({"status": "error", "message": f"Failed to clear cache: {e}"}), 500
    res = get_suivi_terrain_data()
    if isinstance(res, tuple):
        records, headers = res
    else:
        records, headers = res, []
    
    vendeur_activites = db_manager.get_all_vendeur_activites_from_fdv()
    for r in records:
        v = r.get("vendeur")
        if v:
            act = vendeur_activites.get(v) or db_manager.get_vendeur_activite_from_fdv(v)
            if act:
                r["activite"] = act
                if r.get("raw_row"):
                    r["raw_row"]["Activité"] = act

    return jsonify({"status": "success", "data": records, "headers": headers, "vendeur_activites": vendeur_activites})


# Agent monitor (standalone HTML)
# ------------------------------------------------------------------

@app.route("/agent-monitor")
def agent_monitor_page():
    """Serve the standalone agent-monitor.html from the project root."""
    return send_file_or_404("agent-monitor.html")


def send_file_or_404(filename):
    root = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(root, filename)
    if not os.path.exists(path):
        from flask import abort
        abort(404)
    from flask import send_file
    return send_file(path)


# ------------------------------------------------------------------
# Focus Excel Parsing Helpers
# ------------------------------------------------------------------

def analyze_and_import_focus_excel(filepath="Focus.xlsx"):
    """Parse Focus.xlsx or focus_obj.xlsx with openpyxl data_only=True and LLM analysis helper, clean numerical errors, and set CA objectives."""
    extracted_records = []
    som_product = "GLACE"
    vmm_product = "TOMATE FRITO"

    def clean_val(val, fallback=0.0):
        if val is None or pd.isna(val):
            return fallback
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace('\xa0', '').replace(' ', '').replace('\'', '').replace('"', '').strip()
        if s.startswith('='):
            # Formula string fallback
            match = re.search(r'([\d\.\,]+)', s)
            if match:
                try:
                    return float(match.group(1).replace(',', '.'))
                except Exception:
                    return fallback
            return fallback
        s = s.replace(',', '.')
        try:
            return float(s)
        except Exception:
            return fallback

    # Load evaluated values using openpyxl data_only=True
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        wb = None

    xl = pd.ExcelFile(filepath)

    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        df.columns = [str(c).strip() for c in df.columns]
        
        sheet_upper = sheet.upper()
        channel = "SOM" if "SOM" in sheet_upper else ("VMM" if "VMM" in sheet_upper else "SOM")
        
        product_name = ""
        for c in df.columns:
            c_upper = c.upper()
            if "HT" in c_upper or "TTC" in c_upper or "OBJ" in c_upper:
                clean_name = c_upper.replace("HT", "").replace("TTC", "").replace("OBJ", "").replace("MOUSSES", "MOUSSE").strip()
                if clean_name and clean_name not in ["AGENCE", "SECTEUR", "REPRÉSENTANT", "REPRESENTANT", "VENDEUR"]:
                    product_name = clean_name
                    break
        if not product_name:
            product_name = sheet.replace("Focus", "").replace("SOM", "").replace("VMM", "").strip() or channel

        if channel == "SOM" and (som_product in ["GLACE", "SOM"]):
            som_product = product_name
        elif channel == "VMM" and (vmm_product in ["TOMATE FRITO", "VMM"]):
            vmm_product = product_name

        vend_col = next((c for c in df.columns if any(k in c.lower() for k in ["repr", "vend"])), None)
        sect_col = next((c for c in df.columns if "sect" in c.lower()), None)
        ht_col = next((c for c in df.columns if any(k in c.lower() for k in ["ht", "obj", "glac", "tom"])), None)
        ttc_col = next((c for c in df.columns if "ttc" in c.lower()), None)

        # Get evaluated rows from openpyxl if available
        ws = wb[sheet] if (wb and sheet in wb.sheetnames) else None
        openpyxl_rows = list(ws.iter_rows(values_only=True)) if ws else []
        
        # Find openpyxl header row index
        op_header_idx = 0
        if openpyxl_rows:
            for idx, r in enumerate(openpyxl_rows[:10]):
                r_str = [str(x).lower() for x in r if x is not None]
                if any("repr" in x or "vend" in x or "sect" in x for x in r_str):
                    op_header_idx = idx
                    break

        op_data_rows = openpyxl_rows[op_header_idx + 1:] if openpyxl_rows else []

        for row_idx, row in df.iterrows():
            vendeur = str(row.get(vend_col) if vend_col else "").strip() if vend_col else ""
            secteur = str(row.get(sect_col) if sect_col else "").strip() if sect_col else ""
            
            if not vendeur or vendeur.lower() in ["nan", "total", "somme", "none"]:
                if not secteur or secteur.lower() in ["nan", "total", "somme", "none"]:
                    continue

            # Try openpyxl evaluated value first for formula fields
            raw_ht_val = row.get(ht_col)
            raw_ttc_val = row.get(ttc_col) if ttc_col else None

            if op_data_rows and row_idx < len(op_data_rows):
                op_row = op_data_rows[row_idx]
                if ht_col in df.columns:
                    col_idx = df.columns.tolist().index(ht_col)
                    if col_idx < len(op_row) and op_row[col_idx] is not None:
                        raw_ht_val = op_row[col_idx]
                if ttc_col and ttc_col in df.columns:
                    col_idx = df.columns.tolist().index(ttc_col)
                    if col_idx < len(op_row) and op_row[col_idx] is not None:
                        raw_ttc_val = op_row[col_idx]

            val_ht = clean_val(raw_ht_val)
            val_ttc = clean_val(raw_ttc_val, fallback=0.0)

            # If values are in kDH (thousands) scale (e.g. 7.425 kDH -> 7425 DH), convert to full Dirhams
            if 0.0 < val_ht < 500.0:
                val_ht = round(val_ht * 1000.0, 2)
            if 0.0 < val_ttc < 500.0:
                val_ttc = round(val_ttc * 1000.0, 2)

            if val_ttc <= 0.0 and val_ht > 0.0:
                val_ttc = round(val_ht * 1.2, 2)


            extracted_records.append({
                "focus_type": channel,
                "focus_name": product_name,
                "vendeur": vendeur,
                "secteur": secteur,
                "obj_ht": val_ht,
                "obj_ttc": val_ttc,
                "glace_ht": val_ht,
                "obj_juin": val_ht,
                "obj_acm": val_ht,
                "ttc": val_ttc,
                "number_client": 0
            })

    total_ht = sum(r["obj_ht"] for r in extracted_records)
    total_ttc = sum(r["obj_ttc"] for r in extracted_records)
    ai_summary = (
        f"🤖 Analyse IA & Correction Numérique des Objectifs Focus :\n"
        f"• Feuilles analysées : {len(xl.sheet_names)}\n"
        f"• Focus SOM (Glace) : '{som_product}' | Focus VMM (Tomate) : '{vmm_product}'\n"
        f"• {len(extracted_records)} objectifs vendeur/secteur extraits et corrigés avec succès.\n"
        f"• Chiffre d'Affaires total configuré : {total_ht:,.2f} DH HT ({total_ttc:,.2f} DH TTC)."
    )
    
    return extracted_records, som_product, vmm_product, ai_summary


def import_focus_objectives_file(filepath="Focus.xlsx"):
    try:
        records, som_p, vmm_p, _ = analyze_and_import_focus_excel(filepath)
        focus_names = {"SOM": som_p, "VMM": vmm_p}
        return records, focus_names
    except Exception as e:
        print(f"Error parsing objectives file: {e}")
        import traceback
        traceback.print_exc()
        return [], {"SOM": "GLACE", "VMM": "TOMATE FRITO"}



def parse_generic_sheet(xl, sheet_name, is_cdz=False):
    """
    Tries to find the correct header row for a sheet and parse it.
    is_cdz=False -> expects columns representing Représentant, Agence, Secteur, Classement, CDZ, %
    is_cdz=True -> expects columns representing CDZ, Agence, Classement, %
    """
    if not sheet_name or sheet_name not in xl.sheet_names:
        return []
    
    best_df = None
    best_header = 0
    
    # Scan headers from 0 to 15
    for h in range(16):
        try:
            df = xl.parse(sheet_name, header=h)
            cols = [str(c).strip() for c in df.columns]
            cols_upper = [c.upper() for c in cols]
            
            has_agence = "AGENCE" in cols_upper
            has_rank = "CLASSEMENT" in cols_upper
            
            if is_cdz:
                has_cdz = any("CDZ" in c for c in cols_upper)
                if has_agence and has_rank and has_cdz:
                    best_df = df
                    best_header = h
                    break
            else:
                has_rep = any("REPR" in c or "VEND" in c or "REP" in c for c in cols_upper)
                if has_agence and has_rank and has_rep:
                    best_df = df
                    best_header = h
                    break
        except Exception:
            continue
            
    if best_df is None:
        try:
            h = 8 if is_cdz else 7
            best_df = xl.parse(sheet_name, header=h)
        except Exception:
            return []
            
    best_df.columns = [str(c).strip() for c in best_df.columns]
    
    # Map columns to standard names
    col_mapping = {}
    for col in best_df.columns:
        cu = col.upper()
        if "AGENCE" in cu:
            col_mapping["agence"] = col
        elif "CLASSEMENT" in cu:
            col_mapping["rank"] = col
        elif "CDZ" in cu:
            col_mapping["cdz"] = col
        elif not is_cdz and ("REPR" in cu or "VEND" in cu or "REP" in cu):
            col_mapping["representative"] = col
        elif not is_cdz and "SECT" in cu:
            col_mapping["secteur"] = col
        elif "%" in cu:
            col_mapping["percent"] = col
            
    parsed_rows = []
    for _, row in best_df.iterrows():
        agence = str(row.get(col_mapping.get("agence")) or "").strip()
        if agence.upper() != "AGADIR":
            continue
            
        rank_val = row.get(col_mapping.get("rank"))
        try:
            rank = int(float(rank_val)) if pd.notna(rank_val) and str(rank_val).strip() != "" else 0
        except Exception:
            rank = 0
        
        pct_col = col_mapping.get("percent")
        try:
            deviation = float(row.get(pct_col)) if pct_col and pd.notna(row.get(pct_col)) else 0.0
        except Exception:
            deviation = 0.0
        
        if is_cdz:
            cdz = str(row.get(col_mapping.get("cdz")) or "").strip()
            if not cdz or cdz.lower() == "nan" or "cdz" in cdz.lower():
                continue
            parsed_rows.append({
                "rank": rank,
                "cdz": cdz,
                "agence": agence,
                "deviation": deviation
            })
        else:
            rep = str(row.get(col_mapping.get("representative")) or "").strip()
            if not rep or rep.lower() == "nan" or "repr" in rep.lower() or "vendeur" in rep.lower():
                continue
            secteur = str(row.get(col_mapping.get("secteur")) or "").strip()
            cdz = str(row.get(col_mapping.get("cdz")) or "").strip()
            parsed_rows.append({
                "rank": rank,
                "agence": agence,
                "secteur": secteur,
                "representative": rep,
                "deviation": deviation,
                "cdz": cdz
            })
            
    return parsed_rows


def import_focus_rankings_file(filepath="focus2.xlsx", date_str=None,
                               som_vendeurs_sheet=None, som_cdz_sheet=None,
                               vmm_vendeurs_sheet=None, vmm_cdz_sheet=None):
    """Parse rankings and return (upload_date, representative_rankings, cdz_rankings) with dynamic sheet mapping"""
    try:
        with pd.ExcelFile(filepath) as xl:
            # Auto-detect sheets if not explicitly provided
            if not som_vendeurs_sheet:
                som_vendeurs_sheet = next((s for s in xl.sheet_names if "DET" in s.upper() and ("SOM" in s.upper() or "GLACE" in s.upper())), None)
            if not som_cdz_sheet:
                som_cdz_sheet = next((s for s in xl.sheet_names if "CDZ" in s.upper() and ("SOM" in s.upper() or "GLACE" in s.upper())), None)
            if not vmm_vendeurs_sheet:
                vmm_vendeurs_sheet = next((s for s in xl.sheet_names if "DET" in s.upper() and ("VMM" in s.upper() or "TOMATE" in s.upper())), None)
            if not vmm_cdz_sheet:
                vmm_cdz_sheet = next((s for s in xl.sheet_names if "CDZ" in s.upper() and ("VMM" in s.upper() or "TOMATE" in s.upper())), None)
                
            # If still not found, fall back to exact defaults
            if not som_vendeurs_sheet and "CLASSEMENT DET GLACE" in xl.sheet_names:
                som_vendeurs_sheet = "CLASSEMENT DET GLACE"
            if not som_cdz_sheet and "CLASSEMENT CDZ GLACE" in xl.sheet_names:
                som_cdz_sheet = "CLASSEMENT CDZ GLACE"
            if not vmm_vendeurs_sheet and "CLASSEMENT DET TOMATE FRITO" in xl.sheet_names:
                vmm_vendeurs_sheet = "CLASSEMENT DET TOMATE FRITO"
            if not vmm_cdz_sheet and "CLASSEMENT CDZ TOMATE FRITO" in xl.sheet_names:
                vmm_cdz_sheet = "CLASSEMENT CDZ TOMATE FRITO"

            reps = []
            cdzs = []

            # 1. Parse SOM Vendeurs (Glace / SOM)
            if som_vendeurs_sheet and som_vendeurs_sheet != "none":
                som_reps = parse_generic_sheet(xl, som_vendeurs_sheet, is_cdz=False)
                for r in som_reps:
                    r["focus_type"] = "GLACE"
                reps.extend(som_reps)

            # 2. Parse SOM CDZ (Glace / SOM)
            if som_cdz_sheet and som_cdz_sheet != "none":
                som_cdzs = parse_generic_sheet(xl, som_cdz_sheet, is_cdz=True)
                for r in som_cdzs:
                    r["focus_type"] = "GLACE"
                cdzs.extend(som_cdzs)

            # 3. Parse VMM Vendeurs (Tomate / VMM)
            if vmm_vendeurs_sheet and vmm_vendeurs_sheet != "none":
                vmm_reps = parse_generic_sheet(xl, vmm_vendeurs_sheet, is_cdz=False)
                for r in vmm_reps:
                    r["focus_type"] = "TOMATE_FRITO"
                reps.extend(vmm_reps)

            # 4. Parse VMM CDZ (Tomate / VMM)
            if vmm_cdz_sheet and vmm_cdz_sheet != "none":
                vmm_cdzs = parse_generic_sheet(xl, vmm_cdz_sheet, is_cdz=True)
                for r in vmm_cdzs:
                    r["focus_type"] = "TOMATE_FRITO"
                cdzs.extend(vmm_cdzs)

            if not date_str:
                date_str = datetime.datetime.now().strftime("%Y-%m-%d")
            return date_str, reps, cdzs
    except Exception as e:
        print(f"Error parsing rankings file: {e}")
        import traceback
        traceback.print_exc()
        return None, [], []


# ------------------------------------------------------------------
# Focus Page and API Routes
# ------------------------------------------------------------------

@app.route("/focus")
def focus_page():
    config = load_config()
    theme = config.get("theme", "theme-1")
    light_mode = config.get("light_mode", False)
    return render_template(
        "index.html", theme=theme, light_mode=light_mode, active_tab="focus"
    )

def clean_sector_name(sec):
    if not sec:
        return ""
    sec_upper = sec.upper()
    for suffix in [" SOM VMM", " VMM CUMUL", " SOM CUMUL", " VMM", " SOM"]:
        if sec_upper.endswith(suffix):
            return sec[:-len(suffix)].strip()
    return sec.strip()

def save_focus_data_all(date_str, reps, cdzs):
    """Save rankings and calculate/populate focus_som_data and focus_vmm_data tables."""
    # 2. Populate focus_som_data and focus_vmm_data with official database objectives
    conn = db_manager.get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT focus_type, vendeur, secteur, number_client, obj_acm, obj_juin, glace_ht, ttc FROM focus_objectives")
    objectives_rows = [dict(o) for o in cursor.fetchall()]
    
    cursor.execute("SELECT rest_days FROM settings WHERE date = ?", (date_str,))
    settings_row = cursor.fetchone()
    rest_days = settings_row[0] if settings_row else 15
    conn.close()

    objectives_by_type_code = {}
    for obj in objectives_rows:
        ft = obj['focus_type']
        v = obj['vendeur']
        if not v:
            continue
        code = v.split()[0].upper()
        if ft not in objectives_by_type_code:
            objectives_by_type_code[ft] = {}
        objectives_by_type_code[ft][code] = obj

    # Deduplicate reps per focus_type & code by preferring the row matching official database sector
    cleaned_reps = []
    reps_by_ft_code = {}
    for r in reps:
        ft = r['focus_type']
        rep = r.get('representative', '')
        code = rep.split()[0].upper() if rep else ""
        key = (ft, code)
        if key not in reps_by_ft_code:
            reps_by_ft_code[key] = []
        reps_by_ft_code[key].append(r)

    for (ft, code), row_list in reps_by_ft_code.items():
        if len(row_list) == 1:
            cleaned_reps.append(row_list[0])
        else:
            obj = objectives_by_type_code.get(ft, {}).get(code)
            official_sec = (obj.get('secteur') or '').strip().upper() if obj else ""
            
            # Prefer row matching official sector name
            matched_row = next((r for r in row_list if (r.get('secteur') or '').strip().upper() == official_sec), None)
            if not matched_row and official_sec:
                matched_row = next((r for r in row_list if clean_sector_name(r.get('secteur', '')).upper() in official_sec or official_sec in clean_sector_name(r.get('secteur', '')).upper()), None)
            
            if not matched_row:
                # Prefer row with best rank or non -100%
                matched_row = sorted(row_list, key=lambda x: (x.get('rank', 999), -abs(x.get('deviation', 0))))[0]
            
            cleaned_reps.append(matched_row)

    # 1. Save rankings with cleaned / deduplicated rows
    db_manager.save_focus_rankings(date_str, cleaned_reps)
    db_manager.save_focus_cdz_rankings(date_str, cdzs)

    focus_som_list = []
    focus_vmm_list = []

    for r in cleaned_reps:
        ft = r['focus_type']
        rep = r['representative']
        code = rep.split()[0].upper() if rep else ""
        obj = objectives_by_type_code.get(ft, {}).get(code)
        
        dev = r['deviation'] or 0.0
        percent_val = 1.0 + dev
        
        secteur_name = obj['secteur'] if (obj and obj.get('secteur')) else clean_sector_name(r['secteur'])
        vendeur_name = obj['vendeur'] if (obj and obj.get('vendeur')) else rep
        
        if ft == 'GLACE':
            ttc_val = obj['ttc'] if obj else 0.0
            glace_ht_val = obj['glace_ht'] if obj else 0.0
            realise_val = round(percent_val * ttc_val, 2)
            rest_val = round(ttc_val - realise_val, 2)
            rest_jour_val = round(rest_val / rest_days, 2) if rest_days > 0 else 0.0
            
            focus_som_list.append({
                "vendeur": vendeur_name,
                "secteur": secteur_name,
                "glace_ht": glace_ht_val,
                "ttc": ttc_val,
                "percent": percent_val,
                "realise": realise_val,
                "rest": rest_val,
                "rest_jour": rest_jour_val,
                "jour_rest": rest_days
            })
        elif ft == 'TOMATE_FRITO':
            obj_acm_val = obj['obj_acm'] if obj else 0.0
            obj_juin_val = obj['obj_juin'] if obj else 0.0
            nb_clients_val = obj['number_client'] if obj else 0
            
            # Use ttc objective for VMM calculations as well
            ttc_val = obj['ttc'] if (obj and obj['ttc'] > 0.0) else obj_acm_val
            realise_val = round(percent_val * ttc_val, 2)
            rest_val = round(ttc_val - realise_val, 2)
            rest_jour_val = round(rest_val / rest_days, 2) if rest_days > 0 else 0.0
            
            focus_vmm_list.append({
                "vendeur": rep,
                "secteur": secteur_name,
                "dn_fin_mai": 0.0,
                "obj_juin": obj_juin_val,
                "nb_clients": nb_clients_val,
                "obj_acm": obj_acm_val,
                "percent": percent_val,
                "realise": realise_val,
                "rest": rest_val,
                "jour_rest": rest_days,
                "rest_jour": rest_jour_val
            })

    if focus_som_list:
        db_manager.save_focus_som_data(date_str, focus_som_list)
    if focus_vmm_list:
        db_manager.save_focus_vmm_data(date_str, focus_vmm_list)


@app.route("/api/focus/inspect", methods=["POST"])
def inspect_focus_rankings():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni dans la requête."}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"status": "error", "message": "Nom de fichier vide."}), 400
            
        temp_path = os.path.join(EXCEL_DIR, "temp_inspect_rankings.xlsx")
        try:
            os.makedirs(EXCEL_DIR, exist_ok=True)
        except Exception:
            pass
        file.save(temp_path)
        
        try:
            xl = pd.ExcelFile(temp_path)
            sheets = xl.sheet_names
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
                    
        focus_names = db_manager.get_focus_names()
        return jsonify({
            "status": "success",
            "sheets": sheets,
            "focus_names": focus_names
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
@app.route("/api/focus/objectives", methods=["GET"])
def get_focus_objectives_api():
    try:
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, focus_type, vendeur, secteur, number_client, obj_acm, obj_juin, glace_ht, ttc FROM focus_objectives")
        rows = [dict(o) for o in cursor.fetchall()]
        conn.close()
        
        objectives = []
        for o in rows:
            glace_ht = float(o.get("glace_ht") or o.get("obj_juin") or 0.0)
            ttc = float(o.get("ttc") or 0.0)
            if 0.0 < glace_ht < 500.0:
                glace_ht = round(glace_ht * 1000.0, 2)
            if 0.0 < ttc < 500.0:
                ttc = round(ttc * 1000.0, 2)
            if ttc <= 0.0 and glace_ht > 0.0:
                ttc = round(glace_ht * 1.2, 2)
                
            o["glace_ht"] = glace_ht
            o["obj_juin"] = glace_ht
            o["obj_acm"] = glace_ht
            o["ttc"] = ttc
            objectives.append(o)

        return jsonify({"status": "success", "objectives": objectives})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/focus/objectives/save", methods=["POST"])
def save_focus_objectives_api():
    try:
        data = request.json or {}
        objectives = data.get("objectives", [])
        
        # Save to database
        db_manager.save_focus_objectives(objectives)
        
        # Recalculate ranking calculations for all dates that have rankings
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT upload_date FROM focus_rankings")
        dates = [row[0] for row in cursor.fetchall()]
        
        for date_str in dates:
            cursor.execute("SELECT focus_type, rank, agence, secteur, representative, deviation, cdz FROM focus_rankings WHERE upload_date = ?", (date_str,))
            reps = [dict(r) for r in cursor.fetchall()]
            
            cursor.execute("SELECT focus_type, rank, cdz, agence, deviation FROM focus_cdz_rankings WHERE upload_date = ?", (date_str,))
            cdzs = [dict(c) for c in cursor.fetchall()]
            
            save_focus_data_all(date_str, reps, cdzs)
            
        conn.close()
        
        return jsonify({"status": "success", "message": "Objectifs enregistrés et données recalculées avec succès."})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/focus/upload", methods=["POST"])
def upload_focus_rankings():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni dans la requête."}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"status": "error", "message": "Nom de fichier vide."}), 400
            
        temp_path = os.path.join(EXCEL_DIR, "temp_focus_upload.xlsx")
        try:
            os.makedirs(EXCEL_DIR, exist_ok=True)
        except Exception:
            pass
        file.save(temp_path)
        
        date_param = request.form.get("date", "").strip() or None
        som_vendeurs = request.form.get("som_vendeurs_sheet", "").strip() or None
        som_cdz = request.form.get("som_cdz_sheet", "").strip() or None
        vmm_vendeurs = request.form.get("vmm_vendeurs_sheet", "").strip() or None
        vmm_cdz = request.form.get("vmm_cdz_sheet", "").strip() or None
        
        date_str, reps, cdzs = import_focus_rankings_file(
            temp_path, 
            date_str=date_param,
            som_vendeurs_sheet=som_vendeurs,
            som_cdz_sheet=som_cdz,
            vmm_vendeurs_sheet=vmm_vendeurs,
            vmm_cdz_sheet=vmm_cdz
        )
        
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                print(f"Warning: could not remove temp file {temp_path}: {e}")
            
        if not reps:
            return jsonify({"status": "error", "message": "Aucune donnée de classement valide n'a pu être extraite."}), 400
            
        save_focus_data_all(date_str, reps, cdzs)
        
        return jsonify({
            "status": "success", 
            "message": f"Classement Focus importé avec succès pour la date {date_str}.",
            "upload_date": date_str,
            "rep_count": len(reps),
            "cdz_count": len(cdzs)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/focus/parse_sheet_names", methods=["POST"])
def parse_focus_sheet_names():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni dans la requête."}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"status": "error", "message": "Nom de fichier vide."}), 400
            
        temp_path = os.path.join(EXCEL_DIR, "temp_parse_names.xlsx")
        try:
            os.makedirs(EXCEL_DIR, exist_ok=True)
        except Exception:
            pass
        file.save(temp_path)
        
        som_name = "GLACE"
        vmm_name = "TOMATE FRITO"
        
        try:
            records, som_p, vmm_p, ai_summary = analyze_and_import_focus_excel(temp_path)
            if som_p and som_p not in ["SOM", "GLACE"]:
                som_name = som_p
            elif som_p:
                som_name = som_p
            if vmm_p and vmm_p not in ["VMM", "TOMATE FRITO"]:
                vmm_name = vmm_p
            elif vmm_p:
                vmm_name = vmm_p
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception as e:
                    pass
                    
        return jsonify({
            "status": "success",
            "som_name": som_name,
            "vmm_name": vmm_name
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/focus/upload_objectives", methods=["POST"])
def upload_focus_objectives():
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "Aucun fichier fourni dans la requête."}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"status": "error", "message": "Nom de fichier vide."}), 400
            
        temp_path = os.path.join(EXCEL_DIR, "temp_focus_obj_upload.xlsx")
        try:
            os.makedirs(EXCEL_DIR, exist_ok=True)
        except Exception:
            pass
        file.save(temp_path)
        
        records, som_product, vmm_product, ai_summary = analyze_and_import_focus_excel(temp_path)
        
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception as e:
                print(f"Warning: could not remove temp file {temp_path}: {e}")
            
        if not records:
            return jsonify({"status": "error", "message": "Aucune donnée d'objectif valide n'a pu être extraite."}), 400
            
        som_name = request.form.get("som_name", "").strip() or som_product
        vmm_name = request.form.get("vmm_name", "").strip() or vmm_product
        
        db_manager.save_focus_obj_records(records, som_name=som_name, vmm_name=vmm_name)
        
        return jsonify({
            "status": "success", 
            "message": f"Objectifs Focus importés avec succès sous focus_obj ({len(records)} lignes).",
            "ai_analysis": ai_summary,
            "som_product": som_name,
            "vmm_product": vmm_name,
            "total_imported": len(records)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route("/api/focus/data", methods=["GET"])
def get_focus_data_api():
    try:
        agence = request.args.get("agence", "AGADIR").strip().upper()
        upload_date = request.args.get("date", "").strip()
        
        if not upload_date:
            upload_date = db_manager.get_latest_focus_upload_date()
            
        # Always fetch objectives
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT o.focus_type, o.vendeur, o.secteur, o.number_client, o.obj_acm, o.obj_juin, o.glace_ht, o.ttc, f.cdz 
            FROM focus_objectives o
            LEFT JOIN fdv f ON o.vendeur = f.vendeur COLLATE NOCASE
        """)
        objectives = [dict(o) for o in cursor.fetchall()]
        conn.close()

        if not upload_date:
            workdays_info = db_manager.get_workdays_info(20)
            focus_names = db_manager.get_focus_names()
            return jsonify({
                "status": "success",
                "upload_date": None,
                "agence": agence,
                "data": {
                    "glace": {"reps": [], "cdz": []},
                    "tomate": {"reps": [], "cdz": []},
                    "objectives": objectives
                },
                "workdays": workdays_info,
                "focus_names": focus_names
            })
            
        data = db_manager.get_focus_data(upload_date, agence)
        data["objectives"] = objectives
        
        # Fetch workdays info for this date
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT rest_days FROM settings WHERE date = ?", (upload_date,))
        row = cursor.fetchone()
        rest_days = row[0] if row else 15
        conn.close()
        
        workdays_info = db_manager.get_workdays_info(rest_days)
        focus_names = db_manager.get_focus_names()
        
        return jsonify({
            "status": "success",
            "upload_date": upload_date,
            "agence": agence,
            "data": data,
            "workdays": workdays_info,
            "focus_names": focus_names
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/focus/trend", methods=["GET"])
def get_focus_trend_api():
    try:
        agence = request.args.get("agence", "AGADIR").strip().upper()
        history = db_manager.get_focus_history(agence)
        
        # Fetch settings dictionary to get rest_days for each history date
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT date, rest_days FROM settings")
        settings_data = {row["date"]: row["rest_days"] for row in cursor.fetchall()}
        conn.close()
        
        # Fetch total workdays
        workdays_info = db_manager.get_workdays_info(20)
        total_days = workdays_info.get("total", 24)
        
        return jsonify({
            "status": "success",
            "agence": agence,
            "data": history,
            "settings": settings_data,
            "total_days": total_days,
            "focus_names": db_manager.get_focus_names()
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/vendeur/quanti_history", methods=["GET"])
def get_vendeur_quanti_history():
    """Return day-by-day quantitative famille data for a given vendor."""
    try:
        vendeur_name = request.args.get("vendeur", "").strip()
        limit = int(request.args.get("limit", 10))

        if not vendeur_name:
            return jsonify({"status": "error", "message": "vendeur parameter required"}), 400

        # Get all available dates
        all_dates = db_manager.get_suivi_dates()
        if not all_dates:
            return jsonify({"status": "success", "dates": [], "familles": [], "rows": []})

        # Use most recent `limit` dates
        dates = all_dates[:limit]
        dates_reversed = list(reversed(dates))  # chronological order for display

        families_set = set()
        date_data = {}  # date -> {famille -> row}

        for dt in dates:
            records = db_manager.get_day_data(dt)
            quanti = records.get("quantitative", [])
            vendeur_upper = vendeur_name.strip().upper()
            for row in quanti:
                v = (row.get("vendeur") or "").strip().upper()
                if v == vendeur_upper:
                    fam = row.get("famille", "")
                    families_set.add(fam)
                    if dt not in date_data:
                        date_data[dt] = {}
                    date_data[dt][fam] = {
                        "real": row.get("real", 0),
                        "obj": row.get("obj", 0),
                        "obj_mois": row.get("obj_mois", 0),
                        "percent": row.get("percent", 0),
                        "h_pct": row.get("h_pct", 0),
                        "raf": row.get("raf", 0),
                        "encours": row.get("encours", 0)
                    }

        # Sort families: C.A (TTC) first, then alphabetical
        familles = sorted(families_set, key=lambda f: ("" if f == "C.A (TTC)" else f))

        # Build output rows (chronological order)
        rows = []
        for dt in dates_reversed:
            row = {"date": dt, "familles": {}}
            for fam in familles:
                row["familles"][fam] = date_data.get(dt, {}).get(fam, {
                    "real": 0, "obj": 0, "obj_mois": 0,
                    "percent": 0, "h_pct": 0, "raf": 0, "encours": 0
                })
            rows.append(row)

        return jsonify({
            "status": "success",
            "vendeur": vendeur_name,
            "dates": dates_reversed,
            "familles": familles,
            "rows": rows
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
@app.route('/api/login', methods=['POST'])
def login_api():
    try:
        data = request.get_json() or {}
        email = (data.get('email') or '').strip().lower()
        password = data.get('password') or ''
        
        valid_email = "guostonline@gmail.com"
        valid_password = "@Iamsorry123#"
        
        if email == valid_email.lower() and password == valid_password:
            return jsonify({
                "status": "success",
                "message": "Connexion réussie!",
                "user": {
                    "email": valid_email,
                    "name": "Guostonline"
                }
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Email ou mot de passe incorrect. Veuillez réessayer."
            }), 401
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == "__main__":
    db_manager.init_db()
    
    # Startup seeding for Focus objectives only
    try:
        conn = db_manager.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM focus_objectives")
        objs_count = cursor.fetchone()[0]
        conn.close()
        
        if objs_count == 0:
            if os.path.exists("Focus.xlsx"):
                print("[STARTUP] Seeding focus objectives from Focus.xlsx...")
                objs = import_focus_objectives_file("Focus.xlsx")
                if objs:
                    db_manager.save_focus_objectives(objs)
                    print(f"[STARTUP] Seeded {len(objs)} focus objectives.")
    except Exception as e:
        print(f"[STARTUP ERROR] Seeding focus objectives failed: {e}")

    # Perform initial run to check if data is populated
    try:
        p = get_processor()
        p.get_day_work()
        p.fix_sheet()
    except Exception as e:
        print(f"Error during initial processing: {e}")
        
    app.run(host="127.0.0.1", port=5000, debug=True)
