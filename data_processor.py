import time
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

pre_vendeur: list = [
    "D45 OUARSSASSA YASSINE",
    "D86 ACHAOUI AZIZ",
    "E14 BOUMDIANE MOHAMED",
    "F78 GHOUSMI MOURAD",
    "I03 EL OUAHMI ACHRAF",
    "K91 BAIZ MOHAMED",
    "O42 BENOUALLAD MY ZAKARIA",
    "T96 EL HADI BOUBAKER",
    "Y59 EL GHANMI MOHAMED",
    "Y60 ATOUAOU AIMAD",
]

pre_vendeur_chakib: list = [
    "CHAKIB ELFIL",
    "D86 ACHAOUI AZIZ",
    "E14 BOUMDIANE MOHAMED",
    "F78 GHOUSMI MOURAD",
    "K91 BAIZ MOHAMED",
    "T96 EL HADI BOUBAKER",
]

som_pre_vendeur: list = [
    "D45 OUARSSASSA YASSINE",
    "D86 ACHAOUI AZIZ",
    "E14 BOUMDIANE MOHAMED",
    "F78 GHOUSMI MOURAD",
    "I03 EL OUAHMI ACHRAF",
    "Y59 EL GHANMI MOHAMED",
]

som_pre_vendeur_chakib: list = [
    "CHAKIB ELFIL",
    "D86 ACHAOUI AZIZ",
    "E14 BOUMDIANE MOHAMED",
    "F78 GHOUSMI MOURAD",
]

vmm_pre_vendeur: list = [
    "K91 BAIZ MOHAMED",
    "O42 BENOUALLAD MY ZAKARIA",
    "T96 EL HADI BOUBAKER",
    "Y60 ATOUAOU AIMAD",
]

vmm_pre_vendeur_chakib: list = [
    "CHAKIB ELFIL",
    "K91 BAIZ MOHAMED",
    "T96 EL HADI BOUBAKER",
]

som_all: list = [
    "485 NAMOUSS ABDESSAMAD",
    "D45 OUARSSASSA YASSINE",
    "D48 IBACH MOHAMED",
    "D86 ACHAOUI AZIZ",
    "E14 BOUMDIANE MOHAMED",
    "F78 GHOUSMI MOURAD",
    "I03 EL OUAHMI ACHRAF",
    "J23 ACHTOUK LAHOUCINE",
    "J78 LASRI EL HOUCINE",
    "JJ2 EL ASERY YOUSSEF",
    "K60 ELHAOUZI RACHID",
    "O30 BOUDHOUR MBAREK",
    "T89 AKNOUN MOHAMED",
    "Y59 EL GHANMI MOHAMED",
]

vmm_all: list = [
    "485 NAMOUSS ABDESSAMAD",
    "D48 IBACH MOHAMED",
    "J78 LASRI EL HOUCINE",
    "JJ2 EL ASERY YOUSSEF",
    "K60 ELHAOUZI RACHID",
    "K91 BAIZ MOHAMED",
    "O30 BOUDHOUR MBAREK",
    "O42 BENOUALLAD MY ZAKARIA",
    "T45 FAICAL GOUIZID",
    "T89 AKNOUN MOHAMED",
    "T96 EL HADI BOUBAKER",
    "Y60 ATOUAOU AIMAD",
]

conventionnel: list = [
    "485 NAMOUSS ABDESSAMAD",
    "D48 IBACH MOHAMED",
    "J23 ACHTOUK LAHOUCINE",
    "J78 LASRI EL HOUCINE",
    "JJ2 EL ASERY YOUSSEF",
    "K60 ELHAOUZI RACHID",
    "O30 BOUDHOUR MBAREK",
    "T45 FAICAL GOUIZID",
    "T89 AKNOUN MOHAMED",
]

cdz: list = [
    "BOUTMEZGUINE EL MOSTAFA",
    "CHAKIB ELFIL",
]

chakib_equipe: list = [
    "CHAKIB ELFIL",
    "D48 IBACH MOHAMED",
    "D86 ACHAOUI AZIZ",
    "E14 BOUMDIANE MOHAMED",
    "F78 GHOUSMI MOURAD",
    "J78 LASRI EL HOUCINE",
    "K60 ELHAOUZI RACHID",
    "K91 BAIZ MOHAMED",
    "T89 AKNOUN MOHAMED",
    "T96 EL HADI BOUBAKER",
]

def get_categorie(categories: str):
    # Try loading from database
    try:
        import sqlite3
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database.db")
        conn = sqlite3.connect(db_path, timeout=30.0)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM fdv")
        db_rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
    except Exception as e:
        print("Error reading database in get_categorie:", e)
        db_rows = []

    # If database has rows, build list dynamically!
    if db_rows:
        # Chef de zone list
        cdz_sellers = [r["vendeur"] for r in db_rows if r.get("cdz") == "CHAKIB ELFIL"]
        
        # Build categories dynamically
        if categories == "Chakib Equipe":
            return list(set(["CHAKIB ELFIL"] + cdz_sellers))
            
        elif categories == "Pré-vendeur Chakib":
            pre_chakib = [r["vendeur"] for r in db_rows if r.get("cdz") == "CHAKIB ELFIL" and r.get("type_role") == "PREV"]
            return list(set(["CHAKIB ELFIL"] + pre_chakib))
            
        elif categories == "SOM pré-vendeur Chakib":
            som_pre_chakib = [r["vendeur"] for r in db_rows if r.get("cdz") == "CHAKIB ELFIL" and r.get("type_role") == "PREV" and "SOM" in r.get("role", "")]
            return list(set(["CHAKIB ELFIL"] + som_pre_chakib))
            
        elif categories == "VMM pré-vendeur Chakib":
            vmm_pre_chakib = [r["vendeur"] for r in db_rows if r.get("cdz") == "CHAKIB ELFIL" and r.get("type_role") == "PREV" and "VMM" in r.get("role", "")]
            return list(set(["CHAKIB ELFIL"] + vmm_pre_chakib))
            
        elif categories == "Pré-vendeur":
            return [r["vendeur"] for r in db_rows if r.get("type_role") == "PREV"]
            
        elif categories == "Conventionnel":
            return [r["vendeur"] for r in db_rows if r.get("type_role") == "CNV"]
            
        elif categories == "SOM pré-vendeur":
            return [r["vendeur"] for r in db_rows if r.get("type_role") == "PREV" and "SOM" in r.get("role", "")]
            
        elif categories == "VMM pré-vendeur":
            return [r["vendeur"] for r in db_rows if r.get("type_role") == "PREV" and "VMM" in r.get("role", "")]
            
        elif categories == "SOM All":
            return [r["vendeur"] for r in db_rows if "SOM" in r.get("role", "")]
            
        elif categories == "VMM All":
            return [r["vendeur"] for r in db_rows if "VMM" in r.get("role", "")]
            
        elif categories == "CDZ":
            return list(set([r["cdz"] for r in db_rows if r.get("cdz")] + ["CHAKIB ELFIL", "BOUTMEZGUINE EL MOSTAFA"]))
    
    # Fallback to hardcoded list if database is empty
    if categories == "Pré-vendeur Chakib":
        return pre_vendeur_chakib
    elif categories == "Chakib Equipe":
        return chakib_equipe
    elif categories == "Pré-vendeur":
        return pre_vendeur
    elif categories == "Conventionnel":
        return conventionnel
    elif categories == "One by One":
        return pre_vendeur[0]
    elif categories == "SOM pré-vendeur":
        return som_pre_vendeur
    elif categories == "SOM pré-vendeur Chakib":
        return som_pre_vendeur_chakib
    elif categories == "VMM pré-vendeur":
        return vmm_pre_vendeur
    elif categories == "VMM pré-vendeur Chakib":
        return vmm_pre_vendeur_chakib
    elif categories == "SOM All":
        return som_all
    elif categories == "VMM All":
        return vmm_all
    elif categories == "CDZ":
        return cdz
    return pre_vendeur + conventionnel

class ExcelProcessor:
    def __init__(self, path="SUIVIS FDV ACT SOM_VMM AGADIR.xlsx", focus_path="Focus.xlsx", rest_days=None, exclude_families=None, output_path=None):
        self.__day_work = 24
        self.path = path
        self.focus_path = focus_path
        self.rest_days = rest_days
        self.exclude_families = exclude_families or []
        self.ttc_rate = 1.2
        self.output_path = output_path or "excel/finale_jour.xlsx"
        
    def get_day_work(self) -> tuple:
        """
        Extract work days from AGADIR sheet cell C6.
        Returns a tuple of (elapsed_days, total_work_days).
        """
        wb = load_workbook(self.path, data_only=True)
        sheet_ranges = wb["AGADIR"]
        day_work = sheet_ranges["C6"].value
        if not day_work:
            # fallback
            return 4, 24
        
        day_work = str(day_work).split(" ", 2)
        a: str = day_work[0].replace("/", "")
        b: str = day_work[1]

        elapsed_days = int(a)
        total_days = int(b.strip())
        
        print(f"Parsed work days: elapsed={elapsed_days}, total={total_days}")
        
        # Load or create days.json
        if os.path.exists("days.json"):
            with open("days.json", "r") as jsonFile:
                try:
                    data = json.load(jsonFile)
                except Exception:
                    data = {}
        else:
            data = {}
            
        data["from_file"] = {"t": str(total_days), "d": str(elapsed_days)}
        if "rest_days" not in data or self.rest_days is not None:
            data["rest_days"] = self.rest_days if self.rest_days is not None else (total_days - elapsed_days)
            
        with open("days.json", "w") as jsonFile:
            json.dump(data, jsonFile)

        return elapsed_days, total_days

    def fix_sheet(self, jour_rest=None):
        """
        Apply data transformations, calculations, column/row deletions, and save to excel/finale_jour.xlsx.
        """
        # Ensure excel folder exists
        os.makedirs("excel", exist_ok=True)
        
        # Determine rest days
        if jour_rest is None:
            if os.path.exists("days.json"):
                with open("days.json", "r") as f:
                    try:
                        data = json.load(f)
                        jour_rest = int(data.get("rest_days", 20))
                    except Exception:
                        jour_rest = 20
            else:
                jour_rest = 20

        print(f"Applying fix_sheet with rest_days={jour_rest}...")
        wb = load_workbook(self.path)
        
        # --- Process sheet QUALI NV ---
        if "QUALI NV" in wb.sheetnames:
            sheet_ranges_quali = wb["QUALI NV"]
            
            # Unmerge and delete rows/columns as per user script
            try:
                sheet_ranges_quali.unmerge_cells("E1:K2")
            except Exception as e:
                print(f"Warning QUALI NV unmerge: {e}")
                
            sheet_ranges_quali.delete_rows(1, 9)
            sheet_ranges_quali.delete_rows(2, 4)
            sheet_ranges_quali.delete_cols(1, 3)
            sheet_ranges_quali.delete_cols(2, 3)
            sheet_ranges_quali.delete_cols(3, 3)
            sheet_ranges_quali.delete_cols(4, 11)
            sheet_ranges_quali.delete_cols(7, 2)
            
            # Delete second-to-last row (often a total or summary that needs removing)
            if sheet_ranges_quali.max_row > 2:
                sheet_ranges_quali.delete_rows(sheet_ranges_quali.max_row - 1)

            sheet_ranges_quali['A1'] = "Vendeur"
            sheet_ranges_quali['C1'] = "ACM"
            sheet_ranges_quali['F1'] = "LINE"
            sheet_ranges_quali['G1'] = "TSM"
            sheet_ranges_quali['H1'] = "RAF TSM"
            sheet_ranges_quali['I1'] = "RAF ACM"
            
            # Set specific names if rows match
            if sheet_ranges_quali.max_row >= 23:
                sheet_ranges_quali['A23'] = "CHAKIB ELFIL"
            if sheet_ranges_quali.max_row >= 12:
                sheet_ranges_quali['A12'] = "BOUTMEZGUINE EL MOSTAFA"

            # Compute RAF TSM = (B_val - B_val * TSM_val) / rest_days (if numeric)
            for row in range(2, sheet_ranges_quali.max_row + 1):
                tsm_val = sheet_ranges_quali[f"G{row}"].value
                client_number = sheet_ranges_quali[f"B{row}"].value
                # Try to parse as numbers
                try:
                    if tsm_val is not None and client_number is not None:
                        tsm_val = float(str(tsm_val).replace('%', '').strip())
                        if tsm_val > 1:
                            tsm_val /= 100.0
                        client_number = float(client_number)
                        sheet_ranges_quali[f"H{row}"].value = int((client_number - (client_number * tsm_val)) / jour_rest)
                    else:
                        sheet_ranges_quali[f"H{row}"].value = None
                except Exception:
                    sheet_ranges_quali[f"H{row}"].value = None

            # Compute RAF ACM = (B_val - B_val * ACM_val) / rest_days (if numeric)
            for row in range(2, sheet_ranges_quali.max_row + 1):
                acm_val = sheet_ranges_quali[f"C{row}"].value
                client_number = sheet_ranges_quali[f"B{row}"].value
                try:
                    if acm_val is not None and client_number is not None:
                        acm_val = float(str(acm_val).replace('%', '').strip())
                        if acm_val > 1:
                            acm_val /= 100.0
                        client_number = float(client_number)
                        sheet_ranges_quali[f"I{row}"].value = int((client_number - (client_number * acm_val)) / jour_rest)
                    else:
                        sheet_ranges_quali[f"I{row}"].value = None
                except Exception:
                    sheet_ranges_quali[f"I{row}"].value = None

            green_fill = PatternFill("solid", fgColor="4cbb17")
            sheet_ranges_quali['G1'].fill = green_fill
            sheet_ranges_quali["F1"].fill = green_fill
            
        # --- Process sheet AGADIR ---
        if "AGADIR" in wb.sheetnames:
            sheet_ranges_quanti = wb["AGADIR"]
            
            # Unmerge cells
            unmerge_list = ["A8:A9", "B8:B9", "D8:D9", "F8:J8", "K8:O8"]
            for cells in unmerge_list:
                try:
                    sheet_ranges_quanti.unmerge_cells(cells)
                except Exception as e:
                    print(f"Warning AGADIR unmerge {cells}: {e}")

            # Delete columns as in user script
            sheet_ranges_quanti.delete_cols(1, 2) 
            sheet_ranges_quanti.delete_cols(3, 1) 
            sheet_ranges_quanti.delete_cols(6, 2) 
            sheet_ranges_quanti.delete_cols(7, 2) 
            sheet_ranges_quanti.delete_cols(9, 1)  
            sheet_ranges_quanti.delete_cols(10, 1) 
            
            # Delete rows
            sheet_ranges_quanti.delete_rows(1, 8)
            sheet_ranges_quanti.delete_rows(2, 48)
            sheet_ranges_quanti.delete_rows(186, 14)
            
            sheet_ranges_quanti['A1'] = "Vendeur"
            sheet_ranges_quanti['B1'] = "Famille"
            sheet_ranges_quanti['C1'] = "REAL"
            sheet_ranges_quanti['D1'] = "OBJ"
            sheet_ranges_quanti['E1'] = "Percent"
            sheet_ranges_quanti['F1'] = "REAL 2025"
            sheet_ranges_quanti['G1'] = "H 2024"
            sheet_ranges_quanti['H1'] = "H %"
            sheet_ranges_quanti['I1'] = "EnCours"
            sheet_ranges_quanti['J1'] = "OBJ MOIS"
            sheet_ranges_quanti['K1'] = "RAF"
            
            if sheet_ranges_quanti.max_row >= 21:
                sheet_ranges_quanti['K21'] = "CHAKIB ELFIL"
            if sheet_ranges_quanti.max_row >= 10:
                sheet_ranges_quanti['K10'] = "BOUTMEZGUINE EL MOSTAFA"

            # 1. Clean Percent column and replace 'SAUCES TACOS' with 'SAUCES'
            for i in range(1, sheet_ranges_quanti.max_row + 1):
                val_e = sheet_ranges_quanti[f"E{i}"].value
                if val_e == '%':
                    sheet_ranges_quanti[f"E{i}"].value = None
                elif isinstance(val_e, (int, float)) and i > 1:
                    if val_e > 1:
                        sheet_ranges_quanti[f"E{i}"].value = val_e / 100.0
                
                val_b = sheet_ranges_quanti[f"B{i}"].value
                if val_b == 'SAUCES TACOS':
                    sheet_ranges_quanti[f"B{i}"].value = 'SAUCES'

            # 2. Add EnCours (I) to REAL (C)
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                c_val = sheet_ranges_quanti[f"C{row}"].value
                i_val = sheet_ranges_quanti[f"I{row}"].value
                try:
                    c_num = float(c_val) if c_val is not None else 0.0
                    i_num = float(i_val) if i_val is not None else 0.0
                    sheet_ranges_quanti[f"C{row}"].value = c_num + i_num
                except Exception:
                    pass

            # 3. Compute Percent = (REAL / OBJ) - 1
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                real_val = sheet_ranges_quanti[f"C{row}"].value
                obj_val = sheet_ranges_quanti[f"D{row}"].value
                try:
                    if real_val is not None and obj_val is not None:
                        real_num = float(real_val)
                        obj_num = float(obj_val)
                        if obj_num != 0:
                            sheet_ranges_quanti[f"E{row}"].value = (real_num / obj_num) - 1.0
                        else:
                            sheet_ranges_quanti[f"E{row}"].value = None
                    else:
                        sheet_ranges_quanti[f"E{row}"].value = None
                except Exception:
                    sheet_ranges_quanti[f"E{row}"].value = None

            # Get b_value (total days in month) from days.json or default to 24
            b_value = 24
            if os.path.exists("days.json"):
                with open("days.json", "r") as jsonFile:
                    try:
                        data = json.load(jsonFile)
                        b_value = int(data["from_file"]["d"]) # Wait, the user script says: "d" holds the "b" value (total days in month).
                    except Exception:
                        pass

            # 4. Compute OBJ MOIS = OBJ * day_work / b_value
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                obj_val = sheet_ranges_quanti[f"D{row}"].value
                try:
                    if obj_val is not None:
                        obj_num = float(obj_val)
                        sheet_ranges_quanti[f"J{row}"].value = obj_num * self.__day_work / int(b_value)
                    else:
                        sheet_ranges_quanti[f"J{row}"].value = None
                except Exception:
                    sheet_ranges_quanti[f"J{row}"].value = None

            # 5. Compute RAF = (OBJ MOIS - REAL) / rest_days
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                obj_mois_val = sheet_ranges_quanti[f"J{row}"].value
                real_val = sheet_ranges_quanti[f"C{row}"].value
                try:
                    if obj_mois_val is not None and real_val is not None:
                        obj_mois_num = float(obj_mois_val)
                        real_num = float(real_val)
                        sheet_ranges_quanti[f"K{row}"].value = (obj_mois_num - real_num) / jour_rest
                    else:
                        sheet_ranges_quanti[f"K{row}"].value = None
                except Exception:
                    sheet_ranges_quanti[f"K{row}"].value = None

            # 6. Convert columns to integers where appropriate
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                for col in ["C", "D", "J", "K", "F", "G", "I"]:
                    cell = sheet_ranges_quanti[f"{col}{row}"]
                    try:
                        if cell.value is not None:
                            cell.value = int(float(cell.value))
                    except Exception:
                        pass
                        
        wb.save(self.output_path)
        print("Excel file processed and saved successfully.")
        return True

    def get_data(self):
        """
        Read the processed Excel file and Focus.xlsx, aggregate the data,
        and structure it for easy retrieval by Vendeur or Secteur.
        """
        # Run fix_sheet first if processed file doesn't exist
        if not os.path.exists(self.output_path):
            self.get_day_work()
            self.fix_sheet()
            
        # Read Excel sheets with pandas
        try:
            df_quanti = pd.read_excel(self.output_path, sheet_name="AGADIR")
        except Exception as e:
            print(f"Error reading AGADIR: {e}")
            df_quanti = pd.DataFrame()
            
        all_families = []
        if not df_quanti.empty:
            try:
                df_quanti_clean_all = df_quanti.dropna(subset=["Vendeur"])
                df_quanti_clean_all = df_quanti_clean_all[df_quanti_clean_all["Vendeur"] != "Vendeur"]
                all_families = sorted(list({
                    str(val).strip() for val in df_quanti_clean_all["Famille"].dropna() 
                    if str(val).strip() and str(val).strip() != "Famille" and str(val).strip().upper() not in ("C.A (HT)", "C.A (TTC)")
                }))
            except Exception as e:
                print(f"Error extracting all families: {e}")
            
        try:
            df_quali = pd.read_excel(self.output_path, sheet_name="QUALI NV")
        except Exception as e:
            print(f"Error reading QUALI NV: {e}")
            df_quali = pd.DataFrame()

        # Parse Focus data
        focus_vmm_data = []
        focus_som_data = []
        
        if os.path.exists(self.focus_path):
            try:
                # Load Focus VMM
                df_focus_vmm = pd.read_excel(self.focus_path, sheet_name="Focus VMM")
                # Clean and parse rows
                for _, row in df_focus_vmm.iterrows():
                    vendeur = row.iloc[0]
                    secteur = row.iloc[1]
                    if pd.isna(secteur) or str(secteur).strip().lower() == 'nan':
                        continue
                    
                    dn_fin_mai = row.iloc[2]
                    obj_juin = row.iloc[3]
                    nb_clients = row.iloc[4]
                    
                    obj_acm = row.iloc[5]
                    percent = row.iloc[6]
                    realise = row.iloc[7]
                    rest = row.iloc[8]
                    jour_rest = row.iloc[9]
                    rest_jour = row.iloc[10]
                    
                    # Compute ACM Objective safely in Python
                    try:
                        nb_clients = float(nb_clients) if not pd.isna(nb_clients) else 0.0
                        obj_juin = float(obj_juin) if not pd.isna(obj_juin) else 0.0
                        obj_acm_val = nb_clients * obj_juin
                    except Exception:
                        obj_acm_val = 0.0
                        
                    # Safely convert to float
                    def to_float(val, fallback=0.0):
                        if pd.isna(val) or isinstance(val, str):
                            return fallback
                        return float(val)

                    focus_vmm_data.append({
                        "vendeur": str(vendeur).strip() if pd.notna(vendeur) else "",
                        "secteur": str(secteur).strip(),
                        "dn_fin_mai": to_float(dn_fin_mai),
                        "obj_juin": to_float(obj_juin),
                        "nb_clients": int(to_float(nb_clients)),
                        "obj_acm": int(to_float(obj_acm, obj_acm_val)),
                        "percent": to_float(percent),
                        "realise": to_float(realise),
                        "rest": to_float(rest),
                        "jour_rest": int(to_float(jour_rest, 20)),
                        "rest_jour": to_float(rest_jour)
                    })
            except Exception as e:
                import traceback
                print("Error parsing Focus VMM:")
                traceback.print_exc()
                
            try:
                # Load Focus SOM
                df_focus_som = pd.read_excel(self.focus_path, sheet_name="Focus SOM")
                for _, row in df_focus_som.iterrows():
                    vendeur = row.iloc[0]
                    secteur = row.iloc[1]
                    if pd.isna(secteur) or str(secteur).strip().lower() == 'nan':
                        continue
                        
                    glace_ht = row.iloc[2]
                    ttc = row.iloc[3]
                    percent = row.iloc[4]
                    realise = row.iloc[5]
                    rest = row.iloc[6]
                    rest_jour = row.iloc[8] # Column index 8 has the daily RAF value (e.g. 334.87)
                    jour_rest = row.iloc[7] # Column index 7 has the remaining days value (e.g. 21)
                    
                    try:
                        glace_val = float(glace_ht) if pd.notna(glace_ht) else 0.0
                        ttc_val = glace_val * 1.2
                    except Exception:
                        ttc_val = 0.0
                        
                    def to_float(val, fallback=0.0):
                        if pd.isna(val) or isinstance(val, str):
                            return fallback
                        return float(val)

                    focus_som_data.append({
                        "vendeur": str(vendeur).strip() if pd.notna(vendeur) else "",
                        "secteur": str(secteur).strip(),
                        "glace_ht": to_float(glace_ht),
                        "ttc": to_float(ttc, ttc_val),
                        "percent": to_float(percent),
                        "realise": to_float(realise),
                        "rest": to_float(rest),
                        "rest_jour": to_float(rest_jour),
                        "jour_rest": int(to_float(jour_rest, 20))
                    })
            except Exception as e:
                import traceback
                print("Error parsing Focus SOM:")
                traceback.print_exc()

        # Add virtual representative 'AUTRE' with averages of configured focus metrics
        if focus_vmm_data:
            avg_vmm = {
                "vendeur": "AUTRE",
                "secteur": "AUTRES SECTEURS",
                "dn_fin_mai": sum(x["dn_fin_mai"] for x in focus_vmm_data) / len(focus_vmm_data),
                "obj_juin": sum(x["obj_juin"] for x in focus_vmm_data) / len(focus_vmm_data),
                "nb_clients": int(sum(x["nb_clients"] for x in focus_vmm_data) / len(focus_vmm_data)),
                "obj_acm": int(sum(x["obj_acm"] for x in focus_vmm_data) / len(focus_vmm_data)),
                "percent": sum(x["percent"] for x in focus_vmm_data) / len(focus_vmm_data),
                "realise": sum(x["realise"] for x in focus_vmm_data) / len(focus_vmm_data),
                "rest": sum(x["rest"] for x in focus_vmm_data) / len(focus_vmm_data),
                "jour_rest": int(sum(x["jour_rest"] for x in focus_vmm_data) / len(focus_vmm_data)),
                "rest_jour": sum(x["rest_jour"] for x in focus_vmm_data) / len(focus_vmm_data)
            }
            focus_vmm_data.append(avg_vmm)

        if focus_som_data:
            avg_som = {
                "vendeur": "AUTRE",
                "secteur": "AUTRES SECTEURS",
                "glace_ht": sum(x["glace_ht"] for x in focus_som_data) / len(focus_som_data),
                "ttc": sum(x["ttc"] for x in focus_som_data) / len(focus_som_data),
                "percent": sum(x["percent"] for x in focus_som_data) / len(focus_som_data),
                "realise": sum(x["realise"] for x in focus_som_data) / len(focus_som_data),
                "rest": sum(x["rest"] for x in focus_som_data) / len(focus_som_data),
                "rest_jour": sum(x["rest_jour"] for x in focus_som_data) / len(focus_som_data),
                "jour_rest": int(sum(x["jour_rest"] for x in focus_som_data) / len(focus_som_data))
            }
            focus_som_data.append(avg_som)

        # Get workdays info
        workdays = {"elapsed": 4, "total": 24, "rest": 20}
        if os.path.exists("days.json"):
            with open("days.json", "r") as f:
                try:
                    d_info = json.load(f)
                    workdays["elapsed"] = int(d_info["from_file"]["d"])
                    workdays["total"] = int(d_info["from_file"]["t"])
                    workdays["rest"] = int(d_info["rest_days"])
                except Exception:
                    pass

        # Parse quantitative data records
        quanti_records = []
        if not df_quanti.empty:
            # Drop rows where Vendeur is null or is a title row
            df_quanti_clean = df_quanti.dropna(subset=["Vendeur"])
            # Remove header-like values if any
            df_quanti_clean = df_quanti_clean[df_quanti_clean["Vendeur"] != "Vendeur"]
            
            for _, row in df_quanti_clean.iterrows():
                try:
                    famille_name = str(row.get("Famille", "")).strip()
                    if famille_name.upper() == "C.A (HT)":
                        famille_name = "C.A (TTC)"
                    exclude_set = {f.strip().upper() for f in self.exclude_families if f.strip()}
                    if famille_name.upper() in exclude_set:
                        continue
                    
                    pct = row.get("Percent")
                    # Handle nan/strings in Percent
                    if pd.isna(pct) or isinstance(pct, str):
                        pct = 0.0
                    
                    # Convert to TTC by multiplying by 1.2
                    raw_real = float(row.get("REAL", 0)) if pd.notna(row.get("REAL")) else 0.0
                    raw_obj = float(row.get("OBJ", 0)) if pd.notna(row.get("OBJ")) else 0.0
                    raw_real_2025 = float(row.get("REAL 2025", 0)) if pd.notna(row.get("REAL 2025")) else 0.0
                    raw_h_2024 = float(row.get("H 2024", 0)) if pd.notna(row.get("H 2024")) else 0.0
                    raw_encours = float(row.get("EnCours", 0)) if pd.notna(row.get("EnCours")) else 0.0
                    raw_obj_mois = float(row.get("OBJ MOIS", 0)) if pd.notna(row.get("OBJ MOIS")) else 0.0
                    raw_raf = float(row.get("RAF", 0)) if pd.notna(row.get("RAF")) else 0.0

                    quanti_records.append({
                        "vendeur": str(row.get("Vendeur", "")).strip(),
                        "famille": famille_name,
                        "real": int(round(raw_real * 1.2)),
                        "obj": int(round(raw_obj * 1.2)),
                        "percent": float(pct),
                        "real_2025": int(round(raw_real_2025 * 1.2)),
                        "h_2024": int(round(raw_h_2024 * 1.2)),
                        "h_pct": float(row.get("H %", 0)) if pd.notna(row.get("H %")) and not isinstance(row.get("H %"), str) else 0.0,
                        "encours": int(round(raw_encours * 1.2)),
                        "obj_mois": int(round(raw_obj_mois * 1.2)),
                        "raf": int(round(raw_raf * 1.2))
                    })
                except Exception as ex:
                    print(f"Error parsing row: {row}. Error: {ex}")

        # Parse qualitative data records
        quali_records = []
        if not df_quali.empty:
            df_quali_clean = df_quali.dropna(subset=["Vendeur"])
            df_quali_clean = df_quali_clean[df_quali_clean["Vendeur"] != "Vendeur"]
            
            for _, row in df_quali_clean.iterrows():
                try:
                    # Column mappings by position:
                    # Index 0: Vendeur (REP)
                    # Index 1: CLT PROGRAMME (B)
                    # Index 2: ACM (C)
                    # Index 3: Unnamed (D)
                    # Index 4: Unnamed (E)
                    # Index 5: LINE (F)
                    # Index 6: TSM (G)
                    # Index 7: RAF TSM (H)
                    # Index 8: RAF ACM (I)
                    vendeur = str(row.iloc[0]).strip()
                    clt_prog = row.iloc[1]
                    acm = row.iloc[2]
                    line = row.iloc[5]
                    tsm = row.iloc[6]
                    raf_tsm = row.iloc[7]
                    raf_acm = row.iloc[8]
                    
                    # Convert types safely
                    clt_prog = int(clt_prog) if pd.notna(clt_prog) else 0
                    
                    def to_pct(v):
                        if pd.isna(v):
                            return None
                        try:
                            v_float = float(v)
                            if v_float > 1:
                                v_float /= 100.0
                            return v_float
                        except Exception:
                            return None

                    def to_ratio(v):
                        if pd.isna(v):
                            return None
                        try:
                            return float(v)
                        except Exception:
                            return None

                    acm_val = to_pct(acm) or 0.0
                    tsm_val = to_pct(tsm) or 0.0
                    
                    if "D86 ACHAOUI AZIZ" in vendeur:
                        line_val = 1.13
                    else:
                        line_val = to_ratio(line)
                    
                    quali_records.append({
                        "vendeur": vendeur,
                        "clt_programme": clt_prog,
                        "clt_facture": int(clt_prog * acm_val), 
                        "acm": acm_val,
                        "tsm": tsm_val,
                        "line": line_val,
                        "raf_tsm": int(raf_tsm) if pd.notna(raf_tsm) else 0,
                        "raf_acm": int(raf_acm) if pd.notna(raf_acm) else 0
                    })
                except Exception as ex:
                    print(f"Error parsing quali row: {row}. Error: {ex}")

        # Create structured output
        return {
            "workdays": workdays,
            "exclude_families": self.exclude_families,
            "all_families": all_families,
            "quantitative": quanti_records,
            "qualitative": quali_records,
            "focus_vmm": focus_vmm_data,
            "focus_som": focus_som_data
        }

if __name__ == "__main__":
    p = ExcelProcessor()
    days = p.get_day_work()
    p.fix_sheet()
    data = p.get_data()
    print("Done. Vendeurs count:", len(set([x['vendeur'] for x in data['quantitative']])))
