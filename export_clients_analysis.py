import os
import sqlite3
import openpyxl
import pandas as pd

def generate_export():
    print("=== STARTING FULL EXPORT ANALYSIS FROM UPDATED ACM.XLSX & ALL SECTEURS ===")
    
    # 1. Load mapping from clients.xlsx (Secteurs & localités sheet)
    sec_loc_map = {}
    clients_xlsx = r"c:\Users\DELL\Dev\dashboard2027\clients.xlsx"
    
    if os.path.exists(clients_xlsx):
        wb_c = openpyxl.load_workbook(clients_xlsx, data_only=True, read_only=True)
        sheet_n = None
        for name in wb_c.sheetnames:
            if "Secteur" in name:
                sheet_n = name
                break
        if sheet_n:
            ws_sl = wb_c[sheet_n]
            for row in ws_sl.iter_rows(values_only=True):
                if row and len(row) >= 5 and row[0] != "Secteur":
                    sec, loc, _cnt, som, vmm = row[0], row[1], row[2], row[3], row[4]
                    if sec and loc:
                        sec_str = str(sec).strip().upper()
                        loc_str = str(loc).strip().upper()
                        sec_loc_map[(sec_str, loc_str)] = (
                            str(som).strip() if som else "",
                            str(vmm).strip() if vmm else ""
                        )
        wb_c.close()

    # 2. Load Vendeur mapping from database (vendeur_tournees_visits)
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    visit_client_map = {}
    vt_rows = cursor.execute("""
        SELECT client_code, client_name, tournee, vendeur_code, vendeur_name 
        FROM vendeur_tournees_visits
    """).fetchall()

    if not vt_rows:
        vr_rows = cursor.execute("""
            SELECT client_code, client_nom, tournee, vendeur
            FROM visites_rapports
        """).fetchall()
        vt_rows = []
        for r in vr_rows:
            v_raw = str(r[3] or '').strip()
            v_code = v_raw.split()[0] if v_raw else ''
            vt_rows.append((r[0], r[1], r[2], v_code, v_raw))

    for r in vt_rows:
        code = str(r[0]).strip().upper()
        c_name = str(r[1]).strip()
        tournee = str(r[2]).strip()
        v_code = str(r[3]).strip().upper()
        v_name = str(r[4]).strip()
        full_v = f"{v_code} {v_name}".strip() if v_code else v_name
        
        if code not in visit_client_map:
            visit_client_map[code] = {
                "name": c_name,
                "tournee": tournee,
                "vendeur_code": v_code,
                "vendeur_name": v_name,
                "full_vendeur": full_v
            }

    # 3. Load acm.xlsx clients directly with updated "Secteur" column
    acm_xlsx = r"c:\Users\DELL\Dev\dashboard2027\acm.xlsx"
    acm_clients = {}
    
    if os.path.exists(acm_xlsx):
        wb_acm = openpyxl.load_workbook(acm_xlsx, data_only=True, read_only=True)
        ws_acm = wb_acm.active
        
        header_found = False
        for row in ws_acm.iter_rows(values_only=True):
            if not any(row):
                continue
            row_str = [str(v).strip() for v in row if v is not None]
            if "Code Client" in row_str or "Tournee" in row_str or "Secteur" in row_str:
                header_found = True
                continue
            if not header_found:
                continue
            
            if len(row) >= 5:
                agence = str(row[0]).strip() if row[0] else ""
                secteur = str(row[1]).strip() if row[1] else ""
                tournee = str(row[2]).strip() if row[2] else ""
                c_code = str(row[3]).strip() if row[3] else ""
                c_name = str(row[4]).strip() if row[4] else ""
                
                if c_code and c_code not in ["Code Client", "Code"]:
                    c_code_u = c_code.upper()
                    if c_code_u not in acm_clients:
                        acm_clients[c_code_u] = {
                            "code": c_code,
                            "client": c_name,
                            "secteur": secteur,
                            "tournee": tournee
                        }
        wb_acm.close()

    # Combine all unique client codes
    all_client_codes = set(list(acm_clients.keys()) + list(visit_client_map.keys()))

    export_rows = []

    for code in sorted(list(all_client_codes)):
        acm_info = acm_clients.get(code, {})
        visit_info = visit_client_map.get(code, {})
        
        c_code = code
        c_name = acm_info.get("client") or visit_info.get("name") or ""
        tournee = acm_info.get("tournee") or visit_info.get("tournee") or ""
        secteur = acm_info.get("secteur") or ""
        
        if not secteur and tournee:
            if "AIT MELLOUL" in tournee.upper():
                secteur = "AIT MELLOUL SOM"
            elif "INZEGAN" in tournee.upper():
                secteur = "INZEGANE SOM"
            elif "TIKIOUINE" in tournee.upper():
                secteur = "AGADIR TIKIOUINE SOM"
            elif "TAROUDANT" in tournee.upper():
                secteur = "TAROUDANT SV"
            else:
                secteur = "AGADIR CENTRE VILLE SOM"

        # Resolve Vendeur SOM & Vendeur VMM
        vendeur_som = ""
        vendeur_vmm = ""
        
        # Priority 1: Check clients.xlsx (secteur, localite) mapping
        sec_key = secteur.upper()
        loc_key = tournee.upper()
        som_vmm = sec_loc_map.get((sec_key, loc_key))
        
        if not som_vmm:
            for (sk, lk), (s_val, v_val) in sec_loc_map.items():
                if lk in loc_key or loc_key in lk:
                    som_vmm = (s_val, v_val)
                    break
        
        if som_vmm:
            vendeur_som = som_vmm[0] or ""
            vendeur_vmm = som_vmm[1] or ""

        # Priority 2: Use All Secteurs visit log for this client
        if visit_info:
            v_full = visit_info.get("full_vendeur", "")
            if not vendeur_som:
                vendeur_som = v_full
            if not vendeur_vmm:
                vendeur_vmm = v_full

        # Priority 3: Derive from Secteur name string
        if not vendeur_som and secteur:
            if "AIT MELLOUL" in secteur.upper():
                vendeur_som = "F78 GHOUSMI MOURAD"
            elif "INZEGANE" in secteur.upper():
                vendeur_som = "E14 BOUMDIANE MOHAMED"
            elif "TIKIOUINE" in secteur.upper():
                vendeur_som = "D86 ACHAOUI AZIZ"
            elif "TAROUDANT" in secteur.upper():
                vendeur_som = "D48 IBACH MOHAMED"
            else:
                vendeur_som = secteur

        if not vendeur_vmm and secteur:
            if "AIT MELLOUL" in secteur.upper():
                vendeur_vmm = "F78 GHOUSMI MOURAD"
            elif "INZEGANE" in secteur.upper():
                vendeur_vmm = "K91 BAIZ MOHAMED"
            elif "TIKIOUINE" in secteur.upper():
                vendeur_vmm = "T96 EL HADI BOUBAKER"
            elif "TAROUDANT" in secteur.upper():
                vendeur_vmm = "T89 AKNOUN MOHAMED"
            else:
                vendeur_vmm = secteur

        export_rows.append({
            "code": c_code,
            "client": c_name,
            "secteur": secteur,
            "tourné": tournee,
            "vendeur som": vendeur_som,
            "vendeur vmm": vendeur_vmm
        })

    out_file = r"c:\Users\DELL\Dev\dashboard2027\export_clients_all_secteurs_acm_v2.xlsx"
    df = pd.DataFrame(export_rows)
    df = df[["code", "client", "secteur", "tourné", "vendeur som", "vendeur vmm"]]
    
    df.to_excel(out_file, index=False, sheet_name="Clients")
    print(f"\nSUCCESS! Exported {len(df)} rows to {out_file}")
    conn.close()

if __name__ == "__main__":
    generate_export()
